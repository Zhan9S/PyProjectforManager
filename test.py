import logging
import json
import unicodedata
import tempfile
from datetime import datetime
import subprocess
import base64
import requests
import re
from urllib3 import disable_warnings
from urllib3.exceptions import InsecureRequestWarning
from flask import current_app

disable_warnings(InsecureRequestWarning)


# 初始化FFmpeg环境（启动时执行）
def init_ffmpeg_env():
    """初始化FFmpeg环境变量"""
    if not current_app:
        with app.app_context():
            return _init_ffmpeg_env_core()
    else:
        return _init_ffmpeg_env_core()

# 核心逻辑抽离
def _init_ffmpeg_env_core():
    """FFmpeg初始化核心逻辑（需在应用上下文内执行）"""
    # 从系统配置获取FFmpeg路径（优先），无则使用默认路径
    ffmpeg_bin = SystemConfig.get_config('ffmpeg_bin', "D:\\ffmpeg\\bin")
    ffmpeg_exe = SystemConfig.get_config('ffmpeg_exe', "D:\\ffmpeg\\bin\\ffmpeg.exe")
    ffprobe_exe = SystemConfig.get_config('ffprobe_exe', "D:\\ffmpeg\\bin\\ffprobe.exe")

    # 配置环境变量
    os.environ["PATH"] += os.pathsep + ffmpeg_bin
    os.environ["FFMPEG_BINARY"] = ffmpeg_exe
    os.environ["FFPROBE_BINARY"] = ffprobe_exe

    # 验证FFmpeg
    if not os.path.exists(ffmpeg_exe):
        logger.warning(f"FFmpeg文件不存在：{ffmpeg_exe}，语音功能将不可用")
        return False

    try:
        result = subprocess.run(
            [ffmpeg_exe, "-version"],
            capture_output=True,
            text=True,
            timeout=5
        )
        logger.info("✅ FFmpeg 初始化成功")
        return True
    except Exception as e:
        logger.error(f"FFmpeg初始化失败：{str(e)}")
        return False


# 成果类型匹配规则（OCR识别用）
achievement_rules = {
    '期刊论文': {'pattern': r'[学报 | 期刊 | 杂志].*卷.*期 | ISSN:\d+',
                 'keywords': ['期刊', '学报', '论文', '发表', '卷', '期'], 'priority': 1},
    '发明专利': {'pattern': r'ZL\d{4}\d{8}(\.\d+)?|发明专利申请号 | 公开号',
                 'keywords': ['发明', '专利', 'ZL', '申请号', '公开号'], 'priority': 1},
    '实用新型专利': {'pattern': r'实用新型专利|ZL\d{4}2\d{7}', 'keywords': ['实用新型', '专利', 'ZL'], 'priority': 1},
    '会议论文': {'pattern': r'会议论文 | 会议集 |Proceedings', 'keywords': ['会议', '研讨会', '论坛'],
                 'priority': 2},
    '教材': {'pattern': r'教材|ISBN[:：]?\s*\d+|主编 [:：]?|副主编 [:：]?|出版社 [:：]?',
             'keywords': ['教材', '主编', '副主编', 'ISBN', '出版社', '规划教材', '行业教材'],
             'priority': 2},
    '专著': {'pattern': r'专著|ISBN:\d+|独著 | 合著', 'keywords': ['专著', '独著', '合著', 'ISBN'], 'priority': 2},
    '软著': {'pattern': r'计算机软件著作权 | 软著登字第\d+ 号', 'keywords': ['软著', '软件著作权', '著作权'],
             'priority': 3},
    '教学成果获奖': {'pattern': r'成果奖 | 科技奖 | 一等奖 | 二等奖', 'keywords': ['成果', '获奖', '科技奖', '一等奖', '二等奖'],
                 'priority': 3},
    '教学竞赛获奖': {'pattern': r'教学竞赛 | 教学奖 | 课堂教学', 'keywords': ['教学竞赛', '教学奖', '课堂教学'],
                     'priority': 3},
    '指导学生获奖': {'pattern': r'指导老师 | 学生竞赛 | 大学生.*竞赛 | 参赛同学', 'keywords': ['指导老师', '学生', '竞赛', '获奖'],
                     'priority': 2},
    '教研教改和课程建设项目': {'pattern': r'教学改革研究 | 教改 | 课程建设 | 一流本科课程 | 课程思政 | 高等学校教育教学改革',
                                'keywords': ['教学改革', '教改', '课程建设', '一流课程', '课程思政', '立项', '湖南省普通高等学校'],
                                'priority': 1}
}

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ocr_voice.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

from pdf2image import convert_from_path
import tempfile
import os
import json
import csv
from datetime import datetime, date
from io import BytesIO
from flask import Flask, request, redirect, url_for, flash, session, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
import uuid
import pandas as pd
from sqlalchemy import or_, func
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import re
import time
import random
import requests
import json
from datetime import datetime


# 新增：定义允许上传的文件扩展名
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'docx', 'doc', 'xlsx', 'xls'}

# ---------------------- 1. 应用初始化配置 ----------------------
app = Flask(__name__)
DB_FILE = 'teaching_achievement.db'  # 数据库文件路径
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_FILE}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'teaching-achievement-2026-key'  # 生产环境需替换为随机密钥
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB上传限制

# 创建上传目录
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# 数据库初始化
db = SQLAlchemy(app)
migrate = Migrate(app, db)  # 可选，用于生产环境迁移


# ---------------------- 2. 数据库模型设计（修正团队负责人关联） ----------------------
class User(db.Model):
    """用户表（包含所有用户信息字段）"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    gender = db.Column(db.String(10))
    birth_date = db.Column(db.Date)
    # 关键修复：将 id_card 的默认值设为 NULL，且空值不触发唯一约束
    id_card = db.Column(db.String(18), unique=True, nullable=True, default=None)
    email = db.Column(db.String(100), unique=True, nullable=False)
    phone = db.Column(db.String(20))
    office_phone = db.Column(db.String(20))
    school = db.Column(db.String(100))
    college = db.Column(db.String(100))
    department = db.Column(db.String(100))
    research_room = db.Column(db.String(100))
    role = db.Column(db.String(20), nullable=False, default='teacher')
    api_config = db.Column(db.Text, default='{}')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 密码加密/验证
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    # 获取大模型API配置
    def get_api_config(self):
        try:
            return json.loads(self.api_config)
        except:
            return {}

    # 设置大模型API配置
    def set_api_config(self, config):
        self.api_config = json.dumps(config, ensure_ascii=False)

    # 关联关系修正：
    # 1. 反向引用：当前用户管理的所有团队（Team表的leader_id关联）
    managed_teams = db.relationship('Team', backref='leader', foreign_keys='Team.leader_id', lazy='dynamic')
    # 2. 反向引用：当前用户加入的所有团队（UserTeam表关联）
    joined_teams = db.relationship('UserTeam', backref='user', foreign_keys='UserTeam.user_id', lazy='dynamic')


class Team(db.Model):
    """团队表（核心关联团队负责人）"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)  # 团队名称
    leader_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 团队负责人ID（核心外键）
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 关联关系修正：
    # 1. 正向引用：团队负责人（关联User表）
    # leader = db.relationship('User', backref='managed_teams', foreign_keys=[leader_id])  # 原写法保留也可，二选一
    # 2. 反向引用：团队下的所有成员（通过UserTeam关联）
    members = db.relationship('UserTeam', backref='team', foreign_keys='UserTeam.team_id', lazy='dynamic')


class UserTeam(db.Model):
    """用户-团队关联表（多对多，区分负责人和普通成员）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 关联用户
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)  # 关联团队
    join_time = db.Column(db.DateTime, default=datetime.utcnow)
    is_admin = db.Column(db.Boolean, default=False)  # 扩展：是否为团队管理员（非负责人）

    # 联合唯一索引（一个用户只能加入一个团队一次）
    __table_args__ = (db.UniqueConstraint('user_id', 'team_id', name='_user_team_uc'),)


class AchievementContributor(db.Model):
    """成果关联人表（多对多，支持一个成果有多个关联用户）"""
    id = db.Column(db.Integer, primary_key=True)
    achievement_type = db.Column(db.String(50), nullable=False,
                                 comment='成果类型：journal_paper/conference_paper/textbook/monograph/etc.')
    achievement_id = db.Column(db.Integer, nullable=False, comment='成果 ID')
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='关联用户 ID')
    contributor_role = db.Column(db.String(50), default='author',
                                 comment='贡献角色：author/corresponding_author/editor/compile/etc.')
    is_creator = db.Column(db.Boolean, default=False, comment='是否为录入者')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('achievement_type', 'achievement_id', 'user_id', name='_achievement_user_uc'),
    )

    user = db.relationship('User', backref='achievement_contributions')


class InclusionType(db.Model):
    """论文收录类型表（字典表）"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(100), unique=True, nullable=False, comment='收录类型名称')
    type_code = db.Column(db.String(50), unique=True, nullable=False, comment='收录类型代码')
    description = db.Column(db.Text, comment='描述说明')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 反向引用
    journal_papers = db.relationship('JournalPaper', secondary='journal_paper_inclusion_relation', back_populates='inclusion_types')


class JournalPaperInclusionRelation(db.Model):
    """期刊论文 - 收录类型关联表（多对多）"""
    id = db.Column(db.Integer, primary_key=True)
    paper_id = db.Column(db.Integer, db.ForeignKey('journal_paper.id'), nullable=False)
    inclusion_type_id = db.Column(db.Integer, db.ForeignKey('inclusion_type.id'), nullable=False)
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    # 联合唯一索引（同一篇论文的同一收录类型只能有一条记录）
    __table_args__ = (db.UniqueConstraint('paper_id', 'inclusion_type_id', name='_paper_inclusion_uc'),)

    # 关联关系
    paper = db.relationship('JournalPaper', back_populates='inclusion_relations', overlaps='journal_papers')
    inclusion_type = db.relationship('InclusionType', backref=db.backref('paper_relations', overlaps='journal_papers'), overlaps='journal_papers')


class JournalPaper(db.Model):
    """期刊论文表（完整字段）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 论文名称
    authors = db.Column(db.Text, nullable=False)  # 论文作者（逗号分隔）
    corresponding_authors = db.Column(db.Text)  # 通讯作者（逗号分隔）
    journal_name = db.Column(db.String(200), nullable=False)  # 期刊名称
    inclusion_status = db.Column(db.Text)  # 收录情况（逗号分隔，保留兼容旧数据）
    inclusion_type_ids = db.Column(db.Text, default='', comment='收录类型 ID（逗号分隔，关联 inclusion_type 表）')
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和共同作者）')
    year = db.Column(db.Integer)  # 年
    volume = db.Column(db.String(50))  # 卷
    issue = db.Column(db.String(50))  # 期
    page_range = db.Column(db.String(50))  # 起止页码
    doi = db.Column(db.String(200))  # DOI
    publish_year = db.Column(db.Integer)  # 发表年份
    publish_date = db.Column(db.Date)  # 发表日期
    attachment = db.Column(db.String(256))  # 论文附件路径
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='journal_papers_owned', foreign_keys=[user_id])
    inclusion_relations = db.relationship('JournalPaperInclusionRelation', back_populates='paper', cascade='all, delete-orphan', overlaps='journal_papers')
    inclusion_types = db.relationship('InclusionType', secondary='journal_paper_inclusion_relation', back_populates='journal_papers', viewonly=True)
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(JournalPaper.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='journal_paper')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')



class ConferencePaper(db.Model):
    """会议论文表（完整字段）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 论文名称
    authors = db.Column(db.Text, nullable=False)  # 论文作者（逗号分隔）
    corresponding_authors = db.Column(db.Text)  # 通讯作者（逗号分隔）
    conference_name = db.Column(db.String(200), nullable=False)  # 会议名称
    conference_time = db.Column(db.Text)  # 会议时间（文本格式，如"2026.03.11-2026.03.13"）
    conference_start_date = db.Column(db.Date)  # 会议开始日期
    conference_end_date = db.Column(db.Date)  # 会议结束日期
    conference_place = db.Column(db.String(200))  # 会议地点
    page_range = db.Column(db.String(50))  # 起止页码
    doi = db.Column(db.String(200))  # DOI
    publish_year = db.Column(db.Integer)  # 发表年份
    attachment = db.Column(db.String(256))  # 论文附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和共同作者）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='conference_papers_owned', foreign_keys=[user_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(ConferencePaper.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='conference_paper')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class TextbookLevel(db.Model):
    """教材级别表（字典表）"""
    id = db.Column(db.Integer, primary_key=True)
    level_name = db.Column(db.String(100), unique=True, nullable=False, comment='级别名称')
    level_code = db.Column(db.String(50), unique=True, nullable=False, comment='级别代码')
    description = db.Column(db.Text, comment='描述说明')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 反向引用
    textbooks = db.relationship('Textbook', back_populates='level')


class Textbook(db.Model):
    """教材表（更新字段）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 教材名称
    textbook_series = db.Column(db.String(500))  # 教材系列（顿号分隔）
    chief_editor = db.Column(db.Text)  # 主编（顿号分隔）
    associate_editors = db.Column(db.Text)  # 副主编（顿号分隔）
    editorial_board = db.Column(db.Text)  # 编委（顿号分隔）
    publisher = db.Column(db.String(200))  # 出版社
    isbn = db.Column(db.String(50))  # ISBN
    cip_number = db.Column(db.String(100))  # CIP 核字号
    publication_year = db.Column(db.Integer)  # 出版年份
    publication_month = db.Column(db.Integer)  # 出版月份
    publish_date = db.Column(db.Date)  # 出版日期（兼容原有字段）
    edition = db.Column(db.String(50))  # 版次
    word_count = db.Column(db.String(50))  # 字数（如 318 千字）
    price = db.Column(db.String(20))  # 定价（如 49.00）
    textbook_level_id = db.Column(db.Integer, db.ForeignKey('textbook_level.id'), comment='教材级别 ID（关联 textbook_level 表）')
    textbook_type = db.Column(db.String(50))  # 教材类型（下拉选择）
    applicable_majors = db.Column(db.Text)  # 适用专业
    remarks = db.Column(db.Text)  # 备注
    textbook_attachment = db.Column(db.String(256))  # 教材附件（替换原 attachment）
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和编者）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='textbooks_owned', foreign_keys=[user_id])
    level = db.relationship('TextbookLevel', back_populates='textbooks')
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(Textbook.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='textbook')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')



class Monograph(db.Model):
    """专著表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 专著名称
    textbook_series = db.Column(db.String(500))  # 专著系列（顿号分隔）
    chief_editor = db.Column(db.Text)  # 主编（顿号分隔）
    associate_editors = db.Column(db.Text)  # 副主编（顿号分隔）
    editorial_board = db.Column(db.Text)  # 编委（顿号分隔）
    publisher = db.Column(db.String(200))  # 出版社
    isbn = db.Column(db.String(50))  # ISBN
    cip_number = db.Column(db.String(100))  # CIP 核字号
    publication_year = db.Column(db.Integer)  # 出版年份
    publication_month = db.Column(db.Integer)  # 出版月份
    publish_date = db.Column(db.Date)  # 出版日期（兼容扩展）
    edition = db.Column(db.String(50))  # 版次（如第 1 版、修订版）
    word_count = db.Column(db.String(50))  # 字数（如"318 千字"）
    price = db.Column(db.String(20))  # 定价（如"49.00 元"）
    monograph_type = db.Column(db.String(50))  # 专著类型（下拉选择）
    applicable_majors = db.Column(db.Text)  # 适用专业
    remarks = db.Column(db.Text)  # 备注
    monograph_attachment = db.Column(db.String(256))  # 专著附件
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和著者）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='monographs_owned', foreign_keys=[user_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(Monograph.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='monograph')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class ProjectType(db.Model):
    """项目类型表"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(100), nullable=False, comment='项目类型名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectType {self.type_name}>'


class ProjectStatus(db.Model):
    """项目状态表"""
    id = db.Column(db.Integer, primary_key=True)
    status_name = db.Column(db.String(50), nullable=False, comment='项目状态名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectStatus {self.status_name}>'


class ProjectLevel(db.Model):
    """项目级别表"""
    id = db.Column(db.Integer, primary_key=True)
    level_name = db.Column(db.String(50), nullable=False, comment='项目级别名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectLevel {self.level_name}>'


class ProjectCategory(db.Model):
    """项目类别表"""
    id = db.Column(db.Integer, primary_key=True)
    category_name = db.Column(db.String(100), nullable=False, comment='项目类别名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectCategory {self.category_name}>'


class TeachingProject(db.Model):
    """教研教改和课程建设项目表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False, comment='项目名称')
    project_code = db.Column(db.String(100), comment='项目编号')
    project_leader = db.Column(db.String(100), comment='项目负责人')
    project_members = db.Column(db.Text, comment='项目参与人（顿号分隔）')
    approval_department = db.Column(db.String(200), comment='项目批准部门')
    approval_date = db.Column(db.Date, comment='项目立项时间')
    project_type_id = db.Column(db.Integer, db.ForeignKey('project_type.id'), comment='项目类型 ID')
    project_level_id = db.Column(db.Integer, db.ForeignKey('project_level.id'), comment='项目级别 ID')
    project_category_id = db.Column(db.Integer, db.ForeignKey('project_category.id'), comment='项目类别 ID')
    project_status_id = db.Column(db.Integer, db.ForeignKey('project_status.id'), comment='项目状态 ID')
    funding = db.Column(db.Numeric(10, 2), comment='项目经费（元）')
    start_date = db.Column(db.Date, comment='项目开始时间')
    end_date = db.Column(db.Date, comment='项目结束时间')
    attachment = db.Column(db.String(256), comment='附件路径')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    # 关联关系
    user = db.relationship('User', backref='teaching_projects_owned', foreign_keys=[user_id])
    project_type = db.relationship('ProjectType', backref='teaching_projects')
    project_level = db.relationship('ProjectLevel', backref='teaching_projects')
    project_category = db.relationship('ProjectCategory', backref='teaching_projects')
    project_status = db.relationship('ProjectStatus', backref='teaching_projects')

    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(TeachingProject.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='teaching_project')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class PatentType(db.Model):
    """专利类型表"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(50), unique=True, nullable=False, comment='专利类型名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<PatentType {self.type_name}>'


class PatentStatus(db.Model):
    """专利状态表"""
    id = db.Column(db.Integer, primary_key=True)
    status_name = db.Column(db.String(50), unique=True, nullable=False, comment='专利状态名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<PatentStatus {self.status_name}>'


class Patent(db.Model):
    """专利表（发明/实用新型/外观）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 专利名称
    patent_type_id = db.Column(db.Integer, db.ForeignKey('patent_type.id'), nullable=False, comment='专利类型 ID（关联 patent_type 表）')
    patentee = db.Column(db.String(200), comment='专利权人')
    address = db.Column(db.String(500), comment='地址')
    inventors = db.Column(db.Text, comment='发明人（多人用分号分隔）')
    patent_status_id = db.Column(db.Integer, db.ForeignKey('patent_status.id'), comment='专利状态 ID（关联 patent_status 表）')
    patent_number = db.Column(db.String(100), unique=True, comment='专利号')
    grant_announcement_number = db.Column(db.String(100), comment='授权公告号')
    apply_date = db.Column(db.Date, comment='专利申请日')
    grant_announcement_date = db.Column(db.Date, comment='授权公告日')
    applicant_at_apply_date = db.Column(db.String(200), comment='申请日时申请人（发明专利）')
    inventor_at_apply_date = db.Column(db.String(200), comment='申请日时发明人（发明专利）')
    attachment = db.Column(db.String(256), comment='附件路径')
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，发明人）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='patents_owned', foreign_keys=[user_id])
    patent_type = db.relationship('PatentType', backref='patents')
    patent_status = db.relationship('PatentStatus', backref='patents')
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(Patent.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='patent')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')





class SoftwareCopyright(db.Model):
    """软件著作表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 软件名称
    copyright_owner = db.Column(db.Text)  # 著作权人（多人用分号分隔）
    completion_date = db.Column(db.Date)  # 开发完成日期
    first_publication_date = db.Column(db.Date)  # 首次发表日期
    right_acquisition_method = db.Column(db.String(100))  # 权利取得方式
    right_scope = db.Column(db.String(200))  # 权利范围
    copyright_number = db.Column(db.String(100))  # 登记号
    certificate_number = db.Column(db.String(100))  # 证书号
    register_date = db.Column(db.Date)  # 登记日期
    attachment = db.Column(db.String(256))  # 附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，著作权人）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='software_copyrights_owned', foreign_keys=[user_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(SoftwareCopyright.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='software_copyright')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class TeachingAchievementAward(db.Model):
    """教学成果获奖表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False, comment='成果名称')
    achievement_type_id = db.Column(db.Integer, db.ForeignKey('teaching_achievement_type.id'), nullable=True, comment='教学成果奖类型 ID（关联 teaching_achievement_type 表）')
    achievement_level_id = db.Column(db.Integer, db.ForeignKey('achievement_level.id'), nullable=True, comment='成果等级 ID（关联 achievement_level 表）')
    main_contributors = db.Column(db.Text, comment='主要完成人（多人用分号分隔）')
    completing_units = db.Column(db.Text, comment='成果完成单位（多个用分号分隔）')
    award_year = db.Column(db.Integer, comment='获奖年度')
    award_rank_id = db.Column(db.Integer, db.ForeignKey('award_rank.id'), nullable=True, comment='获奖等级 ID（关联 award_rank 表）')
    certificate_number = db.Column(db.String(100), comment='证书编号')
    awarding_unit = db.Column(db.String(200), comment='颁奖单位')
    award_date = db.Column(db.Date, comment='获奖日期')
    attachment = db.Column(db.String(256), comment='附件路径')
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，获奖完成人）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='teaching_achievement_awards_owned', foreign_keys=[user_id])
    achievement_type = db.relationship('TeachingAchievementType', backref='teaching_achievement_awards')
    achievement_level = db.relationship('AchievementLevel', backref='teaching_achievement_awards')
    award_rank = db.relationship('AwardRank', backref='teaching_achievement_awards')
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(TeachingAchievementAward.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='teaching_achievement_award')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class TeachingAchievementType(db.Model):
    """教学成果奖类型表（管理员可维护）"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(100), unique=True, nullable=False, comment='类型名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<TeachingAchievementType {self.type_name}>'


class AchievementLevel(db.Model):
    """成果等级表（管理员可维护）"""
    id = db.Column(db.Integer, primary_key=True)
    level_name = db.Column(db.String(50), unique=True, nullable=False, comment='等级名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<AchievementLevel {self.level_name}>'


class AwardRank(db.Model):
    """获奖等级表（管理员可维护，多处共用）"""
    id = db.Column(db.Integer, primary_key=True)
    rank_name = db.Column(db.String(50), unique=True, nullable=False, comment='等级名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<AwardRank {self.rank_name}>'


class TeachingCompetitionAward(db.Model):
    """教学竞赛获奖表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 竞赛名称
    award_year = db.Column(db.String(50))  # 获奖年度
    competition_level_id = db.Column(db.Integer, db.ForeignKey('achievement_level.id'))  # 竞赛等级 id
    award_rank_id = db.Column(db.Integer, db.ForeignKey('award_rank.id'))  # 获奖等级 id
    winners = db.Column(db.Text)  # 获奖人（多个用分号分隔）
    winner_unit = db.Column(db.String(500))  # 获奖人所在单位
    competition_name = db.Column(db.String(200))  # 竞赛主办方
    award_date = db.Column(db.Date)  # 获奖日期
    certificate_number = db.Column(db.String(100))  # 证书编号
    attachment = db.Column(db.String(256))  # 附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，参赛教师）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='teaching_competition_awards_owned', foreign_keys=[user_id])
    competition_level = db.relationship('AchievementLevel', backref='teaching_competition_awards_competition_level', foreign_keys=[competition_level_id])
    award_rank = db.relationship('AwardRank', backref='teaching_competition_awards_award_rank', foreign_keys=[award_rank_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(TeachingCompetitionAward.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='teaching_competition_award')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class StudentGuidanceAward(db.Model):
    """指导学生获奖表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    award_year = db.Column(db.String(50))  # 获奖年度
    title = db.Column(db.String(500), nullable=False)  # 获奖名称
    competition_name = db.Column(db.String(200))  # 竞赛名称
    competition_level_id = db.Column(db.Integer, db.ForeignKey('achievement_level.id'))  # 竞赛等级 id
    award_rank_id = db.Column(db.Integer, db.ForeignKey('award_rank.id'))  # 获奖等级 id
    student_name = db.Column(db.String(200))  # 获奖学生
    project_name = db.Column(db.String(500))  # 获奖项目名称
    teacher_name = db.Column(db.String(200))  # 指导教师
    student_unit = db.Column(db.String(500))  # 获奖学生所在单位
    organizer = db.Column(db.String(500))  # 竞赛主办方
    certificate_number = db.Column(db.String(100))  # 证书编号
    award_date = db.Column(db.Date)  # 获奖日期
    attachment = db.Column(db.String(256))  # 附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，指导教师）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='student_guidance_awards_owned', foreign_keys=[user_id])
    competition_level = db.relationship('AchievementLevel', backref='student_guidance_awards_competition_level', foreign_keys=[competition_level_id])
    award_rank = db.relationship('AwardRank', backref='student_guidance_awards_award_rank', foreign_keys=[award_rank_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(StudentGuidanceAward.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='student_guidance_award')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')

class SystemConfig(db.Model):
    """系统全局配置表（仅管理员可维护）"""
    id = db.Column(db.Integer, primary_key=True)
    config_key = db.Column(db.String(100), unique=True, nullable=False)  # 配置项标识（如system_name、max_upload_size）
    config_value = db.Column(db.Text)  # 配置值（字符串/JSON）
    config_desc = db.Column(db.String(200))  # 配置项描述
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    @classmethod
    def get_config(cls, key, default=''):
        """获取系统配置值（无则返回默认值）"""
        config = cls.query.filter_by(config_key=key).first()
        return config.config_value if config else default

    @classmethod
    def set_config(cls, key, value, desc=''):
        """设置系统配置值（不存在则创建，存在则更新）"""
        config = cls.query.filter_by(config_key=key).first()
        if not config:
            config = cls(config_key=key, config_value=value, config_desc=desc)
            db.session.add(config)
        else:
            config.config_value = value
            config.config_desc = desc
        db.session.commit()
        return config


# ---------------------- 3. 辅助函数 ----------------------
def get_current_user():
    """获取当前登录用户 - 修复SQLAlchemy 2.0警告"""
    if 'user_id' in session:
        # 替换过时的 Query.get() 为 Session.get()
        return db.session.get(User, session['user_id'])
    return None


def generate_nav_menu(user):
    """生成左侧导航菜单（根据角色）"""
    # 基础菜单（所有登录用户可见）
    base_menu = [
        '<li><a href="/">首页</a></li>'
    ]

    if user.role == 'teacher':
        base_menu.append('<li><a href="/user/settings">个人账户设置</a></li>')
        base_menu.append('<li><a href="/stats/dashboard">📊 数据统计仪表盘</a></li>')
        base_menu.append('<li><a href="/achievement/ocr_import">📷 OCR 智能导入</a></li>')
        base_menu.append('<li><a href="/achievement/voice_export">🎙️ 语音导出</a></li>')



    elif user.role == 'team_leader':
        base_menu.append('<li><a href="/user/settings">个人账户设置</a></li>')
        base_menu.append('<li><a href="/team/voice_export">🎙️ 团队语音导出</a></li>')

    # 成果管理菜单（仅普通教师可见）
    achievement_menu = [
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">📄 论文管理</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/journal_paper">期刊论文</a></li>',
        '    <li><a href="/achievement/conference_paper">会议论文</a></li>',
        '  </ul>',
        '</li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">📚 教材与专著</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/textbook">教材</a></li>',
        '    <li><a href="/achievement/monograph">专著</a></li>',
        '  </ul>',
        '</li>',
        '<li><a href="/achievement/teaching_project">🔬 教研教改和课程建设项目</a></li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">💡 专利与软著</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/patent">专利</a></li>',
        '    <li><a href="/achievement/software_copyright">软件著作</a></li>',
        '  </ul>',
        '</li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">🏆 获奖管理</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/teaching_achievement_award">教学成果获奖</a></li>',
        '    <li><a href="/achievement/teaching_competition_award">教学竞赛获奖</a></li>',
        '    <li><a href="/achievement/student_guidance_award">指导学生获奖</a></li>',
        '  </ul>',
        '</li>',
    ]

    # 管理员专属菜单（核心修改：移除子菜单，直接显示用户/团队管理）
    admin_menu = [
        '<li><a href="/admin/user_manage">👥 用户管理</a></li>',
        '<li><a href="/admin/team_manage">🏢 团队管理</a></li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">⚙️ 字典管理</span>',
        '  <ul class="submenu">',
        '    <li><a href="/admin/dict_manage/achievement_type">教学成果奖类型</a></li>',
        '    <li><a href="/admin/dict_manage/achievement_level">成果等级</a></li>',
        '    <li><a href="/admin/dict_manage/award_rank">获奖等级</a></li>',
        '  </ul>',
        '</li>'
    ]

    # 团队负责人专属菜单（仅保留带自子菜单的团队管理）
    leader_menu = [
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">👥 团队管理</span>',
        '  <ul class="submenu">',
        '    <li><a href="/team/list">📋 查看团队</a></li>',  # 重点：指向新的团队列表页面
        '    <li><a href="/team/achievements">📊 团队成果统计</a></li>',
        '    <li><a href="/team/manage_members">👨‍🏫 团队成员管理</a></li>',
        '    <li><a href="/team/member_achievements">📋 成员成果详情</a></li>',
        '  </ul>',
        '</li>'
    ]

    # 拼接最终菜单
    menu_html = '<ul class="sidebar-menu">'
    menu_html += ''.join(base_menu)

    # 仅普通教师显示成果管理菜单
    if user.role == 'teacher':
        menu_html += ''.join(achievement_menu)
    # 仅团队负责人显示团队管理菜单
    elif user.role == 'team_leader':
        menu_html += ''.join(leader_menu)

    # 管理员显示系统管理菜单
    if user.role == 'admin':
        menu_html += ''.join(admin_menu)

    menu_html += '<li><a href="/logout">🚪 退出登录</a></li></ul>'
    return menu_html


def render_base_layout(title, content, user):
    """渲染基础布局（左右布局，禁止 Jinja）"""
    nav_menu = generate_nav_menu(user) if user else ''
    user_info = f'欢迎，{user.username}（{user.role}）' if user else '未登录'

    # 获取 flash 消息
    from flask import get_flashed_messages
    flashed_messages = get_flashed_messages(with_categories=True)

    # 生成 flash 消息 HTML
    messages_html = ''
    if flashed_messages:
        for category, message in flashed_messages:
            messages_html += f'<div class="alert alert-{category}">{message}</div>'

    html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} - 教学成果管理系统</title>
    <style>
        /* 全局样式 */
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, "Microsoft YaHei", sans-serif;
        }}
        body {{
            background: linear-gradient(135deg, #87CEEB 0%, #4A90E2 100%);
            min-height: 100vh;
        }}
        /* 左侧导航栏 */
        .sidebar {{
            width: 260px;
            background: linear-gradient(180deg, #2c3e50 0%, #1a252f 100%);
            color: white;
            min-height: 100vh;
            padding: 0;
            box-shadow: 4px 0 10px rgba(0,0,0,0.15);
            position: fixed;
            top: 0;
            left: 0;
            z-index: 1000;
            transition: all 0.3s ease;
        }}
        .sidebar:hover {{
            box-shadow: 4px 0 15px rgba(0,0,0,0.25);
        }}
        .sidebar-header {{
            padding: 25px 20px;
            background: rgba(0,0,0,0.2);
            border-bottom: 2px solid rgba(255,255,255,0.1);
            margin-bottom: 0;
        }}
        .sidebar-header h2 {{
            font-size: 20px;
            font-weight: 700;
            color: #3498db;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
            letter-spacing: 1px;
        }}
        .sidebar-menu {{
            list-style: none;
            padding: 15px 10px;
        }}
        .sidebar-menu li {{
            margin: 8px 0;
        }}
        .sidebar-menu a {{
            display: block;
            padding: 14px 20px;
            color: #ecf0f1;
            text-decoration: none;
            border-radius: 8px;
            transition: all 0.3s ease;
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }}
        .sidebar-menu a::before {{
            content: '';
            position: absolute;
            left: 0;
            top: 0;
            height: 100%;
            width: 4px;
            background: #3498db;
            transform: scaleY(0);
            transition: transform 0.3s ease;
        }}
        .sidebar-menu a:hover {{
            background: rgba(52, 152, 219, 0.15);
            color: #3498db;
            padding-left: 25px;
        }}
        .sidebar-menu a:hover::before {{
            transform: scaleY(1);
        }}
        .menu-group span {{
            display: block;
            padding: 14px 20px;
            color: #bdc3c7;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            border-radius: 8px;
            position: relative;
        }}
        .menu-group span:hover {{
            color: #3498db;
            background: rgba(52, 152, 219, 0.1);
        }}
        .submenu {{
            list-style: none;
            padding-left: 15px;
            display: none;
            margin-top: 5px;
        }}
        .submenu.active {{
            display: block;
            animation: slideDown 0.3s ease;
        }}
        @keyframes slideDown {{
            from {{
                opacity: 0;
                transform: translateY(-10px);
            }}
            to {{
                opacity: 1;
                transform: translateY(0);
            }}
        }}
        .submenu a {{
            padding: 10px 20px;
            font-size: 14px;
            color: #95a5a6;
            border-left: 2px solid transparent;
        }}
        .submenu a:hover {{
            color: #3498db;
            border-left-color: #3498db;
            background: rgba(52, 152, 219, 0.05);
        }}
        /* 右侧内容区 */
        .content {{
            padding: 30px;
            margin-left: 260px;
            width: calc(100% - 260px);
            min-height: 100vh;
        }}
        .user-info {{
            text-align: right;
            margin-bottom: 15px;
            color: rgba(255,255,255,0.9);
            font-weight: 600;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
            padding: 10px 20px;
            background: rgba(255,255,255,0.1);
            border-radius: 8px;
            backdrop-filter: blur(10px);
        }}
        .content-header {{
            margin-bottom: 25px;
            padding-bottom: 15px;
            border-bottom: 2px solid rgba(255,255,255,0.2);
        }}
        .content-header h1 {{
            font-size: 28px;
            color: white;
            font-weight: 700;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}
        .container {{
            background: rgba(255, 255, 255, 0.98);
            border-radius: 12px;
            box-shadow: 0 8px 32px rgba(0,0,0,0.1);
            padding: 35px;
            min-height: 500px;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.5);
        }}
        /* 表单样式 */
        .form-group {{
            margin-bottom: 22px;
            position: relative;
        }}
        label {{
            display: block;
            margin-bottom: 10px;
            font-weight: 600;
            color: #2c3e50;
            font-size: 15px;
            transition: color 0.3s ease;
        }}
        .form-group:hover label {{
            color: #3498db;
        }}
        input, select, textarea {{
            width: 100%;
            padding: 12px 16px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 14px;
            transition: all 0.3s ease;
            background: white;
        }}
        input:focus, select:focus, textarea:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 0 4px rgba(52, 152, 219, 0.1);
            transform: translateY(-2px);
        }}
        input[type="file"] {{
            padding: 10px;
            border: 2px dashed #bdc3c7;
            border-radius: 8px;
            background: #fafafa;
            cursor: pointer;
        }}
        input[type="file"]:hover {{
            border-color: #3498db;
            background: #f0f7ff;
        }}
        .btn {{
            display: inline-block;
            padding: 12px 24px;
            background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s ease;
            text-decoration: none;
            box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
            position: relative;
            overflow: hidden;
        }}
        .btn::before {{
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255,255,255,0.2);
            transform: translate(-50%, -50%);
            transition: width 0.6s, height 0.6s;
        }}
        .btn:hover::before {{
            width: 300px;
            height: 300px;
        }}
        .btn:hover {{
            transform: translateY(-3px);
            box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4);
        }}
        .btn:active {{
            transform: translateY(-1px);
        }}
        .btn-success {{
            background: linear-gradient(135deg, #27ae60 0%, #229954 100%);
            box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
        }}
        .btn-success:hover {{
            box-shadow: 0 6px 20px rgba(39, 174, 96, 0.4);
        }}
        .btn-warning {{
            background: linear-gradient(135deg, #f39c12 0%, #d68910 100%);
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.3);
        }}
        .btn-warning:hover {{
            box-shadow: 0 6px 20px rgba(243, 156, 18, 0.4);
        }}
        .btn-danger {{
            background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.3);
        }}
        .btn-danger:hover {{
            box-shadow: 0 6px 20px rgba(231, 76, 60, 0.4);
        }}
        .btn-info {{
            background: linear-gradient(135deg, #17a2b8 0%, #138496 100%);
            box-shadow: 0 4px 15px rgba(23, 162, 184, 0.3);
        }}
        .btn-info:hover {{
            box-shadow: 0 6px 20px rgba(23, 162, 184, 0.4);
        }}
        .btn-secondary {{
            background: linear-gradient(135deg, #6c757d 0%, #5a6268 100%);
            box-shadow: 0 4px 15px rgba(108, 117, 125, 0.3);
        }}
        /* 提示框样式 */
        .alert {{
            padding: 16px 20px;
            margin-bottom: 20px;
            border-radius: 8px;
            border-left: 4px solid;
            animation: slideIn 0.5s ease;
            font-weight: 500;
        }}
        @keyframes slideIn {{
            from {{
                opacity: 0;
                transform: translateX(-20px);
            }}
            to {{
                opacity: 1;
                transform: translateX(0);
            }}
        }}
        .alert-success {{
            background: #d4edda;
            color: #155724;
            border-color: #28a745;
        }}
        .alert-danger {{
            background: #f8d7da;
            color: #721c24;
            border-color: #dc3545;
        }}
        .alert-warning {{
            background: #fff3cd;
            color: #856404;
            border-color: #ffc107;
        }}
        .alert-info {{
            background: #d1ecf1;
            color: #0c5460;
            border-color: #17a2b8;
        }}
        /* 表格样式优化 */
        table {{
            border-collapse: separate;
            border-spacing: 0;
            width: 100%;
            margin-top: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
        }}
        thead tr {{
            background: linear-gradient(135deg, #87CEEB 0%, #4A90E2 100%);
            color: black;
        }}
        thead th {{
            padding: 16px 12px;
            border: none;
            font-weight: 600;
            text-align: left;
            text-transform: uppercase;
            font-size: 13px;
            letter-spacing: 0.5px;
        }}
        tbody tr {{
            transition: all 0.3s ease;
            border-bottom: 1px solid #e9ecef;
        }}
        tbody tr:hover {{
            background: #f8f9fa;
            transform: scale(1.01);
        }}
        tbody tr:last-child {{
            border-bottom: none;
        }}
        tbody td {{
            padding: 14px 12px;
            border: none;
            color: #495057;
            font-size: 14px;
        }}
        /* 滚动条美化 */
        ::-webkit-scrollbar {{
            width: 10px;
            height: 10px;
        }}
        ::-webkit-scrollbar-track {{
            background: #f1f1f1;
            border-radius: 5px;
        }}
        ::-webkit-scrollbar-thumb {{
            background: #3498db;
            border-radius: 5px;
        }}
        ::-webkit-scrollbar-thumb:hover {{
            background: #2980b9;
        }}
        /* 响应式设计 */
        @media (max-width: 768px) {{
            .sidebar {{
                width: 200px;
            }}
            .content {{
                margin-left: 200px;
                width: calc(100% - 200px);
            }}
            .sidebar-header h2 {{
                font-size: 16px;
            }}
        }}
    </style>
</head>
<body>
    <!-- 左侧导航栏 -->
    <div class="sidebar">
        <div class="sidebar-header">
            <h2>🎓 教学成果管理系统</h2>
        </div>
        {nav_menu}
    </div>

    <!-- 右侧内容区 -->
    <div class="content">
        <div class="user-info">👤 {user_info}</div>
        <div class="content-header">
            <h1>{title}</h1>
        </div>
        <div class="container">
            {messages_html}
            {content}
        </div>
    </div>

    <!-- 新增：子菜单切换脚本 -->
    <script>
        function toggleSubmenu(el) {{
            const submenu = el.nextElementSibling;
            if (submenu && submenu.classList.contains('submenu')) {{
                submenu.classList.toggle('active');
                const span = el;
                span.style.color = submenu.classList.contains('active') ? '#3498db' : '#bdc3c7';
            }}
        }}

        // 页面加载完成后的初始化
        document.addEventListener('DOMContentLoaded', function() {{
            // 为所有按钮添加点击波纹效果
            const buttons = document.querySelectorAll('.btn');
            buttons.forEach(btn => {{
                btn.addEventListener('click', function(e) {{
                    const ripple = document.createElement('span');
                    const rect = btn.getBoundingClientRect();
                    const size = Math.max(rect.width, rect.height);
                    const x = e.clientX - rect.left - size/2;
                    const y = e.clientY - rect.top - size/2;
                    ripple.style.width = ripple.style.height = size + 'px';
                    ripple.style.left = x + 'px';
                    ripple.style.top = y + 'px';
                    ripple.classList.add('ripple');
                    btn.appendChild(ripple);
                    setTimeout(() => ripple.remove(), 600);
                }});
            }});
        }});
    </script>

    <style>
        .ripple {{
            position: absolute;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.6);
            transform: scale(0);
            animation: ripple-animation 0.6s ease-out;
            pointer-events: none;
        }}
        @keyframes ripple-animation {{
            to {{
                transform: scale(4);
                opacity: 0;
            }}
        }}
    </style>
</body>
</html>
'''
    return html

def allowed_file(filename):
    """校验上传文件扩展名"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def handle_file_upload(file, sub_folder):
    """通用文件上传处理：返回文件存储路径"""
    if file and allowed_file(file.filename):
        # 创建子目录（按成果类型分类存储）
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], sub_folder)
        if not os.path.exists(upload_path):
            os.makedirs(upload_path)

        # 安全文件名 + 时间戳避免重复
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        new_filename = f"{timestamp}_{filename}"
        file_path = os.path.join(upload_path, new_filename)

        # 保存文件
        file.save(file_path)
        return file_path
    return None


def get_team_user_ids(current_user):
    """获取团队内所有用户 ID（团队负责人用）"""
    if current_user.role != 'team_leader':
        return [current_user.id]

    teams = Team.query.filter_by(leader_id=current_user.id).all()
    team_ids = [t.id for t in teams]

    user_teams = UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()
    team_user_ids = [ut.user_id for ut in user_teams] + [current_user.id]
    return team_user_ids


def auto_link_contributors(achievement, achievement_type, authors_str, creator_user_id):
    """自动关联成果与系统用户作者"""
    if not authors_str:
        return

    # 修复：同时支持逗号、分号、顿号等多种分隔符
    import re
    author_names = [name.strip() for name in re.split(r'[;,;,,]', authors_str) if name.strip()]
    related_user_ids = set()

    for author_name in author_names:
        users = User.query.filter(
            (User.username == author_name) |
            (User.employee_id == author_name) |
            (User.email.like(f'%{author_name}%'))
        ).all()

        for user in users:
            related_user_ids.add(user.id)

            contributor = AchievementContributor.query.filter_by(
                achievement_type=achievement_type,
                achievement_id=achievement.id,
                user_id=user.id
            ).first()

            if not contributor:
                contributor = AchievementContributor(
                    achievement_type=achievement_type,
                    achievement_id=achievement.id,
                    user_id=user.id,
                    contributor_role='author',
                    is_creator=(user.id == creator_user_id)
                )
                db.session.add(contributor)

    if related_user_ids:
        achievement.related_personnel_ids = ','.join(map(str, related_user_ids))


def render_achievement_list(model, title, fields_config, current_user):
    """通用成果列表页面渲染"""
    # 权限过滤：普通用户看自己的，团队负责人看团队的，管理员无权查看
    if current_user.role == 'admin':
        # 管理员无权查看成果，直接返回提示
        content = '<div class="alert alert-danger">管理员无权查看/操作教师个人成果！</div>'
        return render_base_layout(title, content, current_user)

    if current_user.role == 'team_leader':
        # 1. 获取当前用户管理的所有团队ID（核心：仅能看公开给自己团队的成果）
        managed_teams = Team.query.filter_by(leader_id=current_user.id).all()
        managed_team_ids = [str(t.id) for t in managed_teams]

        # 2. 获取团队内所有成员ID（用于基础筛选）
        user_teams = UserTeam.query.filter(UserTeam.team_id.in_([t.id for t in managed_teams])).all()
        team_user_ids = [ut.user_id for ut in user_teams] + [current_user.id]

        # 3. 核心查询逻辑：
        # - 成果属于团队成员
        # - 且（是自己的成果 OR 成果公开给当前用户管理的任意团队）
        query = model.query.filter(model.user_id.in_(team_user_ids))
        or_conditions = [model.user_id == current_user.id]

        # 遍历当前用户管理的每个团队ID，检查是否在public_team_ids中
        for team_id in managed_team_ids:
            # 处理public_team_ids格式：",1,2,3," 避免部分匹配（如1匹配10）
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )

        query = query.filter(or_(*or_conditions))

    elif current_user.role == 'teacher':
        # 普通教师：仅查看自己的成果
        query = model.query.filter_by(user_id=current_user.id)

    # 分页查询
    page = request.args.get('page', 1, type=int)
    per_page = 5
    pagination = query.order_by(model.update_time.desc()).paginate(page=page, per_page=per_page)
    items = pagination.items

    # 生成列表HTML（添加智能导入按钮）
    table_html = '''
            <div style="margin-bottom:20px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;">
            <a href="?action=add" class="btn btn-primary">新增</a>
            <a href="?action=export" class="btn btn-success">导出</a>
            <a href="?action=stats" class="btn btn-warning">统计分析</a>
        '''
    # 根据成果类型添加智能导入按钮
    type_key = ''
    if model.__name__ == 'JournalPaper':
        type_key = 'journal_paper'
        table_html += '<a href="/achievement/journal_paper/import" class="btn btn-danger">知网智能导入</a>'
    elif model.__name__ == 'ConferencePaper':
        type_key = 'conference_paper'
        table_html += '<a href="/achievement/conference_paper/import" class="btn btn-danger">知网智能导入</a>'
    elif model.__name__ == 'Textbook':
        type_key = 'textbook'
    elif model.__name__ == 'Monograph':
        type_key = 'monograph'
    elif model.__name__ == 'TeachingProject':
        type_key = 'teaching_project'
    elif model.__name__ == 'Patent':
        type_key = 'patent'
    elif model.__name__ == 'SoftwareCopyright':
        type_key = 'software_copyright'
    elif model.__name__ == 'TeachingAchievementAward':
        type_key = 'teaching_achievement_award'
    elif model.__name__ == 'TeachingCompetitionAward':
        type_key = 'teaching_competition_award'
    elif model.__name__ == 'StudentGuidanceAward':
        type_key = 'student_guidance_award'

    # 添加批量导入按钮（如果有对应的类型）
    if type_key:
        table_html += f'<a href="/achievement/batch_import/{type_key}" class="btn btn-info">批量导入</a>'

    table_html += '''
    </div>
    <table style="width:100%; border-collapse: collapse; margin-top:20px;">
        <thead>
            <tr style="background:#f5f7fa;">
                <th style="padding:10px; border:1px solid #dee2e6;">序号</th>
    '''
    # 生成表头
    for field in fields_config:
        table_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{field["label"]}</th>'
    table_html += '''
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    # 生成表体
    for idx, item in enumerate(items, 1):
        table_html += '<tr>'
        table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{idx}</td>'

        # 生成字段值
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            # 处理外键关联字段（显示名称而非 ID）
            if field_name == 'achievement_type_id' and value:
                achievement_type = db.session.get(TeachingAchievementType, value)
                value = achievement_type.type_name if achievement_type else value
            elif field_name == 'achievement_level_id' and value:
                achievement_level = db.session.get(AchievementLevel, value)
                value = achievement_level.level_name if achievement_level else value
            elif field_name == 'competition_level_id' and value:
                competition_level = db.session.get(AchievementLevel, value)
                value = competition_level.level_name if competition_level else value
            elif field_name == 'award_rank_id' and value:
                award_rank = db.session.get(AwardRank, value)
                value = award_rank.rank_name if award_rank else value
            elif field_name == 'project_type_id' and value:
                project_type = db.session.get(ProjectType, value)
                value = project_type.type_name if project_type else value
            elif field_name == 'project_level_id' and value:
                project_level = db.session.get(ProjectLevel, value)
                value = project_level.level_name if project_level else value
            elif field_name == 'project_category_id' and value:
                project_category = db.session.get(ProjectCategory, value)
                value = project_category.category_name if project_category else value
            elif field_name == 'project_status_id' and value:
                project_status = db.session.get(ProjectStatus, value)
                value = project_status.status_name if project_status else value
            elif field_name == 'patent_type_id' and value:
                patent_type = db.session.get(PatentType, value)
                value = patent_type.type_name if patent_type else value
            elif field_name == 'patent_status_id' and value:
                patent_status = db.session.get(PatentStatus, value)
                value = patent_status.status_name if patent_status else value
            elif field_name == 'textbook_level_id' and value:
                textbook_level = db.session.get(TextbookLevel, value)
                value = textbook_level.level_name if textbook_level else value

                # 核心修复：处理期刊论文的收录情况字段（将 ID 列表转为名称列表）
            elif field_name == 'inclusion_type_ids' and value:
                inclusion_ids = [id_str.strip() for id_str in str(value).split(',') if id_str.strip()]
                inclusion_names = []
                for inc_id in inclusion_ids:
                    try:
                        inc_type = db.session.get(InclusionType, int(inc_id))
                        if inc_type:
                            inclusion_names.append(inc_type.type_name)
                    except (ValueError, TypeError):
                        continue
                value = ','.join(inclusion_names) if inclusion_names else ''

                # 处理 NULL 值，转为空字符串
            if value is None:
                value = ''
                # 特殊处理日期字段
            elif isinstance(value, date) or isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d') if value else ''
                # 附件字段显示下载链接
            elif field_name == 'attachment' and value:
                filename = os.path.basename(value)
                value = f'<a href="/download?path={value}" target="_blank">📎 {filename}</a>' if value else ''

            table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{value}</td>'
        # 操作列（仅自己的成果可修改/删除）【核心修改：删除语音导出按钮】
        ops = ''
        if item.user_id == current_user.id:
            ops = f'''
                <a href="?action=edit&id={item.id}" class="btn" style="padding:5px 10px; font-size:12px;">修改</a>
                <a href="?action=delete&id={item.id}" class="btn " style="padding:5px 10px; font-size:12px;" onclick="return confirm('确定删除？')">删除</a>
            '''
        else:
            ops = '仅查看'

        table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{ops}</td>'
        table_html += '</tr>'

    table_html += '''
        </tbody>
    </table>
    '''

    # 分页控件
    pagination_html = '''
    <div style="margin-top:20px; text-align:center;">
    '''
    if pagination.has_prev:
        pagination_html += f'<a href="?page={pagination.prev_num}" class="btn " style="margin:0 5px;">上一页</a>'
    if pagination.has_next:
        pagination_html += f'<a href="?page={pagination.next_num}" class="btn " style="margin:0 5px;">下一页</a>'
    pagination_html += f'''
        <span style="margin:0 10px;">第{pagination.page}页 / 共{pagination.pages}页</span>
    </div>
    '''

    content = table_html + pagination_html
    return render_base_layout(title, content, current_user)


def render_achievement_form(model, title, fields_config, item_id=None):
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    item = db.session.get(model, item_id) if item_id else None
    if item and item.user_id != current_user.id:
        flash('无权限修改该成果！', 'danger')
        return redirect(request.referrer or url_for('index'))

    form_html = f'''
    <form method="POST" enctype="multipart/form-data">
        <input type="hidden" name="id" value="{item_id or ''}">
    '''

    for field in fields_config:
        field_name = field['name']
        field_label = field['label']
        field_type = field.get('type', 'text')
        required = 'required' if field.get('required', False) else ''
        current_value = getattr(item, field_name, None) if item else None  # 初始值设为 None

        # 核心修改：处理 NULL 值
        if current_value is None:
            current_value = ''
        # 日期字段处理（核心优化）
        elif field_type == 'date':
            current_value = current_value.strftime('%Y-%m-%d') if current_value else ''  # 空值显示为空字符串，但提交时转为 None

        # 修复 Select 下拉框渲染逻辑（重点）
        if field_type == 'select':
            # 特殊处理：教材级别（从 TextbookLevel 表读取）
            if field_name == 'textbook_level_id':
                levels = TextbookLevel.query.filter_by(is_active=True).order_by(TextbookLevel.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for level in levels:
                    selected = 'selected' if str(level.id) == str(current_value) else ''
                    form_html += f'<option value="{level.id}" {selected}>{level.level_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：竞赛等级（从 AchievementLevel 表读取）
            if field_name == 'competition_level_id':
                levels = AchievementLevel.query.filter_by(is_active=True).order_by(AchievementLevel.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for level in levels:
                    selected = 'selected' if str(level.id) == str(current_value) else ''
                    form_html += f'<option value="{level.id}" {selected}>{level.level_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：获奖等级（从 AwardRank 表读取）
            if field_name == 'award_rank_id':
                ranks = AwardRank.query.filter_by(is_active=True).order_by(AwardRank.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for rank in ranks:
                    selected = 'selected' if str(rank.id) == str(current_value) else ''
                    form_html += f'<option value="{rank.id}" {selected}>{rank.rank_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：专利类型（从 PatentType 表读取，选项为元组格式）
            if field_name == 'patent_type_id':
                patent_types = PatentType.query.order_by(PatentType.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for pt in patent_types:
                    selected = 'selected' if str(pt.id) == str(current_value) else ''
                    form_html += f'<option value="{pt.id}" {selected}>{pt.type_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：专利状态（从 PatentStatus 表读取，选项为元组格式）
            if field_name == 'patent_status_id':
                patent_statuses = PatentStatus.query.order_by(PatentStatus.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for ps in patent_statuses:
                    selected = 'selected' if str(ps.id) == str(current_value) else ''
                    form_html += f'<option value="{ps.id}" {selected}>{ps.status_name}</option>'
                form_html += '</select></div>'
                continue

            options = field.get('options', [])
            form_html += f'<div class="form-group">'
            form_html += f'<label>{field_label} {"*" if required else ""}</label>'
            form_html += f'<select name="{field_name}" {required}>'
            # 先添加默认空选项
            form_html += '<option value="">请选择</option>'
            # 遍历所有选项并正确设置 selected 状态
            for opt in options:
                # 兼容元组格式 (id, name) 和简单字符串格式
                if isinstance(opt, tuple) and len(opt) == 2:
                    opt_id, opt_name = opt
                    selected = 'selected' if str(current_value) == str(opt_id) else ''
                    form_html += f'<option value="{opt_id}" {selected}>{opt_name}</option>'
                else:
                    # 简单字符串格式
                    selected = 'selected' if str(current_value) == str(opt) else ''
                    form_html += f'<option value="{opt}" {selected}>{opt}</option>'
            form_html += '</select></div>'
            continue


        elif field_type == 'select_multiple':
            # 特殊处理：收录类型多选框（从 InclusionType 表读取）
            if field_name == 'inclusion_type_ids':
                inclusion_types = InclusionType.query.filter_by(is_active=True).order_by(InclusionType.sort_order).all()
                selected_ids = []
                if current_value and current_value.strip():
                    selected_ids = [id_str.strip() for id_str in current_value.split(',') if id_str.strip()]

                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" multiple size="10" {required}>'
                form_html += '<option value="" style="display:none;"></option>'
                for inc_type in inclusion_types:
                    selected = 'selected' if str(inc_type.id) in selected_ids else ''
                    form_html += f'<option value="{inc_type.id}" {selected}>{inc_type.type_name} ({inc_type.type_code})</option>'
                form_html += '</select>'
                form_html += '<p style="margin-top:5px; color:#666;">提示：按住 Ctrl 键可多选/取消选择</p></div>'
                continue

            # 普通多选框（团队等）
            current_user = get_current_user()
            teams = Team.query.all()  # 可根据权限过滤（如仅显示用户加入/管理的团队）
            selected_ids = []
            if current_value and current_value.strip():
                selected_ids = [id_str.strip() for id_str in current_value.split(',') if id_str.strip()]
            form_html += f'<div class="form-group">'
            form_html += f'<label>{field_label} {"*" if required else ""}</label>'
            form_html += f'<select name="{field_name}" multiple size="5" {required}>'
            form_html += '<option value="" style="display:none;"></option>'  # 新增：隐藏空选项
            for team in teams:
                selected = 'selected' if str(team.id) in selected_ids else ''
                form_html += f'<option value="{team.id}" {selected}>{team.name}</option>'
            form_html += '</select>'
            form_html += '<p style="margin-top:5px; color:#666;">提示：按住 Ctrl 键可多选/取消选择</p></div>'
            # 核心添加：跳过后续普通输入框渲染，避免重复
            continue

        # 处理文件上传字段
        elif field_type == 'file':
            form_html += f'<div class="form-group">'
            form_html += f'<label>{field_label}</label>'
            form_html += f'<input type="file" name="{field_name}" accept=".pdf,.docx,.doc,.png,.jpg,.jpeg">'
            # 显示已上传的文件
            if item and getattr(item, field_name, ''):
                filename = os.path.basename(getattr(item, field_name))
                form_html += f'<p style="margin-top:5px;">当前文件：<a href="/download?path={getattr(item, field_name)}" target="_blank">{filename}</a></p>'
            form_html += '</div>'
            continue

        # 普通输入框（文本/整数/日期）
        form_html += f'<div class="form-group">'
        form_html += f'<label>{field_label} {"*" if required else ""}</label>'
        form_html += f'<input type="{field_type}" name="{field_name}" value="{current_value or ""}" {required}>'
        form_html += '</div>'

    form_html += '''
        <div class="form-group">
            <button type="submit" class="btn">保存</button>
            <a href="javascript:history.back()" class="btn" style="background-color:#95a5a6; margin-left:10px;">取消</a>
        </div>
    </form>
    '''

    return render_base_layout(title, form_html, current_user)


def handle_achievement_submit(model, fields_config):
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    item_id = request.form.get('id')
    item = db.session.get(model, item_id) if item_id else None

    if item and item.user_id != current_user.id:
        flash('无权限修改该成果！', 'danger')
        return redirect(request.referrer or url_for('index'))

    if not item:
        item = model()
        item.user_id = current_user.id
        item.create_time = datetime.now()

        title_value = request.form.get('title', '').strip()
        if title_value:
            # 根据模型类型构建查重查询
            duplicate_check = None
            if model == JournalPaper:
                duplicate_check = JournalPaper.query.filter_by(title=title_value).first()
            elif model == ConferencePaper:
                duplicate_check = ConferencePaper.query.filter_by(title=title_value).first()
            elif model == Textbook:
                duplicate_check = Textbook.query.filter_by(title=title_value).first()
            elif model == Monograph:
                duplicate_check = Monograph.query.filter_by(title=title_value).first()
            elif model == TeachingProject:
                duplicate_check = TeachingProject.query.filter_by(title=title_value).first()
            elif model == Patent:
                duplicate_check = Patent.query.filter_by(title=title_value).first()
            elif model == SoftwareCopyright:
                duplicate_check = SoftwareCopyright.query.filter_by(title=title_value).first()
            elif model == TeachingAchievementAward:
                duplicate_check = TeachingAchievementAward.query.filter_by(title=title_value).first()
            elif model == TeachingCompetitionAward:
                duplicate_check = TeachingCompetitionAward.query.filter_by(title=title_value).first()
            elif model == StudentGuidanceAward:
                duplicate_check = StudentGuidanceAward.query.filter_by(title=title_value).first()

            if duplicate_check:
                flash(f'❌ 检测到重复成果：《{title_value}》已存在于数据库中，无法重复导入！', 'danger')
                return redirect(request.referrer)

    for field in fields_config:
        field_name = field['name']
        field_type = field.get('type', 'text')
        if field_type == 'file':
            continue

        value = request.form.get(field_name, '').strip()

        # 处理外键字段（select 类型且字段名以_id 结尾）- 只保存 ID，不保存对象
        if field_type == 'select' and field_name.endswith('_id'):
            if value == '' or value is None:
                value = None
            else:
                try:
                    value = int(value)  # 直接转为整数保存
                    # 验证外键是否存在（可选）
                    if field_name == 'achievement_type_id':
                        related_obj = db.session.get(TeachingAchievementType, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'achievement_level_id':
                        related_obj = db.session.get(AchievementLevel, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'competition_level_id':
                        related_obj = db.session.get(AchievementLevel, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'award_rank_id':
                        related_obj = db.session.get(AwardRank, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_type_id':
                        related_obj = db.session.get(ProjectType, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_level_id':
                        related_obj = db.session.get(ProjectLevel, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_category_id':
                        related_obj = db.session.get(ProjectCategory, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_status_id':
                        related_obj = db.session.get(ProjectStatus, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'patent_type_id':
                        related_obj = db.session.get(PatentType, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'patent_status_id':
                        related_obj = db.session.get(PatentStatus, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'textbook_level_id':
                        related_obj = db.session.get(TextbookLevel, value)
                        if not related_obj:
                            value = None
                except Exception:
                    value = None

        if field_type == 'number':
            if value == '':
                value = None
            else:
                try:
                    value = float(value)
                except ValueError:
                    flash(f'{field["label"]}必须是数字！', 'danger')
                    return redirect(request.referrer)

        if field_type == 'integer':
            if value == '':
                value = None
            else:
                try:
                    value = int(value)
                except ValueError:
                    flash(f'{field["label"]}必须是数字！', 'danger')
                    return redirect(request.referrer)


        elif field_type == 'select_multiple':
            selected_ids = request.form.getlist(field_name)
            selected_ids = [id_str.strip() for id_str in selected_ids if id_str.strip() and id_str != '']
            value = ','.join(selected_ids) if selected_ids else ''

            if model == JournalPaper and field_name == 'inclusion_type_ids':
                inclusion_names = []
                for inc_id in selected_ids:
                    inc_type = InclusionType.query.get(int(inc_id))
                    if inc_type:
                        inclusion_names.append(inc_type.type_name)
                item.inclusion_status = ','.join(inclusion_names)

        elif field_type == 'month':
            # 处理 month 类型（格式：YYYY-MM）
            if value == '':
                value = None
            else:
                try:
                    # 将 YYYY-MM 转换为日期对象（默认使用该月第一天）
                    value = datetime.strptime(value, '%Y-%m').date()
                except ValueError:
                    flash(f'{field["label"]}格式错误（需为 YYYY-MM）！', 'danger')
                    return redirect(request.referrer)


        elif field_type == 'date':
            if value == '':
                value = None
            else:
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()
                except ValueError:
                    flash(f'{field["label"]}格式错误（需为 YYYY-MM-DD）！', 'danger')
                    return redirect(request.referrer)

        elif value == '':
            value = None

        setattr(item, field_name, value)


    for field in fields_config:
        if field.get('type') == 'file':
            file = request.files.get(field['name'])
            if file and file.filename:
                old_path = getattr(item, field['name'], '')
                if old_path and os.path.exists(old_path):
                    os.remove(old_path)
                sub_folder = model.__tablename__
                new_path = handle_file_upload(file, sub_folder)
                setattr(item, field['name'], new_path)

    item.update_time = datetime.now()

    try:
        if not item_id:
            db.session.add(item)
        db.session.flush()

        if not item_id:
            achievement_type_map = {
                'journal_paper': JournalPaper,
                'conference_paper': ConferencePaper,
                'textbook': Textbook,
                'monograph': Monograph,
                'teaching_project': TeachingProject,
                'patent': Patent,
                'software_copyright': SoftwareCopyright,
                'teaching_achievement_award': TeachingAchievementAward,
                'teaching_competition_award': TeachingCompetitionAward,
                'student_guidance_award': StudentGuidanceAward
            }

            for ach_type, ach_model in achievement_type_map.items():
                if model == ach_model:
                    authors_field = 'authors' if hasattr(item, 'authors') else 'chief_editor'
                    authors_str = getattr(item, authors_field, '')
                    if authors_str:
                        auto_link_contributors(item, ach_type, authors_str, current_user.id)
                    break

        db.session.commit()

        if model == JournalPaper:
            inclusion_type_ids = request.form.getlist('inclusion_type_ids')
            JournalPaperInclusionRelation.query.filter_by(paper_id=item.id).delete()
            for inc_id in inclusion_type_ids:
                if inc_id.strip():
                    relation = JournalPaperInclusionRelation(paper_id=item.id, inclusion_type_id=int(inc_id.strip()))
                    db.session.add(relation)
            db.session.commit()

        flash(f'{"修改" if item_id else "新增"}成功！', 'success')
        return redirect(url_for(request.endpoint, action='list'))
    except Exception as e:
        db.session.rollback()
        flash(f'操作失败：{str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))


def handle_achievement_delete(model, item_id):
    """通用成果删除"""
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    item = db.session.get(model, item_id)
    if not item or item.user_id != current_user.id:
        flash('无权限删除该成果！', 'danger')
        return redirect(url_for('index'))

    # 删除附件文件
    for field in ['attachment']:
        file_path = getattr(item, field, '')
        if file_path and os.path.exists(file_path):
            os.remove(file_path)

    # 删除数据库记录
    try:
        db.session.delete(item)
        db.session.commit()
        flash('删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'danger')

    return redirect(request.referrer or url_for('index'))


def render_achievement_stats(model, title, fields_config, current_user):
    """通用成果统计分析页面"""
    user_ids = get_team_user_ids(current_user) if current_user.role != 'admin' else []
    items = model.query.filter(model.user_id.in_(user_ids)).all()

    # 基础统计：总数
    total = len(items)

    # 按年份统计（取publish_year/award_date等日期字段）
    year_stats = {}
    date_field = None
    for field in fields_config:
        if 'year' in field['name'] or 'date' in field['name']:
            date_field = field['name']
            break

    for item in items:
        year = getattr(item, date_field, '')
        if isinstance(year, int) and year:
            year_stats[year] = year_stats.get(year, 0) + 1
        elif isinstance(year, date) and year:
            year = year.year
            year_stats[year] = year_stats.get(year, 0) + 1

    # 生成统计HTML
    stats_html = f'''
    <div class="stats-container">
        <h3>基础统计</h3>
        <p>成果总数：<strong>{total}</strong></p>

        <h3 style="margin-top:20px;">按年份统计</h3>
        <ul>
    '''
    for year, count in sorted(year_stats.items()):
        stats_html += f'<li>{year}年：{count}项</li>'
    stats_html += '''
        </ul>

        <a href="javascript:history.back()" class="btn " style="margin-top:20px;">返回列表</a>
    </div>
    '''

    return render_base_layout(f'{title} - 统计分析', stats_html, current_user)


def export_achievement_excel(model, fields_config, current_user, start_date=None, end_date=None):
    """通用成果Excel导出（支持时间范围筛选）"""
    user_ids = get_team_user_ids(current_user) if current_user.role != 'admin' else []

    # 基础查询：用户权限过滤
    query = model.query.filter(model.user_id.in_(user_ids))

    # 时间范围筛选（核心新增）
    if start_date or end_date:
        # 确定日期字段（根据不同模型的日期字段适配）
        date_field_map = {
            JournalPaper: 'publish_date',
            ConferencePaper: 'conference_time',
            Textbook: 'publish_date',
            Monograph: 'publish_date',
            TeachingProject: 'start_date',
            Patent: 'apply_date',
            SoftwareCopyright: 'register_date',
            TeachingAchievementAward: 'award_date',
            TeachingCompetitionAward: 'award_date',
            StudentGuidanceAward: 'award_date'
        }

        date_field = date_field_map.get(model, None)
        if date_field:
            # 转换字符串日期为date对象
            start_date_obj = None
            end_date_obj = None

            if start_date:
                try:
                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                except:
                    pass

            if end_date:
                try:
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                except:
                    pass

            # 添加时间筛选条件
            if start_date_obj:
                query = query.filter(getattr(model, date_field) >= start_date_obj)
            if end_date_obj:
                query = query.filter(getattr(model, date_field) <= end_date_obj)

    items = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成果数据'

    # 表头
    headers = [f['label'] for f in fields_config]
    ws.append(headers)

    # 数据行
    for item in items:
        row = []
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            # 处理外键关联字段（显示名称而非 ID）
            if field_name == 'achievement_type_id' and value:
                achievement_type = db.session.get(TeachingAchievementType, value)
                value = achievement_type.type_name if achievement_type else ''
            elif field_name == 'achievement_level_id' and value:
                achievement_level = db.session.get(AchievementLevel, value)
                value = achievement_level.level_name if achievement_level else ''
            elif field_name == 'competition_level_id' and value:
                competition_level = db.session.get(AchievementLevel, value)
                value = competition_level.level_name if competition_level else ''
            elif field_name == 'award_rank_id' and value:
                award_rank = db.session.get(AwardRank, value)
                value = award_rank.rank_name if award_rank else ''
            elif field_name == 'project_type_id' and value:
                project_type = db.session.get(ProjectType, value)
                value = project_type.type_name if project_type else ''
            elif field_name == 'project_level_id' and value:
                project_level = db.session.get(ProjectLevel, value)
                value = project_level.level_name if project_level else ''
            elif field_name == 'project_category_id' and value:
                project_category = db.session.get(ProjectCategory, value)
                value = project_category.category_name if project_category else ''
            elif field_name == 'project_status_id' and value:
                project_status = db.session.get(ProjectStatus, value)
                value = project_status.status_name if project_status else ''
            elif field_name == 'patent_type_id' and value:
                patent_type = db.session.get(PatentType, value)
                value = patent_type.type_name if patent_type else ''
            elif field_name == 'patent_status_id' and value:
                patent_status = db.session.get(PatentStatus, value)
                value = patent_status.status_name if patent_status else ''
            elif field_name == 'textbook_level_id' and value:
                textbook_level = db.session.get(TextbookLevel, value)
                value = textbook_level.level_name if textbook_level else ''

            if value is None:
                value = ''
            elif isinstance(value, (date, datetime)):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field_name == 'attachment' and value:
                value = os.path.basename(value) if value else ''
            row.append(value)
        ws.append(row)

    # 列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{model.__tablename__}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
def get_zhipu_api_key(current_user):
    """从用户配置中获取智谱AI API Key"""
    api_config = current_user.get_api_config()
    return api_config.get('zhipu', {}).get('api_key', '')

def ai_analyze_journal_full(citation_text, api_key):
    """AI分析期刊论文引用文本"""
    if not citation_text.strip() or not api_key:
        return {"起止页码": "", "年": "", "卷": "", "期": "", "DOI": ""}

    url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    prompt = f"""
【任务】从指定的知网期刊论文引用文本（格式：作者.题名[J].刊名,年,卷(期):起止页码.DOI.）中精准提取以下信息：
1. 起止页码
2. 年
3. 卷
4. 期
5. DOI

【输入文本】
{citation_text}

【输出规则】
1. 仅输出标准JSON字符串，无任何多余文字、注释、反引号、说明。
2. JSON必须包含字段："起止页码"、"年"、"卷"、"期"、"DOI"。
3. 起止页码格式：数字-数字，无则为空。
4. 年、卷、期只保留数字，无则为空。
5. DOI只保留编号，去掉DOI:前缀，无则为空。
6. 严格按JSON输出，不要任何多余内容。
    """

    payload = {
        "model": "glm-4-flash",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0,
        "stream": False
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=15)
        response.raise_for_status()
        result = response.json()
        ai_content = result['choices'][0]['message']['content'].strip()
        ai_content = ai_content.replace('```json', '').replace('```', '').strip()
        data = json.loads(ai_content)

        return {
            "起止页码": data.get("起止页码", "").strip(),
            "年": data.get("年", "").strip(),
            "卷": data.get("卷", "").strip(),
            "期": data.get("期", "").strip(),
            "DOI": data.get("DOI", "").strip()
        }
    except Exception as e:
        print(f"AI分析期刊论文失败：{e}")
        return {"起止页码": "", "年": "", "卷": "", "期": "", "DOI": ""}

def ai_analyze_citation(citation_text, api_key):
    """AI分析会议论文引用文本"""
    if not citation_text.strip() or not api_key:
        return {'会议地点': '', '起止页码': ''}

    url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    prompt = f"""
【任务】从指定的知网会议论文引用文本（格式：作者.题名[C]//会议主办单位.会议论文集名.会议地点;,出版年:起止页码.DOI.）中精准提取「会议地点」和「起止页码」两类核心信息。
【输入文本】
{citation_text}
【输出规则】
1. 仅输出标准JSON字符串，无任何多余文字；
2. JSON必须包含两个字段："会议地点"和"起止页码"；
3. 会议地点：提取引用文本中的地点/机构信息，无则为空；
4. 起止页码：格式为"数字-数字"，无则为空；
5. 严格按格式输出，示例：{{"会议地点":"湖南中医药大学","起止页码":"10-13"}}。
    """

    payload = {
        "model": "glm-4-flash",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0,
        "stream": False
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        result = response.json()
        ai_content = result['choices'][0]['message']['content'].strip()
        ai_content = ai_content.replace('```json', '').replace('```', '').strip()
        parsed_ai = json.loads(ai_content)

        return {
            '会议地点': parsed_ai.get('会议地点', '').strip(),
            '起止页码': parsed_ai.get('起止页码', '').strip()
        }
    except Exception as e:
        print(f"AI分析会议论文失败：{e}")
        return {'会议地点': '', '起止页码': ''}

def crawl_cnki_journal(keyword, max_papers=3, driver_path=r'C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe'):
    """爬取知网期刊论文"""
    # 浏览器配置
    options = webdriver.EdgeOptions()
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument('--disable-blink-features=AutomationControlled')

    try:
        service = Service(driver_path)
        browser = webdriver.Edge(service=service, options=options)
        browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh']});
            """
        })
        browser.implicitly_wait(3)
        actions = ActionChains(browser)
        results = []

        # 提取论文ID
        def extract_paper_id(link):
            try:
                id_match = re.search(r'id=([^&]+)', link)
                return id_match.group(1) if id_match else ""
            except:
                return ""

        # 提取引用格式
        def extract_quote_manual(row):
            citation = ""
            try:
                quote_btn = WebDriverWait(row, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.icon-quote'))
                )
                browser.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_btn)
                quote_btn.click()

                quote_elem = WebDriverWait(browser, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "td.quote-r"))
                )
                citation = quote_elem.text.strip()

                close_btn = WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.layui-layer-close.layui-layer-close1'))
                )
                close_btn.click()
            except:
                pass
            return citation

        # 爬取逻辑
        browser.get("https://kns.cnki.net/kns8s/AdvSearch")
        WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[name="classify"][resource="JOURNAL"]'))
        )
        browser.find_element(By.CSS_SELECTOR, 'a[name="classify"][resource="JOURNAL"]').click()

        search_input = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-tipid="gradetxt-2"]'))
        )
        search_input.clear()
        search_input.send_keys(keyword)

        search_btn = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.btn-search'))
        )
        search_btn.click()

        table = WebDriverWait(browser, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.result-table-list'))
        )

        rows = table.find_elements(By.CSS_SELECTOR, 'tbody tr')[:max_papers]

        for row in rows:
            paper_data = {
                '论文ID': "", '论文名称': "", '论文作者': "", '通讯作者': "",
                '期刊名称': "", '论文收录情况': "", '年': "", '卷': "", '期': "",
                '起止页码': "", '发表年份': "", '发表日期': "", '引用格式': "", 'DOI': ""
            }
            try:
                title_elem = WebDriverWait(row, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'a.fz14'))
                )
                paper_data['论文名称'] = title_elem.text.strip()
                paper_data['论文ID'] = extract_paper_id(title_elem.get_attribute('href'))
                paper_data['论文作者'] = row.find_element(By.CSS_SELECTOR, 'td.author').text.strip().replace('；', ',')
                paper_data['期刊名称'] = row.find_element(By.CSS_SELECTOR, 'td.source').text.strip()

                # 发表日期
                try:
                    date_elem = row.find_element(By.CSS_SELECTOR, 'td.date')
                    paper_data['发表日期'] = date_elem.text.strip()
                    if paper_data['发表日期']:
                        paper_data['发表年份'] = paper_data['发表日期'].split('-')[0]
                except:
                    pass

                # 提取引用格式
                paper_data['引用格式'] = extract_quote_manual(row)
                results.append(paper_data)
                time.sleep(random.uniform(0.5, 1))

            except Exception as e:
                print(f"爬取单篇期刊论文失败：{e}")
                continue

        browser.quit()
        return results
    except Exception as e:
        print(f"爬取知网期刊论文失败：{e}")
        return []

def crawl_cnki_conference(keyword, max_papers=3, driver_path=r'C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe'):
    """爬取知网会议论文"""
    # 浏览器配置
    options = webdriver.EdgeOptions()
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--disable-images')

    try:
        service = Service(driver_path)
        browser = webdriver.Edge(service=service, options=options)
        browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh']});
            """
        })
        browser.implicitly_wait(1)
        results = []

        # 提取引用格式
        def extract_quote_manual(row):
            citation = ""
            try:
                quote_btn = WebDriverWait(row, 8).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.icon-quote[title="引用"]'))
                )
                browser.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_btn)
                browser.execute_script("arguments[0].click();", quote_btn)

                quote_elem = WebDriverWait(browser, 8).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "td.quote-r"))
                )
                citation = quote_elem.text.strip().replace('\n', '').replace('  ', ' ')

                close_btn = WebDriverWait(browser, 8).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.layui-layer-close.layui-layer-close1'))
                )
                close_btn.click()
            except:
                pass
            return citation

        # 爬取逻辑
        browser.get("https://kns.cnki.net/kns8s/AdvSearch")
        WebDriverWait(browser, 8).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[name="classify"]'))
        )

        classify_elem = WebDriverWait(browser, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[name="classify"][resource="CONFERENCE"]'))
        )
        classify_elem.click()

        search_input = WebDriverWait(browser, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-tipid="gradetxt-2"]'))
        )
        search_input.clear()
        search_input.send_keys(keyword)

        search_btn = WebDriverWait(browser, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.btn-search'))
        )
        search_btn.click()

        table = WebDriverWait(browser, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.result-table-list'))
        )

        rows = table.find_elements(By.CSS_SELECTOR, 'tbody tr')[:max_papers]

        for idx, row in enumerate(rows):
            paper_data = {
                '论文名称': "", '论文作者': "", '通讯作者': "", '会议名称': "",
                '会议时间': "", '会议地点': "", '起止页码': "", 'DOI': "",
                '发表年份': "", '引用格式': ""
            }
            try:
                # 论文名称
                title_elem = WebDriverWait(row, 8).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'a.fz14'))
                )
                paper_data['论文名称'] = title_elem.text.strip()

                # 论文作者
                try:
                    author_elem = WebDriverWait(row, 3).until(
                        EC.presence_of_element_located((
                            By.CSS_SELECTOR,
                            'td[aria-describedby*="author"], td.authorname, td:nth-child(3)'
                        ))
                    )
                    paper_data['论文作者'] = author_elem.text.strip().replace('；', ',').replace(' ', '')
                except:
                    paper_data['论文作者'] = ""

                # 通讯作者
                try:
                    author_text = paper_data['论文作者']
                    if '通讯作者：' in author_text:
                        paper_data['通讯作者'] = author_text.split('通讯作者：')[1].split(',')[0].strip()
                        paper_data['论文作者'] = author_text.split('通讯作者：')[0].strip().rstrip(',')
                    else:
                        paper_data['通讯作者'] = ""
                except:
                    paper_data['通讯作者'] = ""

                # 会议名称
                try:
                    source_elem = WebDriverWait(row, 3).until(
                        EC.presence_of_element_located((
                            By.CSS_SELECTOR,
                            'td[aria-describedby*="source"], td.conferencename, td:nth-child(4)'
                        ))
                    )
                    paper_data['会议名称'] = source_elem.text.strip()
                except:
                    paper_data['会议名称'] = ""

                # 会议时间
                try:
                    time_elem = WebDriverWait(row, 3).until(
                        EC.presence_of_element_located((
                            By.CSS_SELECTOR,
                            'td[aria-describedby*="meetetime"], td.meetingtime, td.date'
                        ))
                    )
                    paper_data['会议时间'] = time_elem.text.strip()
                except:
                    paper_data['会议时间'] = ""

                # 发表年份
                try:
                    meeting_time = paper_data['会议时间']
                    if meeting_time and len(meeting_time) >= 4:
                        paper_data['发表年份'] = meeting_time[:4]
                    else:
                        paper_data['发表年份'] = ""
                except:
                    paper_data['发表年份'] = ""

                # 引用格式
                paper_data['引用格式'] = extract_quote_manual(row)

                # DOI解析
                try:
                    doi_pattern = r'DOI[:：]?\s*(\d+\.\d+/[\w\-\.]+)'
                    doi_match = re.search(doi_pattern, paper_data['引用格式'], re.IGNORECASE)
                    if doi_match:
                        paper_data['DOI'] = doi_match.group(1).strip()
                except:
                    paper_data['DOI'] = ""

                results.append(paper_data)
                time.sleep(random.uniform(0.5, 1))

            except Exception as e:
                print(f"爬取单篇会议论文失败：{e}")
                continue

        browser.quit()
        return results
    except Exception as e:
        print(f"爬取知网会议论文失败：{e}")
        return []


# ---------------------- OCR/语音核心函数 ----------------------
def get_baidu_token(current_user):
    """从用户配置获取百度API Token"""
    api_config = current_user.get_api_config()
    baidu_api_key = api_config.get('baidu', {}).get('api_key', '')
    baidu_secret_key = api_config.get('baidu', {}).get('secret_key', '')

    if not baidu_api_key or not baidu_secret_key:
        return None, "未配置百度API Key/Secret Key！请先前往个人设置 > 大模型API配置 中配置百度 OCR API。"

    url = f"https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={baidu_api_key}&client_secret={baidu_secret_key}"
    try:
        response = requests.post(url, verify=False, timeout=10)
        response.raise_for_status()  # 检查 HTTP 状态码
        token_data = response.json()

        # 检查是否成功获取 token
        if "error" in token_data:
            return None, f"百度API 认证失败：{token_data.get('error_description', '未知错误')}"

        return token_data.get("access_token"), None
    except requests.exceptions.Timeout:
        return None
    except requests.exceptions.ConnectionError:
        return None, "无法连接到百度API 服务器，请检查网络"
    except Exception as e:
        logger.error(f"获取百度 Token 异常：{str(e)}")
        return None, f"获取百度 Token 失败：{str(e)}"


def baidu_ocr_recognize(image_path, current_user):
    """百度 OCR识别图片文字"""
    token, err = get_baidu_token(current_user)
    if err:
        return "", err

    try:
        with open(image_path, 'rb') as f:
            image_data = f.read()
        image_base64 = base64.b64encode(image_data).decode('utf-8')
    except FileNotFoundError:
        return "", f"图片文件不存在：{image_path}"
    except Exception as e:
        return "", f"读取图片失败：{str(e)}"

    ocr_url = "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    params = {
        "access_token": token,
        "image": image_base64,
        "language_type": "CHN_ENG"
    }

    try:
        response = requests.post(ocr_url, headers=headers, data=params, timeout=30)
        response.raise_for_status()  # 检查 HTTP 状态码
        result = response.json()

        # 检查百度API 返回的错误
        if "error_code" in result:
            error_code = result.get('error_code', 'unknown')
            error_msg = result.get('error_msg', '未知错误')

            # 常见错误码解释
            error_explanations = {
                '17': 'QPS 超限，请稍后再试',
                '18': '总请求次数超限',
                '19': '图片大小超限',
                '100': 'Token 无效或过期',
                '110': 'Token 已过期',
            }

            explanation = error_explanations.get(str(error_code), '')
            full_error = f"百度 OCR 调用失败 [错误码{error_code}]：{error_msg}"
            if explanation:
                full_error += f"（{explanation}）"

            logger.error(full_error)
            return "", full_error

        # 正常返回 OCR 结果
        if "words_result" not in result:
            logger.error(f"百度 OCR 返回数据格式异常：{result}")
            return "", "百度 OCR 返回数据格式异常"

        ocr_text = "\n".join([item["words"] for item in result.get("words_result", [])])
        return ocr_text.strip(), None

    except requests.exceptions.Timeout:
        return "", "OCR识别超时，请检查网络连接或重试"
    except requests.exceptions.ConnectionError:
        return "", "无法连接到 OCR 服务器，请检查网络"
    except Exception as e:
        logger.error(f"OCR识别异常：{str(e)}")
        return "", f"OCR识别失败：{str(e)}"


def extract_achievement_info(ocr_text):
    """解析 OCR 文本，提取成果信息（仅识别教研教改和课程建设项目）"""
    if not ocr_text.strip():
        return {
            'type_name': '识别失败',
            'title': '',
            'extra_fields': {},
            'raw_text': ocr_text,
            'confidence': 0.0
        }

    # 规则匹配成果类型（仅教研教改和课程建设项目）
    clean_text = unicodedata.normalize('NFKC', ocr_text)
    clean_text = re.sub(r'\s+', ' ', clean_text)
    lines = [line.strip() for line in ocr_text.split('\n') if line.strip()]
    title = lines[0] if lines else ''

    if '教学改革' in clean_text or '教改' in clean_text:
        return {
            'type_name': '教研教改和课程建设项目',
            'title': title,
            'extra_fields': {},
            'raw_text': ocr_text,
            'confidence': 0.95
        }

    # 匹配关键词
    matched_type = None
    confidence = 0.0

    rule = achievement_rules['教研教改和课程建设项目']

    # 关键词匹配（只要包含任意一个关键词即匹配）
    keyword_matched = [kw for kw in rule['keywords'] if kw in clean_text]
    if keyword_matched:
        matched_type = '教研教改和课程建设项目'
        confidence = min(0.7 + len(keyword_matched) * 0.05, 0.95)

    # 正则匹配（兜底）
    if not matched_type and re.search(rule['pattern'], clean_text, re.IGNORECASE | re.MULTILINE):
        matched_type = '教研教改和课程建设项目'
        confidence = 0.9

    # 如果未匹配到，返回"识别失败"，由 AI 进一步分析
    if not matched_type:
        return {
            'type_name': '识别失败',
            'title': title,
            'extra_fields': {},
            'raw_text': ocr_text,
            'confidence': 0.0
        }

    return {
        'type_name': matched_type,
        'title': title,
        'extra_fields': {},
        'raw_text': ocr_text,
        'confidence': round(confidence, 2)
    }

def audio_to_text(audio_data, current_user):
    """音频转文字（百度语音识别）"""
    # 获取Token
    token, err = get_baidu_token(current_user)
    if err:
        return "", f"获取语音识别Token失败：{err}"

    # WebM转WAV
    f_in_name = None
    f_out_name = None
    try:
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as f_in:
            f_in.write(audio_data)
            f_in_name = f_in.name

        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as f_out:
            f_out_name = f_out.name

        # FFmpeg路径
        ffmpeg_exe = SystemConfig.get_config('ffmpeg_exe', "D:\\ffmpeg\\bin\\ffmpeg.exe")

        # 转换格式
        cmd = [
            ffmpeg_exe,
            "-i", f_in_name,
            "-ar", "16000",
            "-ac", "1",
            "-sample_fmt", "s16",
            "-y",
            f_out_name
        ]
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode != 0:
            raise RuntimeError(f"FFmpeg转换失败：{result.stderr}")

        # 读取WAV数据
        with open(f_out_name, "rb") as f:
            wav_data = f.read()

    except Exception as e:
        return "", f"音频格式转换失败：{str(e)}"
    finally:
        # 清理临时文件
        if f_in_name and os.path.exists(f_in_name):
            os.unlink(f_in_name)
        if f_out_name and os.path.exists(f_out_name):
            os.unlink(f_out_name)

    # 调用百度语音识别
    audio_base64 = base64.b64encode(wav_data).decode('utf-8')
    params = {
        "format": "wav",
        "rate": 16000,
        "channel": 1,
        "cuid": f"achievement_{current_user.id}",
        "token": token,
        "speech": audio_base64,
        "len": len(wav_data),
        "dev_pid": 1537
    }

    try:
        response = requests.post("https://vop.baidu.com/server_api",
                                 json=params,
                                 headers={"Content-Type": "application/json"},
                                 timeout=10)
        result = response.json()
        if result.get("err_no") != 0:
            return "", f"语音识别失败：{result.get('err_msg', '未知错误')}"
        text = result.get("result", [""])[0]
        return text.strip(), None
    except Exception as e:
        return "", f"语音识别请求失败：{str(e)}"


def parse_voice_command(text):
    """解析语音指令，适配团队导出（支持识别老师姓名+成果类型）"""
    result = {
        "action": None,
        "start_date": None,
        "end_date": None,
        "type_name": None,
        "teacher_name": None,  # 新增：识别老师姓名
        "achievement_name": None,  # 新增：识别成果名称
        "is_my": True,
        "is_team": False,
        "is_teaching": False,
        "type_names": []
    }

    if not text:
        return result

    # 识别操作类型
    export_keywords = ["导出", "下载", "保存", "导出Excel", "下载Excel"]
    if any(keyword in text for keyword in export_keywords):
        result["action"] = "export"
    elif any(keyword in text for keyword in ["筛选", "查看", "查询"]):
        result["action"] = "filter"
    else:
        result["action"] = "export"

    # 识别教学相关
    if any(keyword in text for keyword in ["教学", "教学成果", "教学类"]):
        result["is_teaching"] = True

    # ========== 核心增强：识别老师姓名 ==========
    # 匹配"XX老师"格式
    teacher_pattern = r'导出([^，。！？\s]+)老师'
    teacher_match = re.search(teacher_pattern, text)
    if teacher_match:
        result["teacher_name"] = teacher_match.group(1).strip()
        result["is_my"] = False  # 指定了老师，不再是导出自己的

    # ========== 核心增强：识别成果名称（可选） ==========
    # 匹配"的XX项目/的XX论文/的XX专利"等格式
    achievement_pattern = r'的([^，。！？\s]+)(项目|论文|专利|软著|教材|专著|获奖)'
    achievement_match = re.search(achievement_pattern, text)
    if achievement_match:
        result["achievement_name"] = achievement_match.group(1).strip() + achievement_match.group(2).strip()

    # ========== 核心修复：增加对「专利」通用关键词的识别 ==========
    type_names = list(achievement_rules.keys())[:-1]  # 排除"其他"
    # 先匹配完整名称（如发明专利）
    for t_name in type_names:
        if t_name in text:
            result["type_name"] = t_name
            break
    # 如果没匹配到，再匹配通用名称
    if not result["type_name"]:
        if "专利" in text:
            result["type_name"] = "专利"  # 匹配通用的"专利"关键词
        elif "软著" in text:
            result["type_name"] = "软著"
        elif "论文" in text:
            if "期刊" in text:
                result["type_name"] = "期刊论文"
            elif "会议" in text:
                result["type_name"] = "会议论文"
            else:
                result["type_name"] = "期刊论文"  # 默认匹配期刊论文
        elif "教材" in text:
            result["type_name"] = "教材"
        elif "专著" in text:
            result["type_name"] = "专著"
        elif "获奖" in text:
            if "教学竞赛" in text:
                result["type_name"] = "教学竞赛获奖"
            elif "指导学生" in text:
                result["type_name"] = "指导学生获奖"
            else:
                result["type_name"] = "教学成果获奖"
        elif "教研" in text or "教改" in text:
            result["type_name"] = "教研教改和课程建设项目"
        elif "项目" in text and ("教研" in text or "教改" in text):
            result["type_name"] = "教研教改和课程建设项目"

    # 识别时间范围
    single_year_pattern = r'(\d{4})年'
    single_year_match = re.search(single_year_pattern, text)
    if single_year_match:
        result["start_date"] = f"{single_year_match.group(1)}-01-01"
        result["end_date"] = f"{single_year_match.group(1)}-12-31"

    # 年份范围
    year_range_pattern = r'(\d{4})年[至|-|到](\d{4})年'
    year_match = re.search(year_range_pattern, text)
    if year_match:
        result["start_date"] = f"{year_match.group(1)}-01-01"
        result["end_date"] = f"{year_match.group(2)}-12-31"

    # 近三年
    if "近三年" in text:
        current_year = datetime.now().year
        result["start_date"] = f"{current_year - 3}-01-01"
        result["end_date"] = f"{current_year}-12-31"

    # 团队/个人
    if any(keyword in text for keyword in ["团队", "集体", "所有成员"]):
        result["is_my"] = False
        result["is_team"] = True

    return result


def create_achievement_from_ocr(ocr_result, current_user):
    """
    根据 OCR+AI 分析结果创建成果记录（支持用户手动修改后的数据）
    ocr_result 结构：
    {
        'type_name': '期刊论文',
        'title': '论文标题',
        'raw_text': '原始识别文本',
        'extra_fields': {'authors': '作者', 'journal_name': '期刊名', ...},  # 用户确认/修改的字段
        'ai_data': {...}  # AI 分析的原始数据（可选）
    }
    """
    type_name = ocr_result.get('type_name', '')
    title = ocr_result.get('title', '')
    raw_text = ocr_result.get('raw_text', '')
    extra_fields = ocr_result.get('extra_fields', {})
    ai_info = ocr_result.get('ai_data', {})

    if not type_name or not title:
        return False, "成果类型和名称不能为空", None, None

    type_model_mapping = {
        '期刊论文': JournalPaper,
        '会议论文': ConferencePaper,
        '教材': Textbook,
        '专著': Monograph,
        '发明专利': Patent,
        '实用新型专利': Patent,
        '软著': SoftwareCopyright,
        '教学成果获奖': TeachingAchievementAward,
        '教学竞赛获奖': TeachingCompetitionAward,
        '指导学生获奖': StudentGuidanceAward,
        '教研教改和课程建设项目': TeachingProject
    }

    if type_name not in type_model_mapping:
        return False, f"暂不支持创建{type_name}类型的成果", None, None

    model = type_model_mapping[type_name]
    try:
        achievement = model()
        achievement.user_id = current_user.id
        achievement.title = title
        achievement.create_time = datetime.now()
        achievement.update_time = datetime.now()

        # 核心修复：优先使用用户确认的 extra_fields，其次使用 AI 分析结果
        field_data = {**ai_info, **extra_fields}

        if type_name == '期刊论文':
            achievement.authors = field_data.get('authors', '')
            achievement.corresponding_authors = field_data.get('corresponding_authors', '')
            achievement.journal_name = field_data.get('journal_name', '')
            achievement.inclusion_status = field_data.get('inclusion_status', '')

            if field_data.get('year'):
                try:
                    achievement.year = int(field_data['year'])
                except:
                    pass
            if field_data.get('publish_year'):
                try:
                    achievement.publish_year = int(field_data['publish_year'])
                except:
                    pass

            achievement.volume = field_data.get('volume', '')
            achievement.issue = field_data.get('issue', '')
            achievement.page_range = field_data.get('page_range', '')
            achievement.doi = field_data.get('doi', '')

            if field_data.get('publish_date'):
                try:
                    achievement.publish_date = datetime.strptime(field_data['publish_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '会议论文':
            achievement.authors = field_data.get('authors', '')
            achievement.corresponding_authors = field_data.get('corresponding_authors', '')
            achievement.conference_name = field_data.get('conference_name', '')

            if field_data.get('publish_year'):
                try:
                    achievement.publish_year = int(field_data['publish_year'])
                except:
                    pass

            achievement.page_range = field_data.get('page_range', '')
            achievement.doi = field_data.get('doi', '')

            if field_data.get('conference_time'):
                achievement.conference_time = field_data.get('conference_time', '')
            achievement.conference_place = field_data.get('conference_place', '')

        elif type_name == '教材':
            achievement.textbook_series = field_data.get('textbook_series', '')
            achievement.chief_editor = field_data.get('chief_editor', '')
            achievement.associate_editors = field_data.get('associate_editors', '')
            achievement.editorial_board = field_data.get('editorial_board', '')
            achievement.publisher = field_data.get('publisher', '')
            achievement.isbn = field_data.get('isbn', '')
            achievement.cip_number = field_data.get('cip_number', '')

            if field_data.get('publication_year'):
                try:
                    achievement.publication_year = int(field_data['publication_year'])
                except:
                    pass
            if field_data.get('publication_month'):
                try:
                    achievement.publication_month = int(field_data['publication_month'])
                except:
                    pass

            achievement.edition = field_data.get('edition', '')
            achievement.word_count = field_data.get('word_count', '')
            achievement.price = field_data.get('price', '')
            achievement.textbook_level = field_data.get('textbook_level', '')
            achievement.textbook_type = field_data.get('textbook_type', '')
            achievement.applicable_majors = field_data.get('applicable_majors', '')
            achievement.remarks = field_data.get('remarks', '')

        elif type_name == '专著':
            achievement.textbook_series = field_data.get('textbook_series', '')
            achievement.chief_editor = field_data.get('chief_editor', '')
            achievement.associate_editors = field_data.get('associate_editors', '')
            achievement.editorial_board = field_data.get('editorial_board', '')
            achievement.publisher = field_data.get('publisher', '')
            achievement.isbn = field_data.get('isbn', '')
            achievement.cip_number = field_data.get('cip_number', '')

            if field_data.get('publication_year'):
                try:
                    achievement.publication_year = int(field_data['publication_year'])
                except:
                    pass
            if field_data.get('publication_month'):
                try:
                    achievement.publication_month = int(field_data['publication_month'])
                except:
                    pass

            achievement.edition = field_data.get('edition', '')
            achievement.word_count = field_data.get('word_count', '')
            achievement.price = field_data.get('price', '')
            achievement.monograph_type = field_data.get('monograph_type', '')
            achievement.applicable_majors = field_data.get('applicable_majors', '')
            achievement.remarks = field_data.get('remarks', '')

        elif type_name in ['发明专利', '实用新型专利', '外观设计专利']:
            patent_type = PatentType.query.filter_by(type_name=type_name).first()
            if patent_type:
                achievement.patent_type_id = patent_type.id

            achievement.patentee = field_data.get('patentee', '')
            achievement.address = field_data.get('address', '')
            achievement.inventors = field_data.get('inventors', '')
            achievement.grant_announcement_number = field_data.get('grant_announcement_number', '')
            achievement.applicant_at_apply_date = field_data.get('applicant_at_apply_date', '')
            achievement.inventor_at_apply_date = field_data.get('inventor_at_apply_date', '')
            achievement.patent_number = field_data.get('patent_number', '')

            status_name = field_data.get('status', '')
            if status_name:
                patent_status = PatentStatus.query.filter_by(status_name=status_name).first()
                if patent_status:
                    achievement.patent_status_id = patent_status.id

            if field_data.get('apply_date'):
                try:
                    achievement.apply_date = datetime.strptime(field_data['apply_date'], '%Y-%m-%d').date()
                except:
                    pass
            if field_data.get('grant_announcement_date'):
                try:
                    achievement.grant_announcement_date = datetime.strptime(field_data['grant_announcement_date'],
                                                                            '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '软著':
            achievement.copyright_owner = field_data.get('copyright_owner', '')
            achievement.right_acquisition_method = field_data.get('right_acquisition_method', '')
            achievement.right_scope = field_data.get('right_scope', '')
            achievement.certificate_number = field_data.get('certificate_number', '')
            achievement.copyright_number = field_data.get('copyright_number', '')

            if field_data.get('completion_date'):
                try:
                    achievement.completion_date = datetime.strptime(field_data['completion_date'], '%Y-%m-%d').date()
                except:
                    pass
            if field_data.get('first_publication_date'):
                try:
                    achievement.first_publication_date = datetime.strptime(field_data['first_publication_date'],
                                                                           '%Y-%m-%d').date()
                except:
                    pass
            if field_data.get('register_date'):
                try:
                    achievement.register_date = datetime.strptime(field_data['register_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '教学成果获奖':
            achievement.main_contributors = field_data.get('main_contributors', '')
            achievement.completing_units = field_data.get('completing_units', '')
            achievement.award_year = field_data.get('award_year', '')
            achievement.certificate_number = field_data.get('certificate_number', '')
            achievement.awarding_unit = field_data.get('awarding_unit', '')

            if field_data.get('award_date'):
                try:
                    achievement.award_date = datetime.strptime(field_data['award_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '教学竞赛获奖':
            achievement.winners = field_data.get('winners', '')
            achievement.winner_unit = field_data.get('winner_unit', '')
            achievement.competition_name = field_data.get('competition_name', '')
            achievement.award_year = field_data.get('award_year', '')
            achievement.certificate_number = field_data.get('certificate_number', '')

            if field_data.get('award_date'):
                try:
                    achievement.award_date = datetime.strptime(field_data['award_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '指导学生获奖':
            achievement.student_name = field_data.get('student_name', '')
            achievement.project_name = field_data.get('project_name', '')
            achievement.teacher_name = field_data.get('teacher_name', '')
            achievement.competition_name = field_data.get('competition_name', '')
            achievement.award_year = field_data.get('award_year', '')
            achievement.certificate_number = field_data.get('certificate_number', '')

            if field_data.get('award_date'):
                try:
                    achievement.award_date = datetime.strptime(field_data['award_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '教研教改和课程建设项目':
            achievement.title = field_data.get('title', title)
            achievement.project_leader = field_data.get('project_leader', '')
            achievement.project_members = field_data.get('project_members', '')
            achievement.approval_department = field_data.get('approval_department', '')

            if field_data.get('approval_date'):
                try:
                    achievement.approval_date = datetime.strptime(field_data['approval_date'], '%Y-%m-%d').date()
                except:
                    pass
            if field_data.get('start_date'):
                try:
                    achievement.start_date = datetime.strptime(field_data['start_date'], '%Y-%m-%d').date()
                except:
                    pass
            if field_data.get('end_date'):
                try:
                    achievement.end_date = datetime.strptime(field_data['end_date'], '%Y-%m-%d').date()
                except:
                    pass

        # 保存到数据库
        db.session.add(achievement)
        db.session.commit()

        return True, f"成功创建{type_name}记录", type_name, achievement.id

    except Exception as e:
        import traceback
        logger.error(f"创建成果记录失败：{str(e)}")
        logger.error(traceback.format_exc())
        db.session.rollback()
        return False, f"创建失败：{str(e)}", None, None


def ai_analyze_achievement_text(ocr_text, api_key, current_user=None):
    """
    增强版：调用智谱 AI 分析 OCR 文本，提取所有成果类型的全量数据库字段
    :param ocr_text: OCR 识别的原始文本
    :param api_key: 智谱 AI API Key
    :param current_user: 当前登录用户（用于筛选教研教改和课程建设项目）
    :return: 包含全量字段的结构化字典
    """
    if not ocr_text.strip() or not api_key:
        # 返回全量空字段（匹配数据库模型）
        return {
            # 通用字段
            'type_name': '识别失败',
            'title': '',
            'confidence': 0.0,
            'raw_data': {},

            # 期刊论文专属
            'authors': '',
            'corresponding_authors': '',
            'journal_name': '',
            'inclusion_status': '',
            'year': '',
            'volume': '',
            'issue': '',
            'page_range': '',
            'doi': '',
            'publish_year': '',
            'publish_date': '',

            # 会议论文专属
            'conference_name': '',
            'conference_time': '',
            'conference_place': '',

            # 教材专属
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'textbook_level': '',
            'textbook_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专著专属
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'publish_date': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'monograph_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专利专属
            'patent_type': '',
            'patent_number': '',
            'apply_date': '',
            'grant_date': '',
            'status': '',

            # 软著专属
            'copyright_owner': '',
            'completion_date': '',
            'first_publication_date': '',
            'right_acquisition_method': '',
            'right_scope': '',
            'copyright_number': '',
            'certificate_number': '',
            'register_date': '',

            # 教研教改和课程建设项目专属
            'project_code': '',
            'project_leader': '',
            'project_members': '',
            'approval_department': '',
            'approval_date': '',
            'project_type': '',
            'project_level': '',
            'project_category': '',
            'funding': '',
            'start_date': '',
            'end_date': '',

            # 获奖类专属
            'award_level': '',
            'award_rank': '',
            'award_date': '',
            'competition_name': '',
            'student_name': '',
        }

    url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    user_name = ''
    if current_user:
        user_name = current_user.username

    # 增强版 Prompt：明确要求提取所有数据库字段
    prompt = f"""
【任务】从以下文本中精准提取教学成果的**所有关键信息**，严格按指定格式输出 JSON 字符串。
【文本内容】
{ocr_text}

【核心要求】
1. 先识别成果类型（必须是以下之一）：
   期刊论文/会议论文/发明专利/实用新型专利/教材/专著/软著/教学成果获奖/教学竞赛获奖/指导学生获奖/教研教改和课程建设项目

2. **重要：如果是教研教改和课程建设项目，且文本中包含多个项目，只提取项目负责人或参与人包含"{user_name}"的项目**
   - 如果表格中有"主持人"或"项目负责人"列，只提取该列包含"{user_name}"的项目
   - 如果表格中有"参加人员"列，也检查是否包含"{user_name}"
   - 如果没有找到"{user_name}"的项目，返回空字段


3. 根据成果类型，提取对应**所有**字段（无信息则为空字符串）：



### 期刊论文字段
- title: 论文名称（必填）
- authors: 作者（多个用逗号分隔）
- corresponding_authors: 通讯作者（多个用逗号分隔）
- journal_name: 期刊名称
- inclusion_status: 收录情况（如 SCI/SSCI/EI/CSSCI/北大核心等）
- year: 发表年（仅数字）
- volume: 卷（仅数字/字符）
- issue: 期（仅数字/字符）
- page_range: 起止页码（如 10-20）
- doi: DOI 编号
- publish_year: 发表年份（仅数字）
- publish_date: 发表日期（格式 YYYY-MM-DD，无则为空）

### 会议论文字段
- title: 论文名称
- authors: 作者
- corresponding_authors: 通讯作者
- conference_name: 会议名称
- conference_time: 会议时间（YYYY-MM-DD）
- conference_place: 会议地点
- page_range: 起止页码
- doi: DOI 编号
- publish_year: 发表年份

### 教材字段
- title: 教材名称
- textbook_series: 教材系列
- chief_editor: 主编
- associate_editors: 副主编
- editorial_board: 编委
- publisher: 出版社
- isbn: ISBN 号
- cip_number: CIP 核字号
- publication_year: 出版年份
- publication_month: 出版月份（仅数字）
- edition: 版次（如第 1 版）
- word_count: 字数（如 318 千字）
- price: 定价（如 49.00）
- textbook_level: 教材级别（国家级规划/全国行业规划/协编/自编/其它）
- textbook_type: 教材类型（纸质/数字）
- applicable_majors: 适用专业
- remarks: 备注

### 专著字段（重点增强）
- title: 专著名称
- textbook_series: 专著系列
- chief_editor: 主编
- associate_editors: 副主编
- editorial_board: 编委
- publisher: 出版社
- isbn: ISBN 号（13 位或 10 位）
- cip_number: CIP 核字号（如"2023 第 XXXXX 号"）
- publication_year: 出版年份（仅数字）
- publication_month: 出版月份（仅数字）
- publish_date: 出版日期（格式 YYYY-MM-DD）
- edition: 版次（如第 1 版、修订版）
- word_count: 字数（如 318 千字）
- price: 定价（如 49.00）
- monograph_type: 专著类型（学术专著/技术专著/科普著作/其它）
- applicable_majors: 适用专业
- remarks: 备注

### 专利字段（发明/实用新型）
- title: 专利名称
- patent_type: 专利类型（发明专利/实用新型专利/外观设计专利）
- patentee: 专利权人
- address: 地址
- inventors: 发明人（多人用分号分隔）
- status: 专利状态（受理/初步审查/公开/实质审查/授权）
- patent_number: 专利号/申请号（如 ZL202412345678.9）
- grant_announcement_number: 授权公告号
- apply_date: 专利申请日（YYYY-MM-DD）
- grant_announcement_date: 授权公告日（YYYY-MM-DD）
- applicant_at_apply_date: 申请日时申请人（发明专利特有）
- inventor_at_apply_date: 申请日时发明人（发明专利特有）


### 软著字段
- title: 软件名称（必填）
- copyright_owner: 著作权人（多人用分号分隔，如"张三；李四；王五"）
- completion_date: 开发完成日期（格式 YYYY-MM-DD）
- first_publication_date: 首次发表日期（格式 YYYY-MM-DD）
- right_acquisition_method: 权利取得方式（如"原始取得"、"受让取得"、"继承取得"等）
- right_scope: 权利范围（如"全部权利"、"部分权利"等）
- copyright_number: 登记号（如"2024SR123456"）
- certificate_number: 证书号（如"软著登字第 1234567 号"）
- register_date: 登记日期（格式 YYYY-MM-DD）


### 教研教改和课程建设项目字段（重点新增）
- title: 项目名称（必填，从项目名单表格中提取）
- project_code: 项目编号（序号或正式编号，如"1"、"2019JG001"等）
- project_leader: 项目负责人/主持人（从表格"主持人"列提取，文本格式，如"李超"）
- project_members: 项目参与人（从表格"参加人员"列提取，多人时用顿号分隔，如"刘增明、黄嘉、赵可、杨华文"）
- approval_department: 项目批准部门（从文件头提取，如"湖南省教育厅"）
- approval_date: 项目立项时间（从文件落款日期提取，格式 YYYY-MM，如"2019-09"）
- project_type_name: 项目类型名称（从以下选择：普通本科高校教学改革研究项目、学位与研究生教育改革研究项目、一流本科课程建设项目、课程思政建设项目、其它）
- project_level_name: 项目级别名称（从以下选择：国家级、省部级、市厅级、校级、院级、其它）
- project_category_name: 项目类别名称（从以下选择：重点项目、一般项目、线上一流课程、线上线下混合式一流课程、线下一流课程、社会实践一流课程、虚拟仿真实验教学一流课程、其它）
- funding: 项目经费（数值型，单位元，如 50000）
- start_date: 项目开始时间（格式 YYYY-MM-DD）
- end_date: 项目结束时间（格式 YYYY-MM-DD）

### 教学成果获奖字段
- title: 成果名称（必填）
- achievement_type_name: 教学成果奖类型（湖南中医药大学教学成果奖/湖南中医药大学研究生教学成果奖/湖南省计算机学会高等教育教学成果奖/其它）
- achievement_level_name: 成果等级（国家级/省部级/市厅级/校级/院级/其它）
- award_rank_name: 获奖等级（特等奖/一等奖/二等奖/三等奖/优秀奖/其它）
- main_contributors: 主要完成人（多人用分号分隔）
- completing_units: 成果完成单位（多个用分号分隔）
- award_year: 获奖年度（仅数字）
- certificate_number: 证书编号
- awarding_unit: 颁奖单位
- award_date: 获奖日期（YYYY-MM-DD）

### 教学竞赛获奖
**特征词**：教师团队、教师竞赛、教学竞赛、特发此证、鼓励
- title: 竞赛名称
- award_year: 获奖年度（仅数字）
- competition_level_name: 竞赛等级（国家级/省部级/市厅级/校级/院级/其它）
- award_rank_name: 获奖等级（特等奖/一等奖/二等奖/三等奖/优秀奖/其它）
- winners: 获奖人（多人用分号分隔）
- winner_unit: 获奖人所在单位
- competition_name: 竞赛主办方
- certificate_number: 证书编号
- award_date: 获奖日期（YYYY-MM-DD）

#### 指导学生获奖
**特征词**：指导老师、学生获奖、指导教师、获奖学生、学生姓名
- title: 获奖名称
- award_year: 获奖年度（仅数字）
- competition_name: 竞赛名称
- competition_level_name: 竞赛等级（国家级/省部级/市厅级/校级/院级/其它）
- award_rank_name: 获奖等级（特等奖/一等奖/二等奖/三等奖/优秀奖/其它）
- student_name: 获奖学生
- project_name: 获奖项目名称
- teacher_name: 指导教师
- student_unit: 获奖学生所在单位
- organizer: 竞赛主办方
- certificate_number: 证书编号
- award_date: 获奖日期（YYYY-MM-DD）

【输出规则】
1. 仅输出标准 JSON 字符串，无任何多余文字、注释、反引号
2. 所有字段值为字符串类型，无信息则为空字符串
3. 必须包含 confidence 字段（0-1，代表识别置信度）
4. type_name 字段必须匹配指定的成果类型列表
5. 对于教研教改和课程建设项目，重点关注以下特征词：
   - "教学改革研究"、"教改"、"课程建设"、"一流本科课程"、"课程思政"
   - "学校名称"、"项目名称"、"主持人"、"参加人员"、"项目类别"
   - "普通教育"、"湖南省普通高等学校"等
"""

    payload = {
        "model": "glm-4-flash",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "stream": False
    }


    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        result = response.json()
        ai_content = result['choices'][0]['message']['content'].strip()

        # 清理 AI 返回格式
        ai_content = ai_content.replace('json', '').replace('```', '').strip()
        ai_data_raw = json.loads(ai_content)


        # 检查 AI 返回的数据类型
        if isinstance(ai_data_raw, list):
            ai_data = ai_data_raw[0] if len(ai_data_raw) > 0 and isinstance(ai_data_raw[0], dict) else {}
        elif isinstance(ai_data_raw, dict):
            ai_data = ai_data_raw
        else:
            ai_data = {}

        type_name = ai_data.get('type_name', '')
        if type_name == '教学竞赛获奖':
            student_name = ai_data.get('student_name', '')
            teacher_name = ai_data.get('teacher_name', '')
            has_student_keyword = '学生' in ocr_text or '指导' in ocr_text
            if (student_name or teacher_name) and has_student_keyword:
                type_name = '指导学生获奖'
                ai_data['type_name'] = type_name
        elif type_name == '指导学生获奖':
            student_name = ai_data.get('student_name', '')
            teacher_name = ai_data.get('teacher_name', '')
            has_student_keyword = '学生' in ocr_text or '指导' in ocr_text
            if not student_name and not teacher_name and not has_student_keyword:
                type_name = '教学竞赛获奖'
                ai_data['type_name'] = type_name

        # 数据清洗：确保所有字段存在且为字符串（强制转换，避免 None）
        result_data = {
            # 通用字段
            'type_name': str(ai_data.get('type_name', '其他') or '其他'),
            'title': str(ai_data.get('title', '') or ''),
            'confidence': float(ai_data.get('confidence', 0.8) or 0.8),
            'raw_data': ai_data,

            # 期刊论文（强制转换为字符串，避免 None）
            'authors': str(ai_data.get('authors', '') or ''),
            'corresponding_authors': str(ai_data.get('corresponding_authors', '') or ''),
            'journal_name': str(ai_data.get('journal_name', '') or ''),
            'inclusion_status': str(ai_data.get('inclusion_status', '') or ''),
            'year': str(ai_data.get('year', '') or ''),
            'volume': str(ai_data.get('volume', '') or ''),
            'issue': str(ai_data.get('issue', '') or ''),
            'page_range': str(ai_data.get('page_range', '') or ''),
            'doi': str(ai_data.get('doi', '') or ''),
            'publish_year': str(ai_data.get('publish_year', '') or ''),
            'publish_date': str(ai_data.get('publish_date', '') or ''),

            # 会议论文
            'conference_name': str(ai_data.get('conference_name', '') or ''),
            'conference_time': str(ai_data.get('conference_time', '') or ''),
            'conference_place': str(ai_data.get('conference_place', '') or ''),

            # 教材
            'textbook_series': str(ai_data.get('textbook_series', '') or ''),
            'chief_editor': str(ai_data.get('chief_editor', '') or ''),
            'associate_editors': str(ai_data.get('associate_editors', '') or ''),
            'editorial_board': str(ai_data.get('editorial_board', '') or ''),
            'publisher': str(ai_data.get('publisher', '') or ''),
            'isbn': str(ai_data.get('isbn', '') or ''),
            'cip_number': str(ai_data.get('cip_number', '') or ''),
            'publication_year': str(ai_data.get('publication_year', '') or ''),
            'publication_month': str(ai_data.get('publication_month', '') or ''),
            'edition': str(ai_data.get('edition', '') or ''),
            'word_count': str(ai_data.get('word_count', '') or ''),
            'price': str(ai_data.get('price', '') or ''),
            'textbook_level': str(ai_data.get('textbook_level', '') or ''),
            'textbook_type': str(ai_data.get('textbook_type', '') or ''),
            'applicable_majors': str(ai_data.get('applicable_majors', '') or ''),
            'remarks': str(ai_data.get('remarks', '') or ''),

            # 专著（增强字段）
            'textbook_series': str(ai_data.get('textbook_series', '') or ''),
            'chief_editor': str(ai_data.get('chief_editor', '') or ''),
            'associate_editors': str(ai_data.get('associate_editors', '') or ''),
            'editorial_board': str(ai_data.get('editorial_board', '') or ''),
            'publisher': str(ai_data.get('publisher', '') or ''),
            'isbn': str(ai_data.get('isbn', '') or ''),
            'cip_number': str(ai_data.get('cip_number', '') or ''),
            'publication_year': str(ai_data.get('publication_year', '') or ''),
            'publication_month': str(ai_data.get('publication_month', '') or ''),
            'publish_date': str(ai_data.get('publish_date', '') or ''),
            'edition': str(ai_data.get('edition', '') or ''),
            'word_count': str(ai_data.get('word_count', '') or ''),
            'price': str(ai_data.get('price', '') or ''),
            'monograph_type': str(ai_data.get('monograph_type', '') or ''),
            'applicable_majors': str(ai_data.get('applicable_majors', '') or ''),
            'remarks': str(ai_data.get('remarks', '') or ''),

            # 专利
            'patent_type': str(ai_data.get('patent_type', '') or ''),
            'patentee': str(ai_data.get('patentee', '') or ''),
            'address': str(ai_data.get('address', '') or ''),
            'inventors': str(ai_data.get('inventors', '') or ''),
            'status': str(ai_data.get('status', '') or ''),
            'patent_number': str(ai_data.get('patent_number', '') or ''),
            'grant_announcement_number': str(ai_data.get('grant_announcement_number', '') or ''),
            'apply_date': str(ai_data.get('apply_date', '') or ''),
            'grant_announcement_date': str(ai_data.get('grant_announcement_date', '') or ''),
            'applicant_at_apply_date': str(ai_data.get('applicant_at_apply_date', '') or ''),
            'inventor_at_apply_date': str(ai_data.get('inventor_at_apply_date', '') or ''),

            # 软著
            'copyright_owner': str(ai_data.get('copyright_owner', '') or ''),
            'completion_date': str(ai_data.get('completion_date', '') or ''),
            'first_publication_date': str(ai_data.get('first_publication_date', '') or ''),
            'right_acquisition_method': str(ai_data.get('right_acquisition_method', '') or ''),
            'right_scope': str(ai_data.get('right_scope', '') or ''),
            'copyright_number': str(ai_data.get('copyright_number', '') or ''),
            'certificate_number': str(ai_data.get('certificate_number', '') or ''),
            'register_date': str(ai_data.get('register_date', '') or ''),

            # 教研教改和课程建设项目（新增字段）
            'project_code': str(ai_data.get('project_code', '') or ''),
            'project_leader': str(ai_data.get('project_leader', '') or ''),
            'project_members': str(ai_data.get('project_members', '') or ''),
            'approval_department': str(ai_data.get('approval_department', '') or ''),
            'approval_date': str(ai_data.get('approval_date', '') or ''),
            'project_type_name': str(ai_data.get('project_type_name', '') or ''),
            'project_level_name': str(ai_data.get('project_level_name', '') or ''),
            'project_category_name': str(ai_data.get('project_category_name', '') or ''),
            'funding': str(ai_data.get('funding', '') or ''),
            'start_date': str(ai_data.get('start_date', '') or ''),
            'end_date': str(ai_data.get('end_date', '') or ''),

            # 教学成果获奖（新增字段）
            'achievement_type_name': str(ai_data.get('achievement_type_name', '') or ''),
            'achievement_level_name': str(ai_data.get('achievement_level_name', '') or ''),
            'award_rank_name': str(ai_data.get('award_rank_name', '') or ''),
            'main_contributors': str(ai_data.get('main_contributors', '') or ''),
            'completing_units': str(ai_data.get('completing_units', '') or ''),
            'award_year': str(ai_data.get('award_year', '') or ''),
            'certificate_number': str(ai_data.get('certificate_number', '') or ''),
            'awarding_unit': str(ai_data.get('awarding_unit', '') or ''),

            # 获奖类通用
            'award_level': str(ai_data.get('award_level', '') or ''),
            'award_rank': str(ai_data.get('award_rank', '') or ''),
            'award_date': str(ai_data.get('award_date', '') or ''),
            'competition_name': str(ai_data.get('competition_name', '') or ''),
            'student_name': str(ai_data.get('student_name', '') or ''),
        }

        return result_data
    except Exception as e:
        logger.error(f"AI 分析成果信息失败：{str(e)}")
        # 兜底返回空字段（所有字段强制为空字符串，绝对不能为 None）
        return {
            # 通用字段
            'type_name': '错误',
            'title': '',
            'confidence': 0.5,
            'raw_data': {},

            # 期刊论文（所有字段默认为空字符串）
            'authors': '',
            'corresponding_authors': '',
            'journal_name': '',
            'inclusion_status': '',
            'year': '',
            'volume': '',
            'issue': '',
            'page_range': '',
            'doi': '',
            'publish_year': '',
            'publish_date': '',

            # 会议论文
            'conference_name': '',
            'conference_time': '',
            'conference_place': '',

            # 教材
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'textbook_level': '',
            'textbook_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专著
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'publish_date': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'monograph_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专利
            'patent_type': '',
            'patentee': '',
            'address': '',
            'inventors': '',
            'status': '',
            'patent_number': '',
            'grant_announcement_number': '',
            'apply_date': '',
            'grant_announcement_date': '',
            'applicant_at_apply_date': '',
            'inventor_at_apply_date': '',

            # 软著
            'copyright_owner': '',
            'completion_date': '',
            'first_publication_date': '',
            'right_acquisition_method': '',
            'right_scope': '',
            'copyright_number': '',
            'certificate_number': '',
            'register_date': '',

            # 教研教改和课程建设项目（新增字段）
            'project_code': '',
            'project_leader': '',
            'project_members': '',
            'approval_department': '',
            'approval_date': '',
            'project_type': '',
            'project_level': '',
            'project_category': '',
            'funding': '',
            'start_date': '',
            'end_date': '',

            # 获奖类
            'award_level': '',
            'award_rank': '',
            'award_date': '',
            'competition_name': '',
            'student_name': '',
        }


def pdf_to_images(pdf_path, output_dir=None):
    """
    将PDF文件转换为图片（每页一张），优化大PDF处理
    :param pdf_path: PDF文件路径
    :param output_dir: 图片输出目录（默认临时目录）
    :return: 图片文件路径列表
    """
    if not output_dir:
        output_dir = tempfile.mkdtemp()  # 创建临时目录

    try:
        # 适配Windows/Linux/Mac
        poppler_path = None
        if os.name == 'nt':  # Windows系统
            poppler_path = r"F:\poppler-25.12.0\Library\bin"  # 替换为你的poppler路径

        # 优化：增加参数减少内存占用，分块处理
        images = convert_from_path(
            pdf_path,
            dpi=200,  # 降低分辨率（从300改为200，可根据需要调整）
            output_folder=output_dir,
            fmt='png',
            poppler_path=poppler_path,
            paths_only=True,  # 只返回文件路径，不加载图片对象
            grayscale=True,  # 转为灰度图，减少文件大小
            thread_count=2  # 多线程处理，提升速度
        )
        return images
    except Exception as e:
        logger.error(f"PDF转图片失败：{str(e)}")
        raise Exception(f"PDF转图片失败：{str(e)}")



def init_project_dictionaries():
    """初始化项目字典表数据（首次运行时调用）"""
    try:
        # 先创建所有数据库表
        db.create_all()

        # 初始化项目类型
        project_types = [
            ('普通本科高校教学改革研究项目', 1),
            ('学位与研究生教育改革研究项目', 2),
            ('一流本科课程建设项目', 3),
            ('课程思政建设项目', 4)
        ]
        for type_name, sort in project_types:
            if not ProjectType.query.filter_by(type_name=type_name).first():
                pt = ProjectType(type_name=type_name, sort_order=sort)
                db.session.add(pt)

        # 初始化项目状态
        project_statuses = [
            ('在研', 1),
            ('结题', 2),
            ('延期', 3)
        ]
        for status_name, sort in project_statuses:
            if not ProjectStatus.query.filter_by(status_name=status_name).first():
                ps = ProjectStatus(status_name=status_name, sort_order=sort)
                db.session.add(ps)

        # 初始化项目级别
        project_levels = [
            ('国家级', 1),
            ('省部级', 2),
            ('市厅级', 3),
            ('校级', 4),
            ('院级', 5)
        ]
        for level_name, sort in project_levels:
            if not ProjectLevel.query.filter_by(level_name=level_name).first():
                pl = ProjectLevel(level_name=level_name, sort_order=sort)
                db.session.add(pl)

        # 初始化项目类别
        project_categories = [
            ('重点项目', 1),
            ('一般项目', 2),
            ('线上一流课程', 3),
            ('线上线下混合式一流课程', 4),
            ('线下一流课程', 5),
            ('社会实践一流课程', 6),
            ('虚拟仿真实验教学一流课程', 7)
        ]
        for category_name, sort in project_categories:
            if not ProjectCategory.query.filter_by(category_name=category_name).first():
                pc = ProjectCategory(category_name=category_name, sort_order=sort)
                db.session.add(pc)

        # 初始化专利类型
        patent_types = [
            ('发明专利', 1),
            ('实用新型专利', 2),
            ('外观设计专利', 3)
        ]
        for type_name, sort in patent_types:
            if not PatentType.query.filter_by(type_name=type_name).first():
                pt = PatentType(type_name=type_name, sort_order=sort)
                db.session.add(pt)

        # 初始化专利状态
        patent_statuses = [
            ('受理', 1),
            ('初步审查', 2),
            ('公开', 3),
            ('实质审查', 4),
            ('授权', 5)
        ]
        for status_name, sort in patent_statuses:
            if not PatentStatus.query.filter_by(status_name=status_name).first():
                ps = PatentStatus(status_name=status_name, sort_order=sort)
                db.session.add(ps)

        db.session.commit()
        logger.info("项目字典表初始化完成")
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ 项目字典表初始化失败：{str(e)}")



# 在应用启动时自动初始化字典表
with app.app_context():
    init_project_dictionaries()

# ---------------------- 4. 核心路由 ----------------------
@app.route('/')
def index():
    """首页"""
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))

    # 首页内容（根据角色显示不同内容）
    if user.role == 'teacher':
        content = '''
        <h2>教师工作台</h2>
        <p>欢迎使用教学成果管理系统！</p>
        <p>您可以通过左侧导航栏管理您的论文、专利、获奖等教学成果。</p>
        <ul>
            <li>📄 录入/编辑个人论文、教材等成果</li>
            <li>📊 查看个人成果统计分析</li>
            <li>📤 导出成果数据用于项目申报</li>
        </ul>
        '''
    elif user.role == 'team_leader':
        content = '''
        <h2>团队负责人工作台</h2>
        <p>您可以管理团队成员并查看团队整体成果数据。</p>
        <ul>
            <li>👨‍🏫 管理团队成员</li>
            <li>📊 查看团队成果统计</li>
            <li>📤 导出团队成果数据</li>
            <li>📄 管理个人教学成果</li>
        </ul>
        '''
    else:  # admin
        content = '''
        <h2>系统管理员工作台</h2>
        <p>您可以管理系统用户、团队和全局配置。</p>
        <ul>
            <li>👥 管理所有用户账号</li>
            <li>🏢 创建/删除团队</li>
            <li>📊 查看系统整体数据统计</li>
        </ul>
        '''

    return render_base_layout('首页', content, user)


@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    # 如果已登录，跳转到首页
    if get_current_user():
        return redirect(url_for('index'))

    # 处理登录提交
    if request.method == 'POST':
        login_id = request.form.get('login_id')  # 用户名/工号
        password = request.form.get('password')

        # 验证用户（支持用户名或工号登录）
        user = User.query.filter(
            (User.username == login_id) | (User.employee_id == login_id)
        ).first()

        if user and user.check_password(password):
            # 登录成功，设置session
            session['user_id'] = user.id
            flash('登录成功！', 'success')
            return redirect(url_for('index'))
        else:
            flash('用户名/工号或密码错误！', 'danger')

    # 登录页面HTML（无Jinja）
    flash_messages = ''
    for category, message in session.pop('_flashes', []):
        flash_messages += f'<div class="alert alert-{category}">{message}</div>'

    login_html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>登录 - 教学成果管理系统</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", sans-serif;
            background: #f5f7fa;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }}
        .login-box {{
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 20px rgba(0,0,0,0.1);
            width: 400px;
        }}
        .login-box h2 {{
            text-align: center;
            margin-bottom: 30px;
            color: #2c3e50;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        label {{
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #495057;
        }}
        input {{
            width: 100%;
            padding: 10px 15px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 14px;
        }}
        input:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
        }}
        .btn {{
            width: 100%;
            padding: 10px;
            background: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 16px;
            margin-top: 10px;
        }}
        .btn:hover {{
            background: #2980b9;
        }}
        .alert {{
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 4px;
        }}
        .alert-success {{
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }}
        .alert-danger {{
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }}
        .register-link {{
            text-align: center;
            margin-top: 20px;
        }}
        .register-link a {{
            color: #3498db;
            text-decoration: none;
        }}
    </style>
</head>
<body>
    <div class="login-box">
        <h2>教学成果管理系统</h2>
        {flash_messages}
        <form method="POST">
            <div class="form-group">
                <label for="login_id">用户名/工号</label>
                <input type="text" id="login_id" name="login_id" required>
            </div>
            <div class="form-group">
                <label for="password">密码</label>
                <input type="password" id="password" name="password" required>
            </div>
            <button type="submit" class="btn">登录</button>
        </form>
        <div class="register-link">
            <a href="/register">还没有账号？点击注册</a>
        </div>
    </div>
</body>
</html>
'''
    return login_html


@app.route('/register', methods=['GET', 'POST'])
def register():
    """注册页面（仅显示必填项）"""
    # 如果已登录，跳转到首页
    if get_current_user():
        return redirect(url_for('index'))

    # 处理注册提交
    if request.method == 'POST':
        try:
            # 获取表单数据（仅保留必填项）
            username = request.form.get('username')
            password = request.form.get('password')
            employee_id = request.form.get('employee_id')
            email = request.form.get('email')
            user_role = request.form.get('role', 'teacher')

            # 检查必填字段唯一性
            if User.query.filter_by(username=username).first():
                flash('用户名已存在！', 'danger')
                return redirect(url_for('register'))

            if User.query.filter_by(employee_id=employee_id).first():
                flash('工号已存在！', 'danger')
                return redirect(url_for('register'))

            if User.query.filter_by(email=email).first():
                flash('邮箱已存在！', 'danger')
                return redirect(url_for('register'))

            # 安全校验：仅允许teacher/team_leader角色
            if user_role not in ['teacher', 'team_leader']:
                user_role = 'teacher'

            # 创建用户（仅初始化必填字段）
            user = User(
                username=username,
                employee_id=employee_id,
                email=email,
                role=user_role
            )
            user.set_password(password)

            # 保存到数据库
            db.session.add(user)
            db.session.commit()

            flash('注册成功！请登录', 'success')
            return redirect(url_for('login'))

        except Exception as e:
            db.session.rollback()
            flash(f'注册失败：{str(e)}', 'danger')

    # 注册页面HTML（仅保留必填项）
    flash_messages = ''
    for category, message in session.pop('_flashes', []):
        flash_messages += f'<div class="alert alert-{category}">{message}</div>'

    register_html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>注册 - 教学成果管理系统</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", sans-serif;
            background: #f5f7fa;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }}
        .register-box {{
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 20px rgba(0,0,0,0.1);
            width: 400px; /* 缩小宽度，适配少字段 */
        }}
        .register-box h2 {{
            text-align: center;
            margin-bottom: 30px;
            color: #2c3e50;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        label {{
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #495057;
        }}
        input, select {{
            width: 100%;
            padding: 10px 15px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 14px;
        }}
        input:focus, select:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
        }}
        .btn {{
            width: 100%;
            padding: 10px;
            background: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 16px;
            margin-top: 20px;
        }}
        .btn:hover {{
            background: #2980b9;
        }}
        .alert {{
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 4px;
        }}
        .alert-success {{
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }}
        .alert-danger {{
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }}
        .login-link {{
            text-align: center;
            margin-top: 20px;
        }}
        .login-link a {{
            color: #3498db;
            text-decoration: none;
        }}
        .required {{
            color: red;
        }}
    </style>
</head>
<body>
    <div class="register-box">
        <h2>用户注册</h2>
        {flash_messages}
        <form method="POST">
            <!-- 仅保留必填字段 -->
            <div class="form-group">
                <label for="username">用户名 <span class="required">*</span></label>
                <input type="text" id="username" name="username" required>
            </div>

            <div class="form-group">
                <label for="employee_id">工号 <span class="required">*</span></label>
                <input type="text" id="employee_id" name="employee_id" required>
            </div>

            <div class="form-group">
                <label for="email">邮箱 <span class="required">*</span></label>
                <input type="email" id="email" name="email" required>
            </div>

            <div class="form-group">
                <label for="role">用户角色 <span class="required">*</span></label>
                <select id="role" name="role" required>
                    <option value="">请选择</option>
                    <option value="teacher">普通教师</option>
                    <option value="team_leader">团队负责人</option>
                </select>
            </div>

            <div class="form-group">
                <label for="password">密码 <span class="required">*</span></label>
                <input type="password" id="password" name="password" required minlength="6">
            </div>

            <button type="submit" class="btn">注册</button>
        </form>
        <div class="login-link">
            <a href="/login">已有账号？返回登录</a>
        </div>
    </div>
</body>
</html>
'''
    return register_html


@app.route('/logout')
def logout():
    """登出（清空所有session数据，包括flash消息）"""
    # 清空整个session，而非仅删除user_id
    session.clear()
    flash('已成功退出登录！', 'success')
    return redirect(url_for('login'))


# ---------------------- 通用成果管理路由 ----------------------
@app.route('/user/settings', methods=['GET', 'POST'])
def user_settings():
    """个人信息修改（邮箱、电话、身份证等）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心新增：管理员禁止访问个人信息设置
    if current_user.role == 'admin':
        flash('管理员无需设置个人账户信息！', 'danger')
        return redirect(url_for('index'))

    # 处理表单提交
    if request.method == 'POST':
        try:
            # 基础信息更新
            current_user.gender = request.form.get('gender') or None
            birth_date_str = request.form.get('birth_date')
            if birth_date_str:
                current_user.birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            else:
                current_user.birth_date = None

            # 关键修复：空身份证号转为 None
            current_user.id_card = request.form.get('id_card').strip() if request.form.get('id_card') else None
            current_user.email = request.form.get('email', '')  # 必填，前端已校验
            current_user.phone = request.form.get('phone') or None
            current_user.office_phone = request.form.get('office_phone') or None
            current_user.school = request.form.get('school') or None
            current_user.college = request.form.get('college') or None
            current_user.department = request.form.get('department') or None
            current_user.research_room = request.form.get('research_room') or None

            db.session.commit()
            flash('个人信息修改成功！', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'修改失败：{str(e)}', 'danger')

    # 渲染个人信息修改页面
    birth_date = current_user.birth_date.strftime('%Y-%m-%d') if current_user.birth_date else ''
    form_html = f'''
    <h2>个人信息修改</h2>
    <form method="POST">
        <div class="form-group">
            <label>用户名（不可修改）</label>
            <input type="text" value="{current_user.username}" disabled>
        </div>
        <div class="form-group">
            <label>工号（不可修改）</label>
            <input type="text" value="{current_user.employee_id}" disabled>
        </div>
        <div class="form-group">
            <label>性别</label>
            <select name="gender">
                <option value="">请选择</option>
                <option value="男" {"selected" if current_user.gender == '男' else ''}>男</option>
                <option value="女" {"selected" if current_user.gender == '女' else ''}>女</option>
            </select>
        </div>
        <div class="form-group">
            <label>出生年月日</label>
            <input type="date" name="birth_date" value="{birth_date}">
        </div>
        <div class="form-group">
            <label>身份证号码</label>
            <input type="text" name="id_card" value="{current_user.id_card or ''}" maxlength="18">
        </div>
        <div class="form-group">
            <label>邮箱 <span class="required">*</span></label>
            <input type="email" name="email" value="{current_user.email}" required>
        </div>
        <div class="form-group">
            <label>手机号</label>
            <input type="tel" name="phone" value="{current_user.phone or ''}">
        </div>
        <div class="form-group">
            <label>办公电话</label>
            <input type="tel" name="office_phone" value="{current_user.office_phone or ''}">
        </div>
        <div class="form-group">
            <label>学校</label>
            <input type="text" name="school" value="{current_user.school or ''}">
        </div>
        <div class="form-group">
            <label>学院</label>
            <input type="text" name="college" value="{current_user.college or ''}">
        </div>
        <div class="form-group">
            <label>系部</label>
            <input type="text" name="department" value="{current_user.department or ''}">
        </div>
        <div class="form-group">
            <label>教研室</label>
            <input type="text" name="research_room" value="{current_user.research_room or ''}">
        </div>
        <button type="submit" class="btn">保存修改</button>
    </form>
    <div style="margin-top:20px;">
        <a href="/user/change_password" class="btn">修改密码</a>
        <a href="/user/api_config" class="btn">大模型API配置</a>
    </div>
    '''
    return render_base_layout('个人信息设置', form_html, current_user)


@app.route('/user/change_password', methods=['GET', 'POST'])
def change_password():
    """密码修改（个人主动改密码）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心新增：管理员禁止修改密码（需通过数据库重置）
    if current_user.role == 'admin':
        flash('管理员密码请通过数据库手动重置！', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        old_pwd = request.form.get('old_password')
        new_pwd = request.form.get('new_password')

        # 校验
        if not current_user.check_password(old_pwd):
            flash('原密码错误！', 'danger')
        elif len(new_pwd) < 6:
            flash('新密码长度不能少于6位！', 'danger')
        else:
            try:
                current_user.set_password(new_pwd)
                db.session.commit()
                flash('密码修改成功，请重新登录！', 'success')
                return redirect(url_for('logout'))
            except Exception as e:
                db.session.rollback()
                flash(f'修改失败：{str(e)}', 'danger')

    # 渲染密码修改页面
    form_html = '''
    <h2>修改密码</h2>
    <form method="POST">
        <div class="form-group">
            <label>原密码 <span class="required">*</span></label>
            <input type="password" name="old_password" required>
        </div>
        <div class="form-group">
            <label>新密码 <span class="required">*</span></label>
            <input type="password" name="new_password" required minlength="6">
        </div>
        <button type="submit" class="btn">确认修改</button>
    </form>
    '''
    return render_base_layout('修改密码', form_html, current_user)


@app.route('/user/api_config', methods=['GET', 'POST'])
def api_config():
    """大模型API配置（仅保留百度+智谱）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心新增：管理员禁止配置API
    if current_user.role == 'admin':
        flash('管理员无需配置个人账户的API！', 'danger')
        return redirect(url_for('index'))

    # 获取现有API配置
    api_config = current_user.get_api_config() or {}

    if request.method == 'POST':
        try:
            # 仅保留百度+智谱API配置
            new_config = {
                'baidu': {
                    'api_key': request.form.get('baidu_api_key', ''),
                    'secret_key': request.form.get('baidu_secret_key', '')
                },
                'zhipu': {
                    'api_key': request.form.get('zhipu_api_key', '')
                }
            }
            current_user.set_api_config(new_config)
            db.session.commit()
            flash('API配置保存成功！', 'success')
            api_config = new_config  # 更新页面展示数据
        except Exception as e:
            db.session.rollback()
            flash(f'保存失败：{str(e)}', 'danger')

    # 渲染简化后的API配置页面（仅百度+智谱）
    form_html = f'''
    <h2>大模型API配置</h2>
    <div class="alert alert-info">
        配置完成后可用于OCR智能导入、语音导入/导出等功能
    </div>
    <form method="POST">
        <h3 style="margin-top:20px;">百度文心一言API配置</h3>
        <div class="form-group">
            <label>API Key</label>
            <input type="text" name="baidu_api_key" value="{api_config.get('baidu', {}).get('api_key', '')}" >
        </div>
        <div class="form-group">
            <label>Secret Key</label>
            <input type="text" name="baidu_secret_key" value="{api_config.get('baidu', {}).get('secret_key', '')}" >
        </div>

        <h3 style="margin-top:30px;">智谱AI（ZHIPU）API配置</h3>
        <div class="form-group">
            <label>API Key</label>
            <input type="text" name="zhipu_api_key" value="{api_config.get('zhipu', {}).get('api_key', '')}" >
        </div>

        <button type="submit" class="btn" style="margin-top:30px;">保存配置</button>
    </form>
    '''
    return render_base_layout('大模型API配置', form_html, current_user)

@app.route('/admin/user_manage', methods=['GET', 'POST'])
def admin_user_manage():
    """管理员-用户管理（仅查看角色，不可修改）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'admin':
        flash('无管理员权限！', 'danger')
        return redirect(url_for('index'))

    # 处理用户创建/删除（保留创建、删除功能，移除角色编辑）
    if request.method == 'POST':
        action = request.form.get('action')
        user_id = request.form.get('user_id')

        try:
            if action == 'create':
                # 创建新用户
                username = request.form.get('username')
                employee_id = request.form.get('employee_id')
                email = request.form.get('email')
                role = request.form.get('role', 'teacher')
                password = request.form.get('password', '123456')  # 默认密码

                # 校验唯一性
                if User.query.filter_by(username=username).first():
                    flash('用户名已存在！', 'danger')
                    return redirect(url_for('admin_user_manage'))
                if User.query.filter_by(employee_id=employee_id).first():
                    flash('工号已存在！', 'danger')
                    return redirect(url_for('admin_user_manage'))

                # 安全校验：仅允许teacher/team_leader角色
                if user_role not in ['teacher', 'team_leader']:
                    user_role = 'teacher'

                # 创建用户
                new_user = User(
                    username=username,
                    employee_id=employee_id,
                    email=email,
                    role=role
                )
                new_user.set_password(password)
                db.session.add(new_user)
                flash(f'用户{username}创建成功（默认密码：{password}）', 'success')

            elif action == 'delete':
                # 删除用户（非管理员）
                user = db.session.get(User, user_id)
                if user and user.username != 'admin':
                    # 删除关联数据（简化版，实际可保留成果数据）
                    db.session.delete(user)
                    flash(f'用户{user.username}删除成功！', 'success')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 查询所有用户
    users = User.query.order_by(User.role, User.username).all()

    # 渲染用户管理页面（移除角色编辑下拉框，仅展示角色）
    user_list_html = '''
    <h2>用户管理</h2>
    <div style="margin-bottom:20px;">
        <button onclick="showCreateForm()" class="btn">新增用户</button>
    </div>

    <!-- 新增用户表单 -->
    <div id="createForm" style="display:none; margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h3>新增用户</h3>
        <form method="POST">
            <input type="hidden" name="action" value="create">
            <div class="form-group">
                <label>用户名 <span class="required">*</span></label>
                <input type="text" name="username" required>
            </div>
            <div class="form-group">
                <label>工号 <span class="required">*</span></label>
                <input type="text" name="employee_id" required>
            </div>
            <div class="form-group">
                <label>邮箱 <span class="required">*</span></label>
                <input type="email" name="email" required>
            </div>
            <div class="form-group">
                <label>角色 <span class="required">*</span></label>
                <select name="role">
                    <option value="teacher">普通教师</option>
                    <option value="team_leader">团队负责人</option>
                    <option value="admin">管理员（谨慎）</option>
                </select>
            </div>
            <div class="form-group">
                <label>初始密码（默认：123456）</label>
                <input type="password" name="password" value="123456">
            </div>
            <button type="submit" class="btn">创建</button>
            <button type="button" onclick="hideCreateForm()" class="btn" style="background:#95a5a6;">取消</button>
        </form>
    </div>

    <!-- 用户列表 -->
    <table style="width:100%; border-collapse:collapse;">
        <thead>
            <tr style="background:#f5f7fa;">
                <th style="padding:10px; border:1px solid #dee2e6;">用户名</th>
                <th style="padding:10px; border:1px solid #dee2e6;">工号</th>
                <th style="padding:10px; border:1px solid #dee2e6;">邮箱</th>
                <th style="padding:10px; border:1px solid #dee2e6;">角色</th>
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    for user in users:
        # 仅展示角色，移除编辑下拉框
        role_display = {
            'teacher': '普通教师',
            'team_leader': '团队负责人',
            'admin': '管理员'
        }.get(user.role, '未知角色')

        # 角色样式（区分不同角色）
        role_style = ''
        if user.role == 'admin':
            role_style = 'style="color: #e74c3c; font-weight: bold;"'
        elif user.role == 'team_leader':
            role_style = 'style="color: #2980b9; font-weight: bold;"'

        # 删除按钮（超级管理员不可删）
        delete_btn = ''
        if user.username != 'admin':
            delete_btn = f'''
            <form method="POST" style="display:inline;" onsubmit="return confirm('确定删除？')">
                <input type="hidden" name="action" value="delete">
                <input type="hidden" name="user_id" value="{user.id}">
                <button type="submit" class="btn" style="padding:5px 10px; font-size:12px; background:#e74c3c;">删除</button>
            </form>
            '''
        else:
            delete_btn = '<span style="color:#999;">不可删除</span>'

        user_list_html += f'''
        <tr>
            <td style="padding:10px; border:1px solid #dee2e6;">{user.username}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{user.employee_id}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{user.email}</td>
            <td style="padding:10px; border:1px solid #dee2e6;" {role_style}>{role_display}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{delete_btn}</td>
        </tr>
        '''

    user_list_html += '''
        </tbody>
    </table>

    <script>
        function showCreateForm() {
            document.getElementById('createForm').style.display = 'block';
        }
        function hideCreateForm() {
            document.getElementById('createForm').style.display = 'none';
        }
    </script>
    '''
    return render_base_layout('用户管理', user_list_html, current_user)


@app.route('/admin/team_manage', methods=['GET', 'POST'])
def admin_team_manage():
    """管理员-团队管理（创建/删除团队、指定负责人、添加成员）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'admin':
        flash('无管理员权限！', 'danger')
        return redirect(url_for('index'))

    # 处理团队操作
    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'create_team':
                # 创建团队
                team_name = request.form.get('team_name')
                leader_id = request.form.get('leader_id')

                if Team.query.filter_by(name=team_name).first():
                    flash('团队名称已存在！', 'danger')
                else:
                    leader_user = db.session.get(User, leader_id)
                    new_team = Team(
                        name=team_name,
                        leader_id=leader_id
                    )
                    db.session.add(new_team)
                    db.session.flush()
                    # 自动将负责人加入团队
                    db.session.add(UserTeam(user_id=leader_id, team_id=new_team.id))
                    flash(f'团队{team_name}创建成功！已将{leader_user.username}设为团队负责人', 'success')

            elif action == 'delete_team':
                # 删除团队
                team_id = request.form.get('team_id')
                team = db.session.get(Team, team_id)
                if team:
                    # 删除团队成员关联
                    UserTeam.query.filter_by(team_id=team_id).delete()
                    # 删除团队
                    db.session.delete(team)
                    flash(f'团队{team.name}删除成功！', 'success')

            elif action == 'add_member':
                # 添加团队成员
                team_id = request.form.get('team_id')
                user_id = request.form.get('user_id')

                if UserTeam.query.filter_by(team_id=team_id, user_id=user_id).first():
                    flash('该用户已在团队中！', 'danger')
                else:
                    db.session.add(UserTeam(team_id=team_id, user_id=user_id))
                    flash('成员添加成功！', 'success')

            elif action == 'remove_member':
                # 移除团队成员
                ut_id = request.form.get('ut_id')
                ut = db.session.get(UserTeam, ut_id)
                if ut:
                    db.session.delete(ut)
                    flash('成员移除成功！', 'success')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 查询所有团队和用户
    teams = Team.query.all()
    all_users = User.query.filter(User.role != 'admin').all()  # 管理员不加入团队

    # 渲染团队管理页面
    team_html = '''
    <h2>团队管理</h2>

    <!-- 创建团队表单 -->
    <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h3>创建新团队</h3>
        <form method="POST">
            <input type="hidden" name="action" value="create_team">
            <div class="form-group">
                <label>团队名称 <span class="required">*</span></label>
                <input type="text" name="team_name" required>
            </div>
            <div class="form-group">
                <label>团队负责人 <span class="required">*</span></label>
                <select name="leader_id" required>
                    <option value="">请选择</option>
    '''
    # 填充负责人选项 - 仅允许选择注册为团队负责人的用户
    for user in all_users:
        if user.role == 'team_leader':  # 仅显示注册时就是团队负责人的用户
            team_html += f'<option value="{user.id}">{user.username}（{user.employee_id}）</option>'

    team_html += '''
                </select>
            </div>
            <button type="submit" class="btn">创建团队</button>
        </form>
    </div>

    <!-- 团队列表 -->
    '''
    for team in teams:
        leader = db.session.get(User, team.leader_id)
        # 查询团队成员
        members = UserTeam.query.filter_by(team_id=team.id).all()
        member_list = []
        for ut in members:
            user = db.session.get(User, ut.user_id)
            member_list.append((ut.id, user))

        # 团队卡片
        team_html += f'''
        <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:20px;">
                <h3>{team.name}</h3>
                <div>
                    <form method="POST" onsubmit="return confirm('确定删除该团队？')" style="display:inline;">
                        <input type="hidden" name="action" value="delete_team">
                        <input type="hidden" name="team_id" value="{team.id}">
                        <button type="submit" class="btn" style="background:#e74c3c;">删除团队</button>
                    </form>
                </div>
            </div>
            <div style="margin-bottom:10px;">
                <strong>团队负责人：</strong>{leader.username}（{leader.employee_id}）
            </div>

            <!-- 添加成员 -->
            <div style="margin-bottom:20px;">
                <form method="POST" style="display:flex; gap:10px; align-items:end;">
                    <input type="hidden" name="action" value="add_member">
                    <input type="hidden" name="team_id" value="{team.id}">
                    <div class="form-group" style="flex:1;">
                        <label>添加团队成员</label>
                        <select name="user_id" required>
                            <option value="">请选择用户</option>
        '''
        # 填充可选用户（排除已加入的）
        for user in all_users:
            is_in_team = any(ut.user_id == user.id for ut in members)
            if not is_in_team:
                team_html += f'<option value="{user.id}">{user.username}（{user.employee_id}）</option>'

        team_html += '''
                        </select>
                    </div>
                    <button type="submit" class="btn">添加</button>
                </form>
            </div>

            <!-- 成员列表 -->
            <div>
                <strong>团队成员：</strong>
                <ul style="margin:10px 0; padding-left:20px;">
        '''
        for ut_id, user in member_list:
            # 移除成员按钮（负责人不可移除）
            remove_btn = ''
            if user.id != team.leader_id:
                remove_btn = f'''
                <form method="POST" style="display:inline; margin-left:10px;">
                    <input type="hidden" name="action" value="remove_member">
                    <input type="hidden" name="ut_id" value="{ut_id}">
                    <button type="submit" class="btn" style="padding:2px 8px; font-size:12px; background:#95a5a6;" onclick="return confirm('确定移除？')">移除</button>
                </form>
                '''
            team_html += f'<li>{user.username}（{user.employee_id}）{remove_btn}</li>'

        team_html += '''
                </ul>
            </div>
        </div>
        '''

    return render_base_layout('团队管理', team_html, current_user)


@app.route('/team/achievements')
def team_achievements():
    """团队负责人-团队成果多维度统计（仅统计公开给本团队的成果）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 获取团队信息和成员ID
    teams = Team.query.filter_by(leader_id=current_user.id).all()
    team_ids = [t.id for t in teams]
    team_user_ids = [ut.user_id for ut in UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()]
    if not team_user_ids:
        content = '<div class="alert alert-warning">暂无团队成员数据！</div>'
        return render_base_layout('团队成果统计', content, current_user)

    # 关键：获取当前管理的团队ID字符串列表（用于过滤public_team_ids）
    managed_team_ids_str = [str(t.id) for t in teams]

    # 多维度统计
    stats = {
        'total': {},  # 总数统计
        'by_type': {},  # 按成果类型统计
        'by_year': {},  # 按年份统计
        'by_member': {}  # 按成员统计
    }

    # 成果模型列表
    achievement_models = [
        ('期刊论文', JournalPaper),
        ('会议论文', ConferencePaper),
        ('教材', Textbook),
        ('专著', Monograph),
        ('教研教改和课程建设项目', TeachingProject),
        ('专利', Patent),
        ('软著', SoftwareCopyright),
        ('教学成果获奖', TeachingAchievementAward),
        ('教学竞赛获奖', TeachingCompetitionAward),
        ('指导学生获奖', StudentGuidanceAward)
    ]

    # 统计总数和按类型（仅统计公开给本团队的成果）
    total_count = 0
    for name, model in achievement_models:
        # 构建过滤条件：1. 属于团队成员 2. 公开给当前管理的任意团队
        or_conditions = []
        for team_id in managed_team_ids_str:
            # 处理格式：",1,2,3," 避免部分匹配（如1匹配10）
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )

        # 执行过滤查询
        query = model.query.filter(
            model.user_id.in_(team_user_ids),
            or_(*or_conditions)
        )
        count = query.count()

        stats['total'][name] = count
        total_count += count
    stats['total']['总计'] = total_count

    # 按年份统计（仅统计公开给本团队的成果）
    year_fields = {
        JournalPaper: 'publish_year',
        ConferencePaper: 'publish_year',
        Textbook: 'publish_date',
        Monograph: 'publish_date',
        TeachingProject: 'start_date',
        Patent: 'apply_date',
        SoftwareCopyright: 'register_date',
        TeachingAchievementAward: 'award_date',
        TeachingCompetitionAward: 'award_date',
        StudentGuidanceAward: 'award_date'
    }
    for name, model in achievement_models:
        # 过滤公开给本团队的成果
        or_conditions = []
        for team_id in managed_team_ids_str:
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )

        items = model.query.filter(
            model.user_id.in_(team_user_ids),
            or_(*or_conditions)
        ).all()

        field = year_fields[model]
        for item in items:
            value = getattr(item, field)
            if value:
                year = value.year if isinstance(value, date) else value
                if year not in stats['by_year']:
                    stats['by_year'][year] = {n: 0 for n, _ in achievement_models}
                stats['by_year'][year][name] += 1

    # 按成员统计（仅统计公开给本团队的成果）
    team_users = User.query.filter(User.id.in_(team_user_ids)).all()
    for user in team_users:
        user_count = {}
        for name, model in achievement_models:
            # 过滤该用户公开给本团队的成果
            or_conditions = []
            for team_id in managed_team_ids_str:
                or_conditions.append(
                    func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
                )

            user_count[name] = model.query.filter(
                model.user_id == user.id,
                or_(*or_conditions)
            ).count()

        user_count['总计'] = sum(user_count.values())
        stats['by_member'][user.username] = user_count

    # 渲染统计页面（添加导出按钮）
    stats_html = f'''
    <h2>团队成果统计（负责人：{current_user.username}）</h2>

    <!-- 总数统计 -->
    <div style="margin-bottom:30px;">
        <h3>成果总数</h3>
        <div style="padding:20px; background:#f5f7fa; border-radius:8px;">
            <p>团队总成果数：<strong>{stats['total']['总计']}</strong> 项</p>
            <table style="width:100%; border-collapse:collapse; margin-top:10px;">
                <thead>
                    <tr style="background:#e9ecef;">
                        <th style="padding:10px; border:1px solid #dee2e6;">成果类型</th>
                        <th style="padding:10px; border:1px solid #dee2e6;">数量（项）</th>
                    </tr>
                </thead>
                <tbody>
    '''
    # 核心修改：为每个成果类型添加导出按钮
    for name, count in stats['total'].items():
        if name != '总计':
            # 仅当有成果时显示导出按钮
            export_btn = ''
            if count > 0 and teams:  # 确保有团队ID
                export_btn = f'''
                <a href="/team/export_achievement?team_id={teams[0].id}&type={name}" 
                   class="btn" 
                   style="padding:5px 10px; font-size:12px; background:#27ae60; margin-left:10px;">
                    导出公开成果
                </a>
                '''
            stats_html += f'''
            <tr>
                <td style="padding:10px; border:1px solid #dee2e6;">{name}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">
                    {count}
                    {export_btn}
                </td>
            </tr>
            '''
    stats_html += '''
                </tbody>
            </table>
        </div>
    </div>

    <!-- 按成员统计 -->
    <div style="margin-bottom:30px;">
        <h3>按成员统计（仅统计公开给本团队的成果）</h3>
        <table style="width:100%; border-collapse:collapse;">
            <thead>
                <tr style="background:#e9ecef;">
                    <th style="padding:10px; border:1px solid #dee2e6;">团队成员</th>
    '''
    # 成员统计表头
    for name, _ in achievement_models:
        stats_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{name}</th>'
    stats_html += '<th style="padding:10px; border:1px solid #dee2e6;">总计</th>'
    stats_html += '''
                </tr>
            </thead>
            <tbody>
    '''
    # 成员统计数据
    for username, counts in stats['by_member'].items():
        stats_html += f'<tr><td style="padding:10px; border:1px solid #dee2e6;">{username}</td>'
        for name, _ in achievement_models:
            stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{counts[name]}</td>'
        stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;"><strong>{counts["总计"]}</strong></td></tr>'
    stats_html += '''
            </tbody>
        </table>
    </div>

    <!-- 按年份统计 -->
    <div>
        <h3>按年份统计（仅统计公开给本团队的成果）</h3>
        <table style="width:100%; border-collapse:collapse;">
            <thead>
                <tr style="background:#e9ecef;">
                    <th style="padding:10px; border:1px solid #dee2e6;">年份</th>
    '''
    # 年份统计表头
    for name, _ in achievement_models:
        stats_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{name}</th>'
    stats_html += '<th style="padding:10px; border:1px solid #dee2e6;">总计</th>'
    stats_html += '''
                </tr>
            </thead>
            <tbody>
    '''
    # 年份统计数据
    for year in sorted(stats['by_year'].keys(), reverse=True):
        year_data = stats['by_year'][year]
        year_total = sum(year_data.values())
        stats_html += f'<tr><td style="padding:10px; border:1px solid #dee2e6;">{year}</td>'
        for name, _ in achievement_models:
            stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{year_data[name]}</td>'
        stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;"><strong>{year_total}</strong></td></tr>'
    stats_html += '''
            </tbody>
        </table>
    </div>
    '''
    return render_base_layout('团队成果统计', stats_html, current_user)


@app.route('/team/export_achievement')
def team_export_achievement():
    """团队负责人导出指定类型的公开成果"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 获取导出参数
    team_id = request.args.get('team_id', type=int)
    achievement_type = request.args.get('type')
    if not team_id or not achievement_type:
        flash('导出参数缺失！', 'danger')
        return redirect(url_for('team_achievements'))

    # 验证团队归属（当前用户是该团队负责人）
    team = db.session.get(Team, team_id)
    if not team or team.leader_id != current_user.id:
        flash('无权限导出该团队成果！', 'danger')
        return redirect(url_for('team_achievements'))

    # 成果类型映射
    type_mapping = {
        '期刊论文': (JournalPaper, 'journal'),
        '会议论文': (ConferencePaper, 'conference'),
        '教材': (Textbook, 'textbook'),
        '专著': (Monograph, 'monograph'),
        '教研教改和课程建设项目': (TeachingProject, 'teaching_project'),
        '专利': (Patent, 'patent'),
        '软著': (SoftwareCopyright, 'software_copyright'),
        '教学成果获奖': (TeachingAchievementAward, 'teaching_achievement_award'),
        '教学竞赛获奖': (TeachingCompetitionAward, 'teaching_competition_award'),
        '指导学生获奖': (StudentGuidanceAward, 'student_guidance_award')
    }

    if achievement_type not in type_mapping:
        flash('不支持的成果类型！', 'danger')
        return redirect(url_for('team_achievements'))

    model, export_type = type_mapping[achievement_type]
    team_id_str = str(team_id)

    # 过滤：仅导出公开给该团队的成果
    query = model.query.filter(
        func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id_str, ',')) > 0
    )

    # 字段配置
    fields_config_map = {
        'journal_paper': [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'journal_name', 'label': '期刊名称'},
            {'name': 'inclusion_status', 'label': '收录情况'},
            {'name': 'year', 'label': '年'},
            {'name': 'volume', 'label': '卷'},
            {'name': 'issue', 'label': '期'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'conference_paper': [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'conference_name', 'label': '会议名称'},
            {'name': 'conference_time', 'label': '会议时间'},
            {'name': 'conference_place', 'label': '会议地点'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'textbook': [
            {'name': 'title', 'label': '教材名称'},
            {'name': 'textbook_series', 'label': '教材系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'textbook_attachment', 'label': '附件'}
        ],
        'monograph': [
            {'name': 'title', 'label': '专著名称'},
            {'name': 'textbook_series', 'label': '专著系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'monograph_attachment', 'label': '附件'}
        ],
        'teaching_project': [
            {'name': 'title', 'label': '项目名称'},
            {'name': 'project_code', 'label': '项目编号'},
            {'name': 'project_leader', 'label': '项目负责人'},
            {'name': 'project_members', 'label': '项目参与人'},
            {'name': 'approval_department', 'label': '批准部门'},
            {'name': 'approval_date', 'label': '立项时间'},
            {'name': 'funding', 'label': '经费'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'patent': [
            {'name': 'title', 'label': '专利名称'},
            {'name': 'inventors', 'label': '发明人'},
            {'name': 'patent_number', 'label': '专利号'},
            {'name': 'apply_date', 'label': '申请日'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'software_copyright': [
            {'name': 'title', 'label': '软件名称'},
            {'name': 'copyright_owner', 'label': '著作权人'},
            {'name': 'copyright_number', 'label': '登记号'},
            {'name': 'certificate_number', 'label': '证书号'},
            {'name': 'register_date', 'label': '登记日期'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'teaching_achievement_award': [
            {'name': 'title', 'label': '成果名称'},
            {'name': 'achievement_type', 'label': '教学成果奖类型', 'relation': 'achievement_type',
             'relation_field': 'type_name'},
            {'name': 'achievement_level', 'label': '成果等级', 'relation': 'achievement_level',
             'relation_field': 'level_name'},
            {'name': 'main_contributors', 'label': '主要完成人'},
            {'name': 'completing_units', 'label': '成果完成单位'},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'award_rank', 'label': '获奖等级', 'relation': 'award_rank', 'relation_field': 'rank_name'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'awarding_unit', 'label': '颁奖单位'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'teaching_competition_award': [
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'title', 'label': '竞赛名称'},
            {'name': 'award_rank', 'label': '获奖等级', 'relation': 'award_rank', 'relation_field': 'rank_name'},
            {'name': 'winners', 'label': '获奖人'},
            {'name': 'winner_unit', 'label': '获奖人所在单位'},
            {'name': 'competition_level', 'label': '竞赛等级', 'relation': 'competition_level', 'relation_field': 'level_name'},
            {'name': 'competition_name', 'label': '竞赛主办方'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'student_guidance_award': [
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'title', 'label': '竞赛名称'},
            {'name': 'award_rank', 'label': '获奖等级', 'relation': 'award_rank', 'relation_field': 'rank_name'},
            {'name': 'student_name', 'label': '获奖学生'},
            {'name': 'project_name', 'label': '获奖项目名称'},
            {'name': 'teacher_name', 'label': '指导教师'},
            {'name': 'student_unit', 'label': '获奖学生所在单位'},
            {'name': 'competition_level', 'label': '竞赛等级', 'relation': 'competition_level', 'relation_field': 'level_name'},
            {'name': 'organizer', 'label': '竞赛主办方'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'attachment', 'label': '附件'}
        ]
    }

    fields_config = fields_config_map.get(export_type, [])
    items = query.all()

    # 核心修改：所有类型统一导出为Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{achievement_type}公开成果'

    # 表头
    headers = [f['label'] for f in fields_config]
    ws.append(headers)

    # 数据行
    for item in items:
        row = []
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            # 处理关联表字段
            if field.get('relation'):
                relation_obj = getattr(item, field['relation'], None)
                if relation_obj:
                    value = getattr(relation_obj, field.get('relation_field', 'name'), '')
                else:
                    value = ''

            if value is None:
                value = ''
            elif isinstance(value, (date, datetime)):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field_name == 'attachment' and value:
                value = os.path.basename(value) if value else ''
            row.append(value)
        ws.append(row)

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{achievement_type}_公开成果_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/team/manage_members', methods=['GET', 'POST'])
def manage_members():
    """团队负责人-团队成员管理（添加/移除）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 处理成员操作
    if request.method == 'POST':
        try:
            action = request.form.get('action')
            if action == 'add_member':
                # 添加成员
                team_id = request.form.get('team_id')
                user_id = request.form.get('user_id')

                if UserTeam.query.filter_by(team_id=team_id, user_id=user_id).first():
                    flash('该用户已在团队中！', 'danger')
                else:
                    db.session.add(UserTeam(team_id=team_id, user_id=user_id))
                    flash('成员添加成功！', 'success')

            elif action == 'remove_member':
                # 移除成员
                ut_id = request.form.get('ut_id')
                ut = db.session.get(UserTeam, ut_id)
                if ut:
                    db.session.delete(ut)
                    flash('成员移除成功！', 'success')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 获取当前用户管理的团队
    teams = Team.query.filter_by(leader_id=current_user.id).all()
    if not teams:
        content = '<div class="alert alert-warning">您尚未管理任何团队！</div>'
        return render_base_layout('团队成员管理', content, current_user)

    # 可添加的用户（非管理员、未加入当前团队）
    all_users = User.query.filter(User.role != 'admin').all()

    # 渲染成员管理页面
    member_html = '''
    <h2>团队成员管理</h2>
    '''
    for team in teams:
        # 查询团队现有成员
        members = UserTeam.query.filter_by(team_id=team.id).all()
        # 可添加的用户（排除已加入的）
        available_users = []
        for user in all_users:
            is_in_team = any(ut.user_id == user.id for ut in members)
            if not is_in_team:
                available_users.append(user)

        # 团队卡片
        member_html += f'''
        <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
            <h3>{team.name}</h3>

            <!-- 添加成员 -->
            <div style="margin-bottom:20px;">
                <form method="POST" style="display:flex; gap:10px; align-items:end;">
                    <input type="hidden" name="action" value="add_member">
                    <input type="hidden" name="team_id" value="{team.id}">
                    <div class="form-group" style="flex:1;">
                        <label>添加团队成员</label>
                        <select name="user_id" required>
                            <option value="">请选择用户</option>
        '''
        # 填充可选用户
        for user in available_users:
            member_html += f'<option value="{user.id}">{user.username}（{user.employee_id}）</option>'
        member_html += '''
                        </select>
                    </div>
                    <button type="submit" class="btn">添加</button>
                </form>
            </div>

            <!-- 现有成员 -->
            <div>
                <strong>当前成员：</strong>
                <ul style="margin:10px 0; padding-left:20px;">
        '''
        # 成员列表
        for ut in members:
            user = db.session.get(User, ut.user_id)
            # 负责人不可移除
            if user.id == team.leader_id:
                member_html += f'<li>{user.username}（{user.employee_id}）<span style="color:#999;">（团队负责人）</span></li>'
            else:
                member_html += f'''
                <li>
                    {user.username}（{user.employee_id}）
                    <form method="POST" style="display:inline; margin-left:10px;">
                        <input type="hidden" name="action" value="remove_member">
                        <input type="hidden" name="ut_id" value="{ut.id}">
                        <button type="submit" class="btn" style="padding:2px 8px; font-size:12px;" onclick="return confirm('确定移除？')">移除</button>
                    </form>
                </li>
                '''
        member_html += '''
                </ul>
            </div>
        </div>
        '''

    return render_base_layout('团队成员管理', member_html, current_user)


@app.route('/download')
def download_file():
    """通用文件下载"""
    file_path = request.args.get('path')
    if not file_path or not os.path.exists(file_path):
        flash('文件不存在！', 'danger')
        return redirect(url_for('index'))
    return send_file(file_path, as_attachment=True)


# 1. 期刊论文管理
@app.route('/achievement/journal_paper', methods=['GET', 'POST'])
def journal_paper_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '论文名称', 'type': 'text', 'required': True},
        {'name': 'authors', 'label': '论文作者', 'type': 'text', 'required': True},
        {'name': 'corresponding_authors', 'label': '通讯作者', 'type': 'text'},
        {'name': 'journal_name', 'label': '期刊名称', 'type': 'text', 'required': True},
        {'name': 'inclusion_type_ids', 'label': '收录情况', 'type': 'select_multiple', 'options': []},
        {'name': 'year', 'label': '年', 'type': 'integer'},
        {'name': 'volume', 'label': '卷', 'type': 'text'},
        {'name': 'issue', 'label': '期', 'type': 'text'},
        {'name': 'page_range', 'label': '起止页码', 'type': 'text'},
        {'name': 'doi', 'label': 'DOI', 'type': 'text'},
        {'name': 'publish_year', 'label': '发表年份', 'type': 'integer'},
        {'name': 'publish_date', 'label': '发表日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '论文附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')

    if request.method == 'POST':
        return handle_achievement_submit(JournalPaper, fields_config)

    if action == 'add':
        # 获取收录类型选项
        inclusion_types = InclusionType.query.filter_by(is_active=True).order_by(InclusionType.sort_order).all()
        fields_config[4]['options'] = [t.type_name for t in inclusion_types]
        return render_achievement_form(JournalPaper, '新增期刊论文', fields_config)
    elif action == 'edit':
        # 获取收录类型选项
        inclusion_types = InclusionType.query.filter_by(is_active=True).order_by(InclusionType.sort_order).all()
        fields_config[4]['options'] = [t.type_name for t in inclusion_types]
        item_id = request.args.get('id')
        return render_achievement_form(JournalPaper, '修改期刊论文', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(JournalPaper, item_id)
    # 核心修改：传递时间筛选参数
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(JournalPaper, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(JournalPaper, '期刊论文', fields_config, current_user)
    else:
        return render_achievement_list(JournalPaper, '期刊论文管理', fields_config, current_user)

# 2. 会议论文管理 - 修改导出逻辑
@app.route('/achievement/conference_paper', methods=['GET', 'POST'])
def conference_paper_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '论文名称', 'type': 'text', 'required': True},
        {'name': 'authors', 'label': '论文作者', 'type': 'text', 'required': True},
        {'name': 'corresponding_authors', 'label': '通讯作者', 'type': 'text'},
        {'name': 'conference_name', 'label': '会议名称', 'type': 'text', 'required': True},
        {'name': 'conference_start_date', 'label': '会议开始日期', 'type': 'date'},
        {'name': 'conference_end_date', 'label': '会议结束日期', 'type': 'date'},
        {'name': 'conference_place', 'label': '会议地点', 'type': 'text'},
        {'name': 'page_range', 'label': '起止页码', 'type': 'text'},
        {'name': 'doi', 'label': 'DOI', 'type': 'text'},
        {'name': 'publish_year', 'label': '发表年份', 'type': 'integer'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '论文附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(ConferencePaper, fields_config)

    if action == 'add':
        return render_achievement_form(ConferencePaper, '新增会议论文', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(ConferencePaper, '修改会议论文', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(ConferencePaper, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(ConferencePaper, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(ConferencePaper, '会议论文', fields_config, current_user)
    else:
        return render_achievement_list(ConferencePaper, '会议论文管理', fields_config, current_user)

# 3. 教材管理（复用通用函数）
@app.route('/achievement/textbook', methods=['GET', 'POST'])
def textbook_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心修改：更新字段配置
    fields_config = [
        {'name': 'title', 'label': '教材名称', 'type': 'text', 'required': True},
        {'name': 'textbook_series', 'label': '教材系列', 'type': 'text'},
        {'name': 'chief_editor', 'label': '主编', 'type': 'text'},
        {'name': 'associate_editors', 'label': '副主编', 'type': 'text'},
        {'name': 'editorial_board', 'label': '编委', 'type': 'text'},
        {'name': 'publisher', 'label': '出版社', 'type': 'text'},
        {'name': 'isbn', 'label': 'ISBN', 'type': 'text'},
        {'name': 'cip_number', 'label': 'CIP 核字号', 'type': 'text'},
        {'name': 'publication_year', 'label': '出版年份', 'type': 'integer'},
        {'name': 'publication_month', 'label': '出版月份', 'type': 'integer'},
        {'name': 'publish_date', 'label': '出版日期', 'type': 'date'},
        {'name': 'edition', 'label': '版次', 'type': 'text'},
        {'name': 'word_count', 'label': '字数', 'type': 'text'},
        {'name': 'price', 'label': '定价', 'type': 'text'},
        # 教材级别下拉框（从数据库读取）
        {'name': 'textbook_level_id', 'label': '教材级别', 'type': 'select', 'options': []},
        # 教材类型下拉框
        {'name': 'textbook_type', 'label': '教材类型', 'type': 'select', 'options': [
            '纸质教材', '数字教材'
        ]},
        {'name': 'applicable_majors', 'label': '适用专业', 'type': 'text'},
        {'name': 'remarks', 'label': '备注', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        # 附件字段名更新
        {'name': 'textbook_attachment', 'label': '教材附件', 'type': 'file'}
    ]

    # 处理表单提交（适配新字段）
    if request.method == 'POST':
        return handle_achievement_submit(Textbook, fields_config)

    # 其余逻辑（action 分支）保持不变，仅渲染和列表展示会自动适配新字段
    action = request.args.get('action', 'list')
    if action == 'add':
        # 获取教材级别选项
        levels = TextbookLevel.query.filter_by(is_active=True).order_by(TextbookLevel.sort_order).all()
        fields_config[14]['options'] = [l.level_name for l in levels]
        return render_achievement_form(Textbook, '新增教材', fields_config)
    elif action == 'edit':
        # 获取教材级别选项
        levels = TextbookLevel.query.filter_by(is_active=True).order_by(TextbookLevel.sort_order).all()
        fields_config[14]['options'] = [l.level_name for l in levels]
        item_id = request.args.get('id')
        return render_achievement_form(Textbook, '修改教材', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(Textbook, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(Textbook, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(Textbook, '教材', fields_config, current_user)
    else:
        return render_achievement_list(Textbook, '教材管理', fields_config, current_user)



# 4. 专著管理（复用通用函数）
@app.route('/achievement/monograph', methods=['GET', 'POST'])
def monograph_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '专著名称', 'type': 'text', 'required': True},
        {'name': 'textbook_series', 'label': '专著系列', 'type': 'text'},
        {'name': 'chief_editor', 'label': '主编', 'type': 'text'},
        {'name': 'associate_editors', 'label': '副主编', 'type': 'text'},
        {'name': 'editorial_board', 'label': '编委', 'type': 'text'},
        {'name': 'publisher', 'label': '出版社', 'type': 'text'},
        {'name': 'isbn', 'label': 'ISBN', 'type': 'text'},
        {'name': 'cip_number', 'label': 'CIP 核字号', 'type': 'text'},
        {'name': 'publication_year', 'label': '出版年份', 'type': 'integer'},
        {'name': 'publication_month', 'label': '出版月份', 'type': 'integer'},
        {'name': 'publish_date', 'label': '出版日期', 'type': 'date'},
        {'name': 'edition', 'label': '版次', 'type': 'text'},
        {'name': 'word_count', 'label': '字数', 'type': 'text'},
        {'name': 'price', 'label': '定价', 'type': 'text'},
        # 专著类型下拉框
        {'name': 'monograph_type', 'label': '专著类型', 'type': 'select', 'options': [
            '学术专著', '技术专著', '科普著作', '其它'
        ]},
        {'name': 'applicable_majors', 'label': '适用专业', 'type': 'text'},
        {'name': 'remarks', 'label': '备注', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        # 附件字段名更新
        {'name': 'monograph_attachment', 'label': '专著附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(Monograph, fields_config)

    # 其余逻辑（action 分支）保持不变，仅渲染和列表展示会自动适配新字段
    action = request.args.get('action', 'list')
    if action == 'add':
        return render_achievement_form(Monograph, '新增专著', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(Monograph, '修改专著', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(Monograph, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(Monograph, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(Monograph, '专著', fields_config, current_user)
    else:
        return render_achievement_list(Monograph, '专著管理', fields_config, current_user)

# 5. 教研教改和课程建设项目管理（复用通用函数）
@app.route('/achievement/teaching_project', methods=['GET', 'POST'])
def teaching_project_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 获取字典表选项
    project_types = [t.type_name for t in ProjectType.query.order_by(ProjectType.sort_order).all()]
    project_levels = [l.level_name for l in ProjectLevel.query.order_by(ProjectLevel.sort_order).all()]
    project_categories = [c.category_name for c in ProjectCategory.query.order_by(ProjectCategory.sort_order).all()]
    project_statuses = [s.status_name for s in ProjectStatus.query.order_by(ProjectStatus.sort_order).all()]
    
    fields_config = [
        {'name': 'title', 'label': '项目名称', 'type': 'text', 'required': True},
        {'name': 'project_code', 'label': '项目编号', 'type': 'text'},
        {'name': 'project_type_id', 'label': '项目类型', 'type': 'select', 'options': project_types},
        {'name': 'project_leader', 'label': '项目负责人', 'type': 'text'},
        {'name': 'project_members', 'label': '项目参与人', 'type': 'text', 'placeholder': '多人请用顿号分隔'},
        {'name': 'approval_department', 'label': '项目批准部门', 'type': 'text'},
        {'name': 'approval_date', 'label': '项目立项时间', 'type': 'date'},
        {'name': 'project_level_id', 'label': '项目级别', 'type': 'select', 'options': project_levels},
        {'name': 'project_category_id', 'label': '项目类别', 'type': 'select', 'options': project_categories},
        {'name': 'funding', 'label': '项目经费（元）', 'type': 'number', 'step': '0.01'},
        {'name': 'start_date', 'label': '项目开始时间', 'type': 'date'},
        {'name': 'end_date', 'label': '项目结束时间', 'type': 'date'},
        {'name': 'project_status_id', 'label': '项目状态', 'type': 'select', 'options': project_statuses},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '项目附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(TeachingProject, fields_config)

    if action == 'add':
        return render_achievement_form(TeachingProject, '新增教研教改和课程建设项目', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(TeachingProject, '修改教研教改和课程建设项目', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(TeachingProject, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(TeachingProject, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(TeachingProject, '教研教改和课程建设项目', fields_config, current_user)
    else:
        return render_achievement_list(TeachingProject, '教研教改和课程建设项目管理', fields_config, current_user)

# 6. 专利管理（复用通用函数）

@app.route('/achievement/patent', methods=['GET', 'POST'])
def patent_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '专利名称', 'type': 'text', 'required': True},
        {'name': 'patent_type_id', 'label': '专利类型', 'type': 'select',
         'options': [(t.id, t.type_name) for t in PatentType.query.order_by(PatentType.sort_order).all()], 'required': True},
        {'name': 'patentee', 'label': '专利权人', 'type': 'text'},
        {'name': 'address', 'label': '地址', 'type': 'text'},
        {'name': 'inventors', 'label': '发明人', 'type': 'text'},
        {'name': 'patent_status_id', 'label': '专利状态', 'type': 'select',
         'options': [(s.id, s.status_name) for s in PatentStatus.query.order_by(PatentStatus.sort_order).all()]},
        {'name': 'patent_number', 'label': '专利号', 'type': 'text'},
        {'name': 'grant_announcement_number', 'label': '授权公告号', 'type': 'text'},
        {'name': 'apply_date', 'label': '专利申请日', 'type': 'date'},
        {'name': 'grant_announcement_date', 'label': '授权公告日', 'type': 'date'},
        {'name': 'applicant_at_apply_date', 'label': '申请日时申请人', 'type': 'text'},
        {'name': 'inventor_at_apply_date', 'label': '申请日时发明人', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(Patent, fields_config)

    if action == 'add':
        return render_achievement_form(Patent, '新增专利', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(Patent, '修改专利', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(Patent, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(Patent, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(Patent, '专利', fields_config, current_user)
    else:
        return render_achievement_list(Patent, '专利管理', fields_config, current_user)


# 7. 软件著作管理（复用通用函数）
@app.route('/achievement/software_copyright', methods=['GET', 'POST'])
def software_copyright_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '软件名称', 'type': 'text', 'required': True},
        {'name': 'copyright_owner', 'label': '著作权人', 'type': 'text'},
        {'name': 'completion_date', 'label': '开发完成日期', 'type': 'date'},
        {'name': 'first_publication_date', 'label': '首次发表日期', 'type': 'date'},
        {'name': 'right_acquisition_method', 'label': '权利取得方式', 'type': 'select', 'options': ['原始取得', '受让取得', '继承取得', '其他']},
        {'name': 'right_scope', 'label': '权利范围', 'type': 'select', 'options': ['全部权利', '部分权利']},
        {'name': 'copyright_number', 'label': '登记号', 'type': 'text'},
        {'name': 'certificate_number', 'label': '证书号', 'type': 'text'},
        {'name': 'register_date', 'label': '登记日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(SoftwareCopyright, fields_config)

    if action == 'add':
        return render_achievement_form(SoftwareCopyright, '新增软件著作', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(SoftwareCopyright, '修改软件著作', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(SoftwareCopyright, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(SoftwareCopyright, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(SoftwareCopyright, '软件著作', fields_config, current_user)
    else:
        return render_achievement_list(SoftwareCopyright, '软件著作管理', fields_config, current_user)


# 8. 教学成果获奖管理（复用通用函数）
@app.route('/achievement/teaching_achievement_award', methods=['GET', 'POST'])
def teaching_achievement_award_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '成果名称', 'type': 'text', 'required': True},
        {'name': 'achievement_type_id', 'label': '教学成果奖类型', 'type': 'select', 'options': []},
        {'name': 'achievement_level_id', 'label': '成果等级', 'type': 'select', 'options': []},
        {'name': 'main_contributors', 'label': '主要完成人', 'type': 'text', 'placeholder': '多人用分号分隔'},
        {'name': 'completing_units', 'label': '成果完成单位', 'type': 'text', 'placeholder': '多个用分号分隔'},
        {'name': 'award_year', 'label': '获奖年度', 'type': 'integer'},
        {'name': 'award_rank_id', 'label': '获奖等级', 'type': 'select', 'options': []},
        {'name': 'certificate_number', 'label': '证书编号', 'type': 'text'},
        {'name': 'awarding_unit', 'label': '颁奖单位', 'type': 'text'},
        {'name': 'award_date', 'label': '获奖日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(TeachingAchievementAward, fields_config)

    if action == 'add':
        achievement_types = TeachingAchievementType.query.filter_by(is_active=True).order_by(TeachingAchievementType.sort_order).all()
        achievement_levels = AchievementLevel.query.filter_by(is_active=True).order_by(AchievementLevel.sort_order).all()
        award_ranks = AwardRank.query.filter_by(is_active=True).order_by(AwardRank.sort_order).all()
        fields_config[1]['options'] = [(t.id, t.type_name) for t in achievement_types]
        fields_config[2]['options'] = [(l.id, l.level_name) for l in achievement_levels]
        fields_config[6]['options'] = [(r.id, r.rank_name) for r in award_ranks]
        return render_achievement_form(TeachingAchievementAward, '新增教学成果获奖', fields_config)
    elif action == 'edit':
        achievement_types = TeachingAchievementType.query.filter_by(is_active=True).order_by(TeachingAchievementType.sort_order).all()
        achievement_levels = AchievementLevel.query.filter_by(is_active=True).order_by(AchievementLevel.sort_order).all()
        award_ranks = AwardRank.query.filter_by(is_active=True).order_by(AwardRank.sort_order).all()
        fields_config[1]['options'] = [(t.id, t.type_name) for t in achievement_types]
        fields_config[2]['options'] = [(l.id, l.level_name) for l in achievement_levels]
        fields_config[6]['options'] = [(r.id, r.rank_name) for r in award_ranks]
        item_id = request.args.get('id')
        return render_achievement_form(TeachingAchievementAward, '修改教学成果获奖', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(TeachingAchievementAward, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(TeachingAchievementAward, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(TeachingAchievementAward, '教学成果获奖', fields_config, current_user)
    else:
        return render_achievement_list(TeachingAchievementAward, '教学成果获奖管理', fields_config, current_user)


# 9. 教学竞赛获奖管理（复用通用函数）
@app.route('/achievement/teaching_competition_award', methods=['GET', 'POST'])
def teaching_competition_award_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '竞赛名称', 'type': 'text', 'required': True},
        {'name': 'award_year', 'label': '获奖年度', 'type': 'text'},
        {'name': 'competition_level_id', 'label': '竞赛等级', 'type': 'select',
         'options_from_model': 'AchievementLevel'},
        {'name': 'award_rank_id', 'label': '获奖等级', 'type': 'select', 'options_from_model': 'AwardRank'},
        {'name': 'winners', 'label': '获奖人', 'type': 'text'},
        {'name': 'winner_unit', 'label': '获奖人所在单位', 'type': 'text'},
        {'name': 'competition_name', 'label': '竞赛主办方', 'type': 'text'},
        {'name': 'award_date', 'label': '获奖日期', 'type': 'date'},
        {'name': 'certificate_number', 'label': '证书编号', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(TeachingCompetitionAward, fields_config)

    if action == 'add':
        return render_achievement_form(TeachingCompetitionAward, '新增教学竞赛获奖', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(TeachingCompetitionAward, '修改教学竞赛获奖', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(TeachingCompetitionAward, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(TeachingCompetitionAward, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(TeachingCompetitionAward, '教学竞赛获奖', fields_config, current_user)
    else:
        return render_achievement_list(TeachingCompetitionAward, '教学竞赛获奖管理', fields_config, current_user)


# 10. 指导学生获奖管理（复用通用函数）
@app.route('/achievement/student_guidance_award', methods=['GET', 'POST'])
def student_guidance_award_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '获奖名称', 'type': 'text', 'required': True},
        {'name': 'award_year', 'label': '获奖年度', 'type': 'text'},
        {'name': 'competition_name', 'label': '竞赛名称', 'type': 'text'},
        {'name': 'competition_level_id', 'label': '竞赛等级', 'type': 'select',
         'options_from_model': 'AchievementLevel'},
        {'name': 'award_rank_id', 'label': '获奖等级', 'type': 'select', 'options_from_model': 'AwardRank'},
        {'name': 'student_name', 'label': '获奖学生', 'type': 'text'},
        {'name': 'project_name', 'label': '获奖项目名称', 'type': 'text'},
        {'name': 'teacher_name', 'label': '指导教师', 'type': 'text'},
        {'name': 'student_unit', 'label': '获奖学生所在单位', 'type': 'text'},
        {'name': 'organizer', 'label': '竞赛主办方', 'type': 'text'},
        {'name': 'certificate_number', 'label': '证书编号', 'type': 'text'},
        {'name': 'award_date', 'label': '获奖日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(StudentGuidanceAward, fields_config)

    if action == 'add':
        return render_achievement_form(StudentGuidanceAward, '新增指导学生获奖', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(StudentGuidanceAward, '修改指导学生获奖', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(StudentGuidanceAward, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(StudentGuidanceAward, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(StudentGuidanceAward, '指导学生获奖', fields_config, current_user)
    else:
        return render_achievement_list(StudentGuidanceAward, '指导学生获奖管理', fields_config, current_user)


@app.route('/team/member_achievements')
def member_achievements():
    """团队负责人-查看成员具体成果详情"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 获取筛选参数
    member_id = request.args.get('member_id')
    achievement_type = request.args.get('type', 'all')

    # 获取当前用户管理的团队ID
    managed_teams = Team.query.filter_by(leader_id=current_user.id).all()
    managed_team_ids = [str(t.id) for t in managed_teams]

    # 获取团队信息
    teams = Team.query.filter_by(leader_id=current_user.id).all()
    team_ids = [t.id for t in teams]
    team_user_ids = [ut.user_id for ut in UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()]

    if not team_user_ids:
        content = '<div class="alert alert-warning">暂无团队成员数据！</div>'
        return render_base_layout('团队成员成果详情', content, current_user)

    # 获取团队成员列表
    team_members = User.query.filter(User.id.in_(team_user_ids)).all()

    # 成果类型映射
    type_mapping = {
        'all': '所有成果',
        'journal_paper': '期刊论文',
        'conference_paper': '会议论文',
        'textbook': '教材',
        'monograph': '专著',
        'teaching_project': '教研教改和课程建设项目',
        'patent': '专利',
        'software_copyright': '软件著作',
        'teaching_achievement_award': '教学成果获奖',
        'teaching_competition_award': '教学竞赛获奖',
        'student_guidance_award': '指导学生获奖'
    }

    model_mapping = {
        'journal_paper': JournalPaper,
        'conference_paper': ConferencePaper,
        'textbook': Textbook,
        'monograph': Monograph,
        'teaching_project': TeachingProject,
        'patent': Patent,
        'software_copyright': SoftwareCopyright,
        'teaching_achievement_award': TeachingAchievementAward,
        'teaching_competition_award': TeachingCompetitionAward,
        'student_guidance_award': StudentGuidanceAward
    }

    # 生成筛选表单
    filter_html = f'''
    <div style="margin-bottom:20px; padding:20px; background:#f5f7fa; border-radius:8px;">
        <form method="GET">
            <div class="form-row" style="display:flex; gap:20px; margin-bottom:10px;">
                <div class="form-group" style="flex:1;">
                    <label>选择团队成员</label>
                    <select name="member_id" required onchange="this.form.submit()">
                        <option value="">全部成员</option>
    '''
    for member in team_members:
        selected = 'selected' if str(member.id) == member_id else ''
        filter_html += f'<option value="{member.id}" {selected}>{member.username}（{member.employee_id}）</option>'

    filter_html += f'''
                    </select>
                </div>
                <div class="form-group" style="flex:1;">
                    <label>成果类型</label>
                    <select name="type" onchange="this.form.submit()">
    '''
    for type_key, type_name in type_mapping.items():
        selected = 'selected' if type_key == achievement_type else ''
        filter_html += f'<option value="{type_key}" {selected}>{type_name}</option>'

    filter_html += '''
                    </select>
                </div>
            </div>
        </form>
    </div>
    '''

    # 查询成果数据（核心：仅显示公开给当前团队的成果）
    achievements = []

    def filter_public_achievements(query):
        """过滤出公开给当前团队的成果"""
        or_conditions = []
        for team_id in managed_team_ids:
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )
        return query.filter(or_(*or_conditions))

    if member_id and achievement_type != 'all' and achievement_type in model_mapping:
        # 筛选指定成员的指定类型成果（仅公开给当前团队的）
        model = model_mapping[achievement_type]
        query = model.query.filter_by(user_id=member_id)
        query = filter_public_achievements(query)
        achievements = query.order_by(model.update_time.desc()).all()

    elif member_id and achievement_type == 'all':
        # 筛选指定成员的所有成果（仅公开给当前团队的）
        for model in model_mapping.values():
            query = model.query.filter_by(user_id=member_id)
            query = filter_public_achievements(query)
            items = query.order_by(model.update_time.desc()).all()
            for item in items:
                item.type_name = [k for k, v in model_mapping.items() if v == model][0]
                achievements.append(item)

    elif achievement_type != 'all' and achievement_type in model_mapping:
        # 筛选所有成员的指定类型成果（仅公开给当前团队的）
        model = model_mapping[achievement_type]
        query = model.query.filter(model.user_id.in_(team_user_ids))
        query = filter_public_achievements(query)
        achievements = query.order_by(model.update_time.desc()).all()

    else:
        # 所有成果（仅公开给当前团队的）
        for model in model_mapping.values():
            query = model.query.filter(model.user_id.in_(team_user_ids))
            query = filter_public_achievements(query)
            items = query.order_by(model.update_time.desc()).all()
            for item in items:
                item.type_name = [k for k, v in model_mapping.items() if v == model][0]
                achievements.append(item)

    # 生成成果列表
    list_html = '''
    <table style="width:100%; border-collapse:collapse;">
        <thead>
            <tr style="background:#e9ecef;">
                <th style="padding:10px; border:1px solid #dee2e6;">成果类型</th>
                <th style="padding:10px; border:1px solid #dee2e6;">成果名称</th>
                <th style="padding:10px; border:1px solid #dee2e6;">所属成员</th>
                <th style="padding:10px; border:1px solid #dee2e6;">创建时间</th>
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    if not achievements:
        list_html += '''
        <tr>
            <td colspan="5" style="padding:20px; text-align:center; border:1px solid #dee2e6;">暂无成果数据（仅显示成员公开给本团队的成果）</td>
        </tr>
        '''
    else:
        for item in achievements:
            # 获取成果类型名称
            if hasattr(item, 'type_name'):
                type_name = type_mapping.get(item.type_name, '未知类型')
            else:
                type_name = type_mapping.get(achievement_type, '未知类型')

            # 获取所属成员
            member = User.query.get(item.user_id)
            member_name = f'{member.username}（{member.employee_id}）' if member else '未知成员'

            # 创建时间
            create_time = item.create_time.strftime('%Y-%m-%d %H:%M') if hasattr(item, 'create_time') else ''

            # 查看详情链接
            detail_link = f'/achievement/{achievement_type if achievement_type != "all" else item.type_name}?action=edit&id={item.id}'

            list_html += f'''
            <tr>
                <td style="padding:10px; border:1px solid #dee2e6;">{type_name}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">{getattr(item, 'title', '无名称')}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">{member_name}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">{create_time}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">
                    <a href="{detail_link}" class="btn" style="padding:5px 10px; font-size:12px;">查看详情</a>
                </td>
            </tr>
            '''

    list_html += '''
        </tbody>
    </table>
    '''

    content = filter_html + list_html
    return render_base_layout('团队成员成果详情', content, current_user)


# ---------------------- 多维度统计 + 图表展示 ----------------------
@app.route('/stats/dashboard')
def stats_dashboard():
    """多维度统计仪表盘（个人/团队）- 移除年度成果趋势"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 确定统计范围
    if current_user.role == 'team_leader':
        # 团队负责人：统计团队数据
        teams = Team.query.filter_by(leader_id=current_user.id).all()
        team_ids = [t.id for t in teams]
        user_ids = [ut.user_id for ut in UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()] + [
            current_user.id]
        stats_scope = '团队'
    else:
        # 普通教师：统计个人数据
        user_ids = [current_user.id]
        stats_scope = '个人'

    def count_user_achievements(model_class):
        """统计用户参与的成果数量"""
        # 统计直接拥有的成果数量（user_id 在 user_ids 中）
        count = model_class.query.filter(model_class.user_id.in_(user_ids)).count()
        return count

    # 成果类型列表（移除作者关联表）
    achievement_types = [
        ('期刊论文', JournalPaper),
        ('会议论文', ConferencePaper),
        ('教材', Textbook),
        ('专著', Monograph),
        ('教研教改和课程建设项目', TeachingProject),
        ('专利', Patent),
        ('软著', SoftwareCopyright),
        ('教学成果获奖', TeachingAchievementAward),
        ('教学竞赛获奖', TeachingCompetitionAward),
        ('指导学生获奖', StudentGuidanceAward)
    ]

    type_stats = []
    total_count = 0

    for name, model in achievement_types:
        count = count_user_achievements(model)
        type_stats.append({'name': name, 'value': count})
        total_count += count


    # 准备饼图数据
    type_labels = [item['name'] for item in type_stats]
    type_values = [item['value'] for item in type_stats]

    # 渲染统计仪表盘（仅保留成果类型分布饼图）
    content = f'''
    <h2>{stats_scope}成果统计仪表盘</h2>
    <div style="margin-bottom:30px; font-size:18px;">
        成果总数：<strong style="color:#3498db; font-size:24px;">{total_count}</strong> 项
    </div>

    <!-- 成果类型分布（饼图） -->
    <div style="margin-bottom:40px; height:400px;">
        <h3 style="margin-bottom:10px;">成果类型分布</h3>
        <canvas id="typeChart"></canvas>
    </div>

    <!-- 引入Chart.js -->
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script>
        // 饼图 - 成果类型分布
        const typeCtx = document.getElementById('typeChart').getContext('2d');
        new Chart(typeCtx, {{
            type: 'pie',
            data: {{
                labels: {json.dumps(type_labels)},
                datasets: [{{
                    label: '成果数量',
                    data: {json.dumps(type_values)},
                    backgroundColor: [
                        '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
                        '#C9CBCF', '#FF9F40', '#FFCD56', '#45B7D1', '#66AA00'
                    ],
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{
                        position: 'right',
                    }}
                }}
            }}
        }});
    </script>
    '''

    return render_base_layout(f'{stats_scope}成果统计仪表盘', content, current_user)

    # ---------------------- 团队列表与创建功能 ----------------------


@app.route('/team/list', methods=['GET', 'POST'])
def team_list():
    """团队列表（查看所有团队）+ 创建团队按钮"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 处理创建团队请求（仅管理员/团队负责人可创建）
    if request.method == 'POST':
        # 权限控制：仅管理员或团队负责人可创建团队
        if current_user.role not in ['admin', 'team_leader']:
            flash('无创建团队权限！', 'danger')
            return redirect(url_for('team_list'))

        team_name = request.form.get('team_name', '').strip()
        if not team_name:
            flash('团队名称不能为空！', 'danger')
            return redirect(url_for('team_list'))

        # 检查团队名称是否重复
        if Team.query.filter_by(name=team_name).first():
            flash('团队名称已存在！', 'danger')
            return redirect(url_for('team_list'))

        # 创建团队（负责人为当前用户）
        try:
            new_team = Team(
                name=team_name,
                leader_id=current_user.id
            )
            db.session.add(new_team)
            # 先提交获取team_id
            db.session.flush()  # 关键：先刷新会话，生成new_team.id但不提交事务
            # 再创建用户-团队关联
            db.session.add(UserTeam(user_id=current_user.id, team_id=new_team.id))
            db.session.commit()
            flash(f'团队「{team_name}」创建成功！', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'创建团队失败：{str(e)}', 'danger')

        return redirect(url_for('team_list'))

    # 根据用户角色筛选团队列表
    if current_user.role == 'admin':
        # 管理员：查看所有团队
        teams = Team.query.order_by(Team.create_time.desc()).all()
    elif current_user.role == 'team_leader':
        # 团队负责人：查看自己管理的团队
        teams = Team.query.filter_by(leader_id=current_user.id).order_by(Team.create_time.desc()).all()
    else:
        # 普通教师：查看自己加入的团队
        user_teams = UserTeam.query.filter_by(user_id=current_user.id).all()
        team_ids = [ut.team_id for ut in user_teams]
        teams = Team.query.filter(Team.id.in_(team_ids)).order_by(Team.create_time.desc()).all()

    # 构建团队列表HTML
    team_list_html = f'''
        <h2>团队管理</h2>

        <!-- 创建团队按钮 + 表单 -->
        <div style="margin-bottom:30px;">
            <button onclick="toggleCreateForm()" class="btn" style="background:#27ae60;">📝 创建新团队</button>

            <!-- 创建团队表单（默认隐藏） -->
            <div id="createTeamForm" style="display:none; margin-top:20px; padding:20px; border:1px solid #eee; border-radius:8px;">
                <h3 style="margin-bottom:20px;">创建新团队</h3>
                <form method="POST">
                    <div class="form-group">
                        <label>团队名称 <span class="required" style="color:red;">*</span></label>
                        <input type="text" name="team_name" required placeholder="请输入团队名称">
                    </div>
                    <button type="submit" class="btn">确认创建</button>
                    <button type="button" onclick="toggleCreateForm()" class="btn" style="background:#95a5a6; margin-left:10px;">取消</button>
                </form>
            </div>
        </div>

        <!-- 团队列表 -->
        <div style="margin-top:20px;">
            <h3>{"所有团队" if current_user.role == 'admin' else "我的团队"}</h3>
            {f'<div class="alert alert-info">暂无团队数据</div>' if not teams else ''}

            <div style="display:grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap:20px; margin-top:20px;">
        '''

    # 渲染每个团队卡片
    for team in teams:
        leader = User.query.get(team.leader_id)
        leader_name = leader.username if leader else '未知'

        # 获取团队成员数量
        member_count = UserTeam.query.filter_by(team_id=team.id).count()

        # 团队操作按钮
        action_buttons = ''
        if current_user.id == team.leader_id or current_user.role == 'admin':
            action_buttons = f'''
                <a href="/team/manage_members?team_id={team.id}" class="btn" style="padding:5px 10px; font-size:12px; margin-right:5px;">管理成员</a>
                <a href="/team/achievements?team_id={team.id}" class="btn" style="padding:5px 10px; font-size:12px; margin-right:5px;">成果统计</a>
                <a href="/team/member_achievements?team_id={team.id}" class="btn" style="padding:5px 10px; font-size:12px;">成员成果</a>
                '''
        elif current_user.role == 'teacher':
            action_buttons = '<span style="color:#7f8c8d;">普通成员（仅查看）</span>'

        # 团队卡片
        team_list_html += f'''
            <div style="border:1px solid #eee; border-radius:8px; padding:20px; background:white; box-shadow:0 2px 5px rgba(0,0,0,0.05);">
                <h4 style="margin-bottom:10px; color:#2c3e50;">{team.name}</h4>
                <p><strong>负责人：</strong>{leader_name}</p>
                <p><strong>创建时间：</strong>{team.create_time.strftime('%Y-%m-%d')}</p>
                <p><strong>成员数量：</strong>{member_count} 人</p>
                <div style="margin-top:15px;">{action_buttons}</div>
            </div>
            '''

    team_list_html += '''
            </div>
        </div>

        <script>
            // 显示/隐藏创建团队表单
            function toggleCreateForm() {
                const form = document.getElementById('createTeamForm');
                form.style.display = form.style.display === 'none' ? 'block' : 'none';
            }
        </script>
        '''

    return render_base_layout('团队列表', team_list_html, current_user)

# ========== 期刊论文智能导入路由 ==========
@app.route('/achievement/journal_paper/import', methods=['GET', 'POST'])
def journal_paper_import():
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('login'))

    zhipu_api_key = get_zhipu_api_key(current_user)
    if not zhipu_api_key:
        content = '''
        <div class="alert alert-danger">
            未配置智谱 AI API Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型 API 配置</a> 配置智谱 AI API Key。
        </div>
        <a href="/achievement/journal_paper" class="btn">返回列表</a>
        '''
        return render_base_layout('期刊论文智能导入', content, current_user)

    # POST 请求：处理爬取和确认
    if request.method == 'POST':
        action = request.form.get('action', '')

        # 用户确认导入
        if action == 'confirm_import':
            paper_indices = request.form.getlist('selected_papers')
            papers_json = request.form.get('papers_data', '[]')

            try:
                papers = json.loads(papers_json)
            except:
                flash('数据格式错误，请重新操作！', 'danger')
                return redirect('/achievement/journal_paper/import')

            if not paper_indices:
                flash('请至少选择一篇论文进行导入！', 'warning')
                return redirect('/achievement/journal_paper/import')

            success_count = 0
            duplicate_count = 0
            imported_titles = []

            for idx in paper_indices:
                try:
                    idx = int(idx)
                    if idx < 0 or idx >= len(papers):
                        continue

                    paper = papers[idx]

                    # 检查数据库中是否已有相同论文名称
                    existing = JournalPaper.query.filter_by(title=paper['论文名称'], user_id=current_user.id).first()
                    if existing:
                        duplicate_count += 1
                        print(f"跳过重复论文：{paper['论文名称']}")
                        continue

                    ai_result = ai_analyze_journal_full(paper['引用格式'], zhipu_api_key)

                    publish_date = None
                    if paper.get('发表日期'):
                        try:
                            publish_date = datetime.strptime(paper['发表日期'], '%Y-%m-%d').date()
                        except:
                            pass

                    journal_paper = JournalPaper(
                        user_id=current_user.id,
                        title=paper['论文名称'],
                        authors=paper['论文作者'],
                        corresponding_authors=paper.get('通讯作者', ''),
                        journal_name=paper['期刊名称'],
                        inclusion_status=paper.get('论文收录情况', ''),
                        year=paper.get('年') or ai_result.get('卷'),
                        volume=paper.get('卷') or ai_result.get('卷'),
                        issue=paper.get('期') or ai_result.get('期'),
                        page_range=paper.get('起止页码') or ai_result.get('起止页码'),
                        doi=paper.get('DOI') or ai_result.get('DOI'),
                        publish_year=paper.get('发表年份') or (ai_result.get('年') if ai_result.get('年') else None),
                        publish_date=publish_date,
                        create_time=datetime.now(),
                        update_time=datetime.now()
                    )
                    db.session.add(journal_paper)
                    db.session.flush()

                    auto_link_contributors(journal_paper, 'journal_paper', paper['论文作者'], current_user.id)

                    success_count += 1
                    imported_titles.append(paper['论文名称'])
                except Exception as e:
                    print(f"导入期刊论文失败：{e}")
                    continue

            db.session.commit()

            msg = f"导入完成！成功导入 {success_count} 篇论文"
            if duplicate_count > 0:
                msg += f"，跳过 {duplicate_count} 篇重复论文"
            msg += "。"

            content = f'''
            <div class="alert alert-success">
                {msg}<br>
            </div>
            <a href="/achievement/journal_paper" class="btn">查看论文列表</a>
            <a href="/achievement/journal_paper/import" class="btn">继续导入</a>
            '''
            return render_base_layout('期刊论文智能导入', content, current_user)

        # 初始爬取请求
        keyword = request.form.get('keyword', '').strip()
        max_papers = request.form.get('max_papers', 3, type=int)
        driver_path = request.form.get('driver_path',
                                       r'C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe')

        if not keyword:
            flash('搜索关键词不能为空！', 'danger')
            return redirect('/achievement/journal_paper/import')

        flash('开始爬取知网数据，请稍候...', 'success')
        papers = crawl_cnki_journal(keyword, max_papers, driver_path)

        if not papers:
            content = '''
            <div class="alert alert-warning">未爬取到任何期刊论文数据！</div>
            <a href="/achievement/journal_paper/import" class="btn">重新导入</a>
            <a href="/achievement/journal_paper" class="btn">返回列表</a>
            '''
            return render_base_layout('期刊论文智能导入', content, current_user)

        # 检查哪些论文已经在数据库中
        for i, paper in enumerate(papers):
            existing = JournalPaper.query.filter_by(title=paper['论文名称'], user_id=current_user.id).first()
            papers[i]['exists_in_db'] = bool(existing)

        # 生成用户确认界面
        papers_json = json.dumps(papers, ensure_ascii=False)

        table_rows = ''
        for i, paper in enumerate(papers):
            checkbox_disabled = 'disabled' if paper.get('exists_in_db') else ''
            checkbox_checked = '' if paper.get('exists_in_db') else 'checked'
            status_badge = '<span style="color:#e74c3c;">❌ 已存在</span>' if paper.get(
                'exists_in_db') else '<span style="color:#27ae60;">✅ 可导入</span>'

            table_rows += f'''
            <tr>
                <td style="padding:8px; border-bottom:1px solid #ddd;">
                    <input type="checkbox" name="selected_papers" value="{i}" {checkbox_checked} {checkbox_disabled}>
                </td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{status_badge}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;"><strong>{paper['论文名称']}</strong></td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper['论文作者']}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper['期刊名称']}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper.get('年', '')} 年 {paper.get('卷', '')} 卷 {paper.get('期', '')} 期</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper.get('DOI', '')}</td>
            </tr>
            '''

        content = f'''
                <h2>期刊论文智能导入 - 用户确认</h2>
                <div class="alert alert-info">
                    <strong>说明：</strong><br>
                    1. 系统检测到以下论文，请勾选需要导入的论文（默认全选可导入的论文）<br>
                    2. <span style="color:#e74c3c;">❌ 已存在</span> 表示数据库中已有相同名称的论文，不能重复导入<br>
                    3. <span style="color:#27ae60;">✅ 可导入</span> 表示数据库中无重复，可以导入<br>
                    4. 您可以手动取消勾选某些不想导入的论文
                </div>

                <form method="POST">
                    <input type="hidden" name="action" value="confirm_import">
                    <textarea name="papers_data" style="display:none;">{papers_json}</textarea>

                    <table style="width:100%; border-collapse:collapse; margin:20px 0;">
                        <thead>
                            <tr style="background:#f8f9fa;">
                                <th style="padding:10px; border:1px solid #ddd;">选择</th>
                                <th style="padding:10px; border:1px solid #ddd;">状态</th>
                                <th style="padding:10px; border:1px solid #ddd;">论文名称</th>
                                <th style="padding:10px; border:1px solid #ddd;">作者</th>
                                <th style="padding:10px; border:1px solid #ddd;">期刊名称</th>
                                <th style="padding:10px; border:1px solid #ddd;">卷期</th>
                                <th style="padding:10px; border:1px solid #ddd;">DOI</th>
                            </tr>
                        </thead>
                        <tbody>
                            {table_rows}
                        </tbody>
                    </table>

                    <div style="margin:20px 0;">
                        <button type="submit" class="btn" style="background:#27ae60; padding:10px 30px;">✅ 确认导入选中的论文</button>
                        <a href="/achievement/journal_paper/import" class="btn" style="background:#95a5a6; margin-left:10px; padding:10px 30px;">返回重新搜索</a>
                    </div>
                </form>
                '''
        return render_base_layout('期刊论文智能导入', content, current_user)

    # GET 请求：显示导入表单
    form_html = '''
    <h2>期刊论文智能导入（知网爬取）</h2>
    <form method="POST">
        <div class="form-group">
            <label>搜索作者 <span style="color:red;">*</span></label>
            <input type="text" name="keyword" placeholder="作者名" required>
        </div>
        <div class="form-group">
            <label>最大导入数量</label>
            <input type="number" name="max_papers" value="3" min="1" max="10">
        </div>
        <div class="form-group">
            <name="driver_path" value="C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe" style="width:100%;">
        </div>
        <div class="form-group">
            <button type="submit" class="btn" style="background:#27ae60;">开始智能导入</button>
            <a href="/achievement/journal_paper" class="btn" style="background:#95a5a6; margin-left:10px;">取消</a>
        </div>
    </form>
    '''
    return render_base_layout('期刊论文智能导入', form_html, current_user)


# ========== 会议论文智能导入路由 ==========
@app.route('/achievement/conference_paper/import', methods=['GET', 'POST'])
def conference_paper_import():
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('login'))

    zhipu_api_key = get_zhipu_api_key(current_user)
    if not zhipu_api_key:
        content = '''
        <div class="alert alert-danger">
            未配置智谱 AI API Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型 API 配置</a> 配置智谱 AI API Key。
        </div>
        <a href="/achievement/conference_paper" class="btn">返回列表</a>
        '''
        return render_base_layout('会议论文智能导入', content, current_user)

    # POST 请求：处理爬取和确认
    if request.method == 'POST':
        action = request.form.get('action', '')

        # 用户确认导入
        if action == 'confirm_import':
            paper_indices = request.form.getlist('selected_papers')
            papers_json = request.form.get('papers_data', '[]')

            try:
                papers = json.loads(papers_json)
            except:
                flash('数据格式错误，请重新操作！', 'danger')
                return redirect('/achievement/conference_paper/import')

            if not paper_indices:
                flash('请至少选择一篇论文进行导入！', 'warning')
                return redirect('/achievement/conference_paper/import')

            success_count = 0
            duplicate_count = 0

            for idx in paper_indices:
                try:
                    idx = int(idx)
                    if idx < 0 or idx >= len(papers):
                        continue

                    paper = papers[idx]

                    # 检查数据库中是否已有相同论文名称
                    existing = ConferencePaper.query.filter_by(title=paper['论文名称']).first()
                    if existing:
                        duplicate_count += 1
                        print(f"跳过重复论文：{paper['论文名称']}")
                        continue

                    ai_result = ai_analyze_citation(paper['引用格式'], zhipu_api_key)

                    conference_start_date = None
                    conference_end_date = None
                    conference_time_str = None

                    if paper.get('会议时间'):
                        time_text = paper['会议时间'].strip()
                        conference_time_str = time_text

                        import re
                        date_range_pattern = r'(\\d{{4}})[.\\-/](\\d{{1,2}})[.\\-/](\\d{{1,2}})\\s*[-–—]\\s*(\\d{{4}})[.\\-/](\\d{{1,2}})[.\\-/](\\d{{1,2}})'
                        match = re.search(date_range_pattern, time_text)

                        if match:
                            start_year, start_month, start_day = match.group(1), match.group(2), match.group(3)
                            end_year, end_month, end_day = match.group(4), match.group(5), match.group(6)
                            try:
                                conference_start_date = datetime(int(start_year), int(start_month),
                                                                 int(start_day)).date()
                                conference_end_date = datetime(int(end_year), int(end_month), int(end_day)).date()
                            except:
                                pass
                        else:
                            single_date_pattern = r'(\\d{{4}})[.\\-/](\\d{{1,2}})[.\\-/](\\d{{1,2}})'
                            match = re.search(single_date_pattern, time_text)
                            if match:
                                try:
                                    year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
                                    conference_start_date = datetime(year, month, day).date()
                                    conference_end_date = conference_start_date
                                except:
                                    pass

                    conference_paper = ConferencePaper(
                        user_id=current_user.id,
                        title=paper['论文名称'],
                        authors=paper['论文作者'],
                        corresponding_authors=paper.get('通讯作者', ''),
                        conference_name=paper['会议名称'],
                        conference_time=conference_time_str,
                        conference_start_date=conference_start_date,
                        conference_end_date=conference_end_date,
                        conference_place=paper.get('会议地点'),
                        page_range=paper.get('起止页码') or ai_result.get('起止页码'),
                        doi=paper.get('DOI') or ai_result.get('DOI'),
                        publish_year=paper.get('发表年份') or (ai_result.get('年') if ai_result.get('年') else None),
                        create_time=datetime.now(),
                        update_time=datetime.now()
                    )
                    db.session.add(conference_paper)
                    db.session.flush()

                    auto_link_contributors(conference_paper, 'conference_paper', paper['论文作者'], current_user.id)

                    success_count += 1
                except Exception as e:
                    print(f"导入会议论文失败：{e}")
                    continue

            db.session.commit()

            msg = f"导入完成！成功导入 {success_count} 篇论文"
            if duplicate_count > 0:
                msg += f"，跳过 {duplicate_count} 篇重复论文"
            msg += "。"

            content = f'''
            <div class="alert alert-success">
                {msg}<br>
            </div>
            <a href="/achievement/conference_paper" class="btn">查看论文列表</a>
            <a href="/achievement/conference_paper/import" class="btn">继续导入</a>
            '''
            return render_base_layout('会议论文智能导入', content, current_user)

        # 初始爬取请求
        keyword = request.form.get('keyword', '').strip()
        max_papers = request.form.get('max_papers', 3, type=int)
        driver_path = request.form.get('driver_path',
                                       r'C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe')

        if not keyword:
            flash('搜索关键词不能为空！', 'danger')
            return redirect('/achievement/conference_paper/import')

        flash('开始爬取知网数据，请稍候...', 'success')
        papers = crawl_cnki_conference(keyword, max_papers, driver_path)

        if not papers:
            content = '''
            <div class="alert alert-warning">未爬取到任何会议论文数据！</div>
            <a href="/achievement/conference_paper/import" class="btn">重新导入</a>
            <a href="/achievement/conference_paper" class="btn">返回列表</a>
            '''
            return render_base_layout('会议论文智能导入', content, current_user)

        # 检查哪些论文已经在数据库中
        for i, paper in enumerate(papers):
            existing = ConferencePaper.query.filter_by(title=paper['论文名称']).first()
            papers[i]['exists_in_db'] = bool(existing)

        # 生成用户确认界面
        papers_json = json.dumps(papers, ensure_ascii=False)

        table_rows = ''
        for i, paper in enumerate(papers):
            checkbox_disabled = 'disabled' if paper.get('exists_in_db') else ''
            checkbox_checked = '' if paper.get('exists_in_db') else 'checked'
            status_badge = '<span style="color:#e74c3c;">❌ 已存在</span>' if paper.get(
                'exists_in_db') else '<span style="color:#27ae60;">✅ 可导入</span>'

            table_rows += f'''
            <tr>
                <td style="padding:8px; border-bottom:1px solid #ddd;">
                    <input type="checkbox" name="selected_papers" value="{i}" {checkbox_checked} {checkbox_disabled}>
                </td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{status_badge}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;"><strong>{paper['论文名称']}</strong></td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper['论文作者']}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper['会议名称']}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper.get('会议时间', '')}</td>
                <td style="padding:8px; border-bottom:1px solid #ddd;">{paper.get('DOI', '')}</td>
            </tr>
            '''

        content = f'''
                <h2>会议论文智能导入 - 用户确认</h2>
                <div class="alert alert-info">
                    <strong>说明：</strong><br>
                    1. 系统检测到以下论文，请勾选需要导入的论文（默认全选可导入的论文）<br>
                    2. <span style="color:#e74c3c;">❌ 已存在</span> 表示数据库中已有相同名称的论文，不能重复导入<br>
                    3. <span style="color:#27ae60;">✅ 可导入</span> 表示数据库中无重复，可以导入<br>
                    4. 您可以手动取消勾选某些不想导入的论文
                </div>

                <form method="POST">
                    <input type="hidden" name="action" value="confirm_import">
                    <textarea name="papers_data" style="display:none;">{papers_json}</textarea>

                    <table style="width:100%; border-collapse:collapse; margin:20px 0;">
                        <thead>
                            <tr style="background:#f8f9fa;">
                                <th style="padding:10px; border:1px solid #ddd;">选择</th>
                                <th style="padding:10px; border:1px solid #ddd;">状态</th>
                                <th style="padding:10px; border:1px solid #ddd;">论文名称</th>
                                <th style="padding:10px; border:1px solid #ddd;">作者</th>
                                <th style="padding:10px; border:1px solid #ddd;">会议名称</th>
                                <th style="padding:10px; border:1px solid #ddd;">会议时间</th>
                                <th style="padding:10px; border:1px solid #ddd;">DOI</th>
                            </tr>
                        </thead>
                        <tbody>
                            {table_rows}
                        </tbody>
                    </table>

                    <div style="margin:20px 0;">
                        <button type="submit" class="btn" style="background:#27ae60; padding:10px 30px;">✅ 确认导入选中的论文</button>
                        <a href="/achievement/conference_paper/import" class="btn" style="background:#95a5a6; margin-left:10px; padding:10px 30px;">返回重新搜索</a>
                    </div>
                </form>
                '''
        return render_base_layout('会议论文智能导入', content, current_user)

    # GET 请求：显示导入表单
    form_html = '''
    <h2>会议论文智能导入（知网爬取）</h2>
    <form method="POST">
        <div class="form-group">
            <label>搜索作者 <span style="color:red;">*</span></label>
            <input type="text" name="keyword" placeholder="作者名" required>
        </div>
        <div class="form-group">
            <label>最大导入数量</label>
            <input type="number" name="max_papers" value="3" min="1" max="10">
        </div>
        <div class="form-group">
            <name="driver_path" value="C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe" style="width:100%;">
        </div>
        <div class="form-group">
            <button type="submit" class="btn" style="background:#27ae60;">开始智能导入</button>
            <a href="/achievement/conference_paper" class="btn" style="background:#95a5a6; margin-left:10px;">取消</a>
        </div>
    </form>
    '''
    return render_base_layout('会议论文智能导入', form_html, current_user)


@app.route('/achievement/ocr_import', methods=['GET', 'POST'])
def ocr_import():
    """OCR 智能导入成果（支持图片/PDF，识别结果可编辑）"""
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    # 检查百度 API 配置（OCR 必需）
    api_config = current_user.get_api_config()
    if not api_config.get('baidu', {}).get('api_key') or not api_config.get('baidu', {}).get('secret_key'):
        content = '''
        <div class="alert alert-danger">
            未配置百度 API Key/Secret Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型 API 配置</a> 配置百度 API。
        </div>
        <a href="/" class="btn">返回首页</a>
        '''
        return render_base_layout('OCR 智能导入', content, current_user)

    # 检查智谱 API 配置（AI 分析可选）
    zhipu_configured = bool(get_zhipu_api_key(current_user))

    # GET 请求：显示上传表单
    if request.method != 'POST':
        ai_tip = ""
        if not zhipu_configured:
            ai_tip = '''
            <div class="alert alert-warning">
                未配置智谱 AI API Key，将使用基础 OCR 识别（无 AI 智能分析）<br>
                配置地址：<a href="/user/api_config">个人设置 > 大模型 API 配置</a>
            </div>
            '''

        form_html = f'''
            <h2>OCR 智能导入成果（支持图片/PDF）</h2>
            {ai_tip}
            <form method="POST" enctype="multipart/form-data">
                <div class="form-group">
                    <label>上传成果图片/PDF <span style="color:red;">*</span></label>
                    <input type="file" name="image_file" accept="image/*,.pdf" required>
                    <p style="margin-top:5px; color:#666;">
                        支持格式：JPG/PNG/GIF/PDF，PDF 文件会自动转换为图片逐页识别<br>
                        <strong>文件大小限制：100MB</strong>，过大的 PDF 建议先拆分后上传
                    </p>
                </div>
                <button type="submit" class="btn" style="background:#27ae60;">开始识别</button>
                <a href="/" class="btn" style="background:#95a5a6; margin-left:10px;">取消</a>
            </form>
            '''
        return render_base_layout('OCR 智能导入', form_html, current_user)

    # POST 请求：处理文件上传和识别
    success = False
    ocr_text = ""
    temp_images = []
    achievement_info = {}
    msg = ''

    if request.method == 'POST':
        # 处理文件上传
        if 'image_file' not in request.files:
            flash('请上传图片/PDF 文件！', 'danger')
            return redirect(request.url)

        file = request.files['image_file']
        if file.filename == '':
            flash('请选择图片/PDF 文件！', 'danger')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            # 保存上传文件到临时目录
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_temp')
            if not os.path.exists(temp_path):
                os.makedirs(temp_path)

            original_filename = file.filename

            if '.' not in original_filename:
                flash('文件格式错误，请重新上传！', 'danger')
                return redirect(request.url)

            file_ext = original_filename.rsplit('.', 1)[1].lower()
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{timestamp}.{file_ext}"
            upload_path = os.path.join(temp_path, filename)
            file.save(upload_path)

            try:
                # PDF 处理逻辑
                if file_ext == 'pdf':
                    try:
                        temp_images = pdf_to_images(upload_path, temp_path)
                    except Exception as e:
                        flash(f'PDF 转换失败：{str(e)}<br>请安装 poppler 并配置路径！', 'danger')
                        if os.path.exists(upload_path):
                            os.remove(upload_path)
                        content = f'''
                        <a href="/achievement/ocr_import" class="btn">重新上传</a>
                        '''
                        return render_base_layout('OCR 智能导入', content, current_user)

                    for idx, img_path in enumerate(temp_images):
                        page_text, err = baidu_ocr_recognize(img_path, current_user)
                        if err:
                            flash(f'第{idx + 1}页识别失败：{err}', 'warning')
                            continue
                        ocr_text += f"\n=== 第{idx + 1}页 ===\n{page_text}"
                else:
                    ocr_text, err = baidu_ocr_recognize(upload_path, current_user)
                    if err:
                        flash(f'OCR 识别失败：{err}', 'danger')
                        if os.path.exists(upload_path):
                            os.remove(upload_path)
                        return redirect(request.url)

                if not ocr_text.strip():
                    flash('未识别到任何文本！', 'warning')
                    if os.path.exists(upload_path):
                        os.remove(upload_path)
                    return redirect(request.url)

                # AI 分析（如果配置了智谱 API）
                ai_info = {}
                if zhipu_configured:
                    ai_info = ai_analyze_achievement_text(ocr_text, get_zhipu_api_key(current_user))
                    achievement_info = {
                        'type_name': ai_info.get('type_name'),
                        'title': ai_info.get('title'),
                        'raw_text': ocr_text,
                        'confidence': ai_info.get('confidence'),
                        'ai_data': ai_info
                    }
                else:
                    achievement_info = extract_achievement_info(ocr_text)
                    achievement_info['raw_text'] = ocr_text

                # ========== 核心修改：显示可编辑文本框，不直接导入 ==========
                # 生成成果类型选项
                type_options = ''
                for type_name in achievement_rules.keys():
                    selected = 'selected' if type_name == achievement_info.get('type_name', '') else ''
                    type_options += f'<option value="{type_name}" {selected}>{type_name}</option>'

                # 生成详细字段表单（根据 AI 分析结果）
                detail_fields_html = ''
                if zhipu_configured and ai_info:
                    # 根据成果类型生成对应所有字段
                    field_mapping = {
                        '期刊论文': [
                            ('title', '论文名称', 'text', ai_info.get('title', '')),
                            ('authors', '作者', 'text', ai_info.get('authors', '')),
                            ('corresponding_authors', '通讯作者', 'text', ai_info.get('corresponding_authors', '')),
                            ('journal_name', '期刊名称', 'text', ai_info.get('journal_name', '')),
                            ('inclusion_status', '收录情况', 'text', ai_info.get('inclusion_status', '')),
                            ('year', '年', 'number', ai_info.get('year', '')),
                            ('volume', '卷', 'text', ai_info.get('volume', '')),
                            ('issue', '期', 'text', ai_info.get('issue', '')),
                            ('page_range', '起止页码', 'text', ai_info.get('page_range', '')),
                            ('doi', 'DOI', 'text', ai_info.get('doi', '')),
                            ('publish_year', '发表年份', 'number', ai_info.get('publish_year', '')),
                            ('publish_date', '发表日期', 'date', ai_info.get('publish_date', '')),
                        ],
                        '会议论文': [
                            ('title', '论文名称', 'text', ai_info.get('title', '')),
                            ('authors', '作者', 'text', ai_info.get('authors', '')),
                            ('corresponding_authors', '通讯作者', 'text', ai_info.get('corresponding_authors', '')),
                            ('conference_name', '会议名称', 'text', ai_info.get('conference_name', '')),
                            ('conference_time', '会议时间', 'text', ai_info.get('conference_time', '')),
                            ('conference_place', '会议地点', 'text', ai_info.get('conference_place', '')),
                            ('page_range', '起止页码', 'text', ai_info.get('page_range', '')),
                            ('doi', 'DOI', 'text', ai_info.get('doi', '')),
                            ('publish_year', '发表年份', 'number', ai_info.get('publish_year', '')),
                        ],
                        '教材': [
                            ('title', '教材名称', 'text', ai_info.get('title', '')),
                            ('textbook_series', '教材系列', 'text', ai_info.get('textbook_series', '')),
                            ('chief_editor', '主编', 'text', ai_info.get('chief_editor', '')),
                            ('associate_editors', '副主编', 'text', ai_info.get('associate_editors', '')),
                            ('editorial_board', '编委', 'text', ai_info.get('editorial_board', '')),
                            ('publisher', '出版社', 'text', ai_info.get('publisher', '')),
                            ('isbn', 'ISBN', 'text', ai_info.get('isbn', '')),
                            ('cip_number', 'CIP 核字号', 'text', ai_info.get('cip_number', '')),
                            ('publication_year', '出版年份', 'number', ai_info.get('publication_year', '')),
                            ('publication_month', '出版月份', 'number', ai_info.get('publication_month', '')),
                            ('edition', '版次', 'text', ai_info.get('edition', '')),
                            ('word_count', '字数', 'text', ai_info.get('word_count', '')),
                            ('price', '定价', 'text', ai_info.get('price', '')),
                            ('textbook_level', '教材级别', 'text', ai_info.get('textbook_level', '')),
                            ('textbook_type', '教材类型', 'text', ai_info.get('textbook_type', '')),
                            ('applicable_majors', '适用专业', 'text', ai_info.get('applicable_majors', '')),
                            ('remarks', '备注', 'textarea', ai_info.get('remarks', '')),
                        ],
                        '专著': [
                            ('title', '专著名称', 'text', ai_info.get('title', '')),
                            ('textbook_series', '专著系列', 'text', ai_info.get('textbook_series', '')),
                            ('chief_editor', '主编', 'text', ai_info.get('chief_editor', '')),
                            ('associate_editors', '副主编', 'text', ai_info.get('associate_editors', '')),
                            ('editorial_board', '编委', 'text', ai_info.get('editorial_board', '')),
                            ('publisher', '出版社', 'text', ai_info.get('publisher', '')),
                            ('isbn', 'ISBN', 'text', ai_info.get('isbn', '')),
                            ('cip_number', 'CIP 核字号', 'text', ai_info.get('cip_number', '')),
                            ('publication_year', '出版年份', 'number', ai_info.get('publication_year', '')),
                            ('publication_month', '出版月份', 'number', ai_info.get('publication_month', '')),
                            ('edition', '版次', 'text', ai_info.get('edition', '')),
                            ('word_count', '字数', 'text', ai_info.get('word_count', '')),
                            ('price', '定价', 'text', ai_info.get('price', '')),
                            ('monograph_type', '专著类型', 'text', ai_info.get('monograph_type', '')),
                            ('applicable_majors', '适用专业', 'text', ai_info.get('applicable_majors', '')),
                            ('remarks', '备注', 'textarea', ai_info.get('remarks', '')),
                        ],
                        '发明专利': [
                            ('title', '专利名称', 'text', ai_info.get('title', '')),
                            ('patentee', '专利权人', 'text', ai_info.get('patentee', '')),
                            ('address', '地址', 'text', ai_info.get('address', '')),
                            ('inventors', '发明人', 'text', ai_info.get('inventors', '')),
                            ('patent_number', '专利号', 'text', ai_info.get('patent_number', '')),
                            ('grant_announcement_number', '授权公告号', 'text',
                             ai_info.get('grant_announcement_number', '')),
                            ('apply_date', '申请日', 'date', ai_info.get('apply_date', '')),
                            ('grant_announcement_date', '授权公告日', 'date',
                             ai_info.get('grant_announcement_date', '')),
                            ('applicant_at_apply_date', '申请日时申请人', 'text',
                             ai_info.get('applicant_at_apply_date', '')),
                            ('inventor_at_apply_date', '申请日时发明人', 'text',
                             ai_info.get('inventor_at_apply_date', '')),
                        ],
                        '实用新型专利': [
                            ('title', '专利名称', 'text', ai_info.get('title', '')),
                            ('patentee', '专利权人', 'text', ai_info.get('patentee', '')),
                            ('address', '地址', 'text', ai_info.get('address', '')),
                            ('inventors', '发明人', 'text', ai_info.get('inventors', '')),
                            ('patent_number', '专利号', 'text', ai_info.get('patent_number', '')),
                            ('grant_announcement_number', '授权公告号', 'text',
                             ai_info.get('grant_announcement_number', '')),
                            ('apply_date', '申请日', 'date', ai_info.get('apply_date', '')),
                            ('grant_announcement_date', '授权公告日', 'date',
                             ai_info.get('grant_announcement_date', '')),
                        ],
                        '软著': [
                            ('title', '软件名称', 'text', ai_info.get('title', '')),
                            ('copyright_owner', '著作权人', 'text', ai_info.get('copyright_owner', '')),
                            ('completion_date', '开发完成日期', 'date', ai_info.get('completion_date', '')),
                            ('first_publication_date', '首次发表日期', 'date',
                             ai_info.get('first_publication_date', '')),
                            ('right_acquisition_method', '权利取得方式', 'text',
                             ai_info.get('right_acquisition_method', '')),
                            ('right_scope', '权利范围', 'text', ai_info.get('right_scope', '')),
                            ('copyright_number', '登记号', 'text', ai_info.get('copyright_number', '')),
                            ('certificate_number', '证书号', 'text', ai_info.get('certificate_number', '')),
                            ('register_date', '登记日期', 'date', ai_info.get('register_date', '')),
                        ],
                        '教学成果获奖': [
                            ('title', '成果名称', 'text', ai_info.get('title', '')),
                            ('main_contributors', '主要完成人', 'text', ai_info.get('main_contributors', '')),
                            ('completing_units', '成果完成单位', 'text', ai_info.get('completing_units', '')),
                            ('award_year', '获奖年度', 'text', ai_info.get('award_year', '')),
                            ('certificate_number', '证书编号', 'text', ai_info.get('certificate_number', '')),
                            ('awarding_unit', '颁奖单位', 'text', ai_info.get('awarding_unit', '')),
                            ('award_date', '获奖日期', 'date', ai_info.get('award_date', '')),
                        ],
                        '教学竞赛获奖': [
                            ('title', '竞赛名称', 'text', ai_info.get('title', '')),
                            ('award_year', '获奖年度', 'text', ai_info.get('award_year', '')),
                            ('winners', '获奖人', 'text', ai_info.get('winners', '')),
                            ('winner_unit', '获奖人所在单位', 'text', ai_info.get('winner_unit', '')),
                            ('competition_name', '竞赛主办方', 'text', ai_info.get('competition_name', '')),
                            ('award_date', '获奖日期', 'date', ai_info.get('award_date', '')),
                            ('certificate_number', '证书编号', 'text', ai_info.get('certificate_number', '')),
                        ],
                        '指导学生获奖': [
                            ('title', '获奖名称', 'text', ai_info.get('title', '')),
                            ('award_year', '获奖年度', 'text', ai_info.get('award_year', '')),
                            ('student_name', '获奖学生', 'text', ai_info.get('student_name', '')),
                            ('project_name', '项目名称', 'text', ai_info.get('project_name', '')),
                            ('teacher_name', '指导教师', 'text', ai_info.get('teacher_name', '')),
                            ('competition_name', '竞赛名称', 'text', ai_info.get('competition_name', '')),
                            ('award_date', '获奖日期', 'date', ai_info.get('award_date', '')),
                            ('certificate_number', '证书编号', 'text', ai_info.get('certificate_number', '')),
                        ],
                        '教研教改和课程建设项目': [
                            ('title', '项目名称', 'text', ai_info.get('title', '')),
                            ('project_code', '项目编号', 'text', ai_info.get('project_code', '')),
                            ('project_leader', '项目负责人', 'text', ai_info.get('project_leader', '')),
                            ('project_members', '项目参与人', 'text', ai_info.get('project_members', '')),
                            ('approval_department', '批准部门', 'text', ai_info.get('approval_department', '')),
                            ('approval_date', '立项时间', 'date', ai_info.get('approval_date', '')),
                            ('start_date', '开始时间', 'date', ai_info.get('start_date', '')),
                            ('end_date', '结束时间', 'date', ai_info.get('end_date', '')),
                            ('funding', '经费', 'number', ai_info.get('funding', '')),
                        ],
                    }

                    fields = field_mapping.get(ai_info.get('type_name', ''), [])
                    for field_name, field_label, field_type, field_value in fields:
                        if field_type == 'textarea':
                            detail_fields_html += f'''
                                                                <div class="form-group">
                                                                    <label>{field_label}</label>
                                                                    <textarea name="detail_{field_name}" rows="3" style="width:100%; padding:10px; font-size:13px;" placeholder="请输入{field_label}">{field_value}</textarea>
                                                                </div>
                                                                '''
                        else:
                            detail_fields_html += f'''
                                                                <div class="form-group">
                                                                    <label>{field_label}</label>
                                                                    <input type="{field_type}" name="detail_{field_name}" value="{field_value}" placeholder="请输入{field_label}">
                                                                </div>
                                                                '''

                            # 构建确认导入表单页面
                            # 将 field_mapping 转换为 JSON 传递给前端 JavaScript
                            field_mapping_json = json.dumps(field_mapping, ensure_ascii=False)

                            # 提取 AI 初始值（去掉 type_name、title、confidence 等控制字段）
                            ai_initial_values = {k: v for k, v in ai_info.items() if
                                                 k not in ['type_name', 'title', 'confidence', 'raw_text', 'ai_data']}
                            ai_initial_values_json = json.dumps(ai_initial_values, ensure_ascii=False)

                            form_content = f'''
                                <form method="POST" action="/achievement/ocr_import/confirm" id="confirmForm">
                                    <input type="hidden" name="ocr_text" id="ocr_text" value="{ocr_text.replace('"', '&quot;')}">
                                    <input type="hidden" name="ai_data" id="ai_data" value='{json.dumps(ai_info, ensure_ascii=False) if ai_info else "{}"}'>

                                    <div style="margin-bottom:20px;">
                                        <div class="form-group">
                                            <label>成果类型 <span style="color:red;">*</span></label>
                                            <select name="type_name" id="type_name" onchange="updateDetailFields()" required>
                                                {type_options}
                                            </select>
                                        </div>

                                        <div class="form-group">
                                            <label>成果名称 <span style="color:red;">*</span></label>
                                            <input type="text" name="title" id="title" value="{achievement_info.get('title', '')}" required placeholder="请输入成果名称" onblur="checkDuplicate()">
                                            <div id="duplicateTip" style="margin-top:8px; font-size:14px; padding:10px; border-radius:4px; display:none;"></div>
                                        </div>
                                    </div>

                                    <!-- 详细字段区域 -->
                                    <div id="detailFieldsArea" style="margin-bottom:20px; padding:15px; background:#f8f9fa; border-radius:5px;">
                                        <h5 style="margin-bottom:15px;">📋 详细字段（AI 自动提取，可修改）</h5>
                                        {detail_fields_html if detail_fields_html else '<p style="color:#999;">暂无详细字段，请在下方原始识别文本中提取</p>'}
                                    </div>

                                    <!-- 原始识别文本区域 -->
                                    <div class="form-group">
                                        <label>原始识别文本（可编辑）<span style="color:red;">*</span></label>
                                        <textarea name="raw_text" id="raw_text" rows="15" style="width:100%; padding:10px; font-family:monospace; font-size:13px;" required>{ocr_text}</textarea>
                                        <p style="margin-top:5px; color:#666;">
                                            💡 提示：可直接在此编辑识别文本，或从上方详细字段中补充信息
                                        </p>
                                    </div>

                                    <!-- 置信度提示 -->
                                    {f'<div class="alert alert-warning">AI 识别置信度：<strong>{achievement_info.get("confidence", 0):.2f}</strong>（低于 0.6 建议仔细核对）</div>' if achievement_info.get('confidence', 0) > 0 else ''}

                                    <div style="margin-top:30px;">
                                        <button type="submit" name="action" value="confirm" class="btn" style="background:#27ae60; margin-right:10px;" id="submitBtn">✅ 确认导入</button>
                                        <button type="button" onclick="window.location.href='/achievement/ocr_import'" class="btn" style="background:#95a5a6;">🔄 重新识别</button>
                                        <a href="/" class="btn" style="background:#3498db; margin-left:10px;">🏠 返回首页</a>
                                    </div>
                                </form>

                                <script>
                                    let isDuplicate = false;
                                    
                                    // 存储 AI 提取的初始值
                                    const aiInitialValues = {ai_initial_values_json};
                                    
                                    // 存储字段映射关系
                                    const fieldMapping = {field_mapping_json};
                                    
                                    // 查重函数
                                    async function checkDuplicate() {{
                                        const title = document.getElementById('title').value.trim();
                                        const typeName = document.getElementById('type_name').value;
                                        const tipDiv = document.getElementById('duplicateTip');
                                        const submitBtn = document.getElementById('submitBtn');
                                        
                                        if (!title) {{
                                            tipDiv.style.display = 'none';
                                            tipDiv.innerHTML = '';
                                            isDuplicate = false;
                                            return;
                                        }}
                                        
                                        tipDiv.style.display = 'block';
                                        tipDiv.style.backgroundColor = '#fff3cd';
                                        tipDiv.style.color = '#856404';
                                        tipDiv.style.border = '1px solid #ffc107';
                                        tipDiv.innerHTML = '⏳ 正在检查重复...';
                                        
                                        try {{
                                            const response = await fetch('/achievement/check_duplicate', {{
                                                method: 'POST',
                                                headers: {{'Content-Type': 'application/json'}},
                                                body: JSON.stringify({{title: title, type_name: typeName}})
                                            }});
                                            
                                            const result = await response.json();
                                            
                                            if (result.exists) {{
                                                tipDiv.style.backgroundColor = '#f8d7da';
                                                tipDiv.style.color = '#721c24';
                                                tipDiv.style.border = '1px solid #f5c6cb';
                                                tipDiv.innerHTML = '<strong>❌ 检测到重复成果！</strong><br>《' + title + '》已存在于数据库中，无法重复导入！<br><small>' + (result.message || '') + '</small>';
                                                isDuplicate = true;
                                                submitBtn.disabled = true;
                                                submitBtn.style.background = '#ccc';
                                                submitBtn.style.cursor = 'not-allowed';
                                                submitBtn.title = '检测到重复成果，无法导入';
                                            }} else {{
                                                tipDiv.style.backgroundColor = '#d4edda';
                                                tipDiv.style.color = '#155724';
                                                tipDiv.style.border = '1px solid #c3e6cb';
                                                tipDiv.innerHTML = '<strong>✅ 未检测到重复</strong><br>可以导入该成果';
                                                isDuplicate = false;
                                                submitBtn.disabled = false;
                                                submitBtn.style.background = '#27ae60';
                                                submitBtn.style.cursor = 'pointer';
                                                submitBtn.removeAttribute('title');
                                            }}
                                        }} catch (error) {{
                                            console.error('查重检查失败:', error);
                                            tipDiv.style.display = 'none';
                                            tipDiv.innerHTML = '';
                                            isDuplicate = false;
                                            submitBtn.disabled = false;
                                        }}
                                    }}
                                    
                                    // 根据成果类型动态更新详细字段
                                    function updateDetailFields() {{
                                        const typeName = document.getElementById('type_name').value;
                                        const detailFieldsDiv = document.getElementById('detailFieldsArea');
                                        
                                        if (!typeName) {{
                                            detailFieldsDiv.innerHTML = '<p style="color:#999;">暂无详细字段</p>';
                                            return;
                                        }}
                                        
                                        const fields = fieldMapping[typeName] || [];
                                        let html = '<h5 style="margin-bottom:15px;">📋 详细字段（AI 自动提取，可修改）</h5>';
                                        
                                        if (fields.length === 0) {{
                                            html += '<p style="color:#999;">暂无详细字段</p>';
                                        }} else {{
                                            fields.forEach(function(field) {{
                                                const fieldName = field[0];
                                                const fieldLabel = field[1];
                                                const fieldType = field[2];
                                                const initialValue = aiInitialValues[fieldName] || '';
                                                
                                                if (fieldType === 'textarea') {{
                                                    html += `                                                        <div class="form-group">
                                                            <label>${{fieldLabel}}</label>
                                                            <textarea name="detail_${{fieldName}}" rows="3" style="width:100%; padding:10px; font-size:13px;" placeholder="请输入${{fieldLabel}}">${{initialValue}}</textarea>
                                                        </div>
                                                    `;
                                                }} else {{
                                                    html += `                                                        <div class="form-group">
                                                            <label>${{fieldLabel}}</label>
                                                            <input type="${{fieldType}}" name="detail_${{fieldName}}" value="${{initialValue}}" placeholder="请输入${{fieldLabel}}">
                                                        </div>
                                                    `;
                                                }}
                                            }});
                                        }}
                                        
                                        detailFieldsDiv.innerHTML = html;
                                    }}
                                    
                                    // 表单提交前再次检查
                                    document.getElementById('confirmForm').addEventListener('submit', function(e) {{
                                        if (isDuplicate) {{
                                            e.preventDefault();
                                            alert('检测到重复成果，无法导入！请修改成果名称。');
                                            return false;
                                        }}
                                    }});
                                    
                                    // 页面加载时初始化一次字段
                                    window.addEventListener('DOMContentLoaded', function() {{
                                        updateDetailFields();
                                    }});
                                </script>
                                '''

                            return render_base_layout('OCR 智能导入 - 结果确认', form_content, current_user)


            except Exception as e:
                success = False
                msg = f'处理失败：{str(e)}'
                logger.error(f"OCR 导入处理异常：{str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                flash(msg, 'danger')
            finally:
                # 清理临时文件
                if os.path.exists(upload_path):
                    os.remove(upload_path)
                for img_path in temp_images:
                    if os.path.exists(img_path):
                        os.remove(img_path)

    # 错误情况处理
    error_message = locals().get('msg', '处理过程中发生未知错误')
    content = f'''
        <div class="alert alert-danger">
            <h4>操作失败</h4>
            <p>{error_message}</p>
        </div>
        <a href="/achievement/ocr_import" class="btn">重新识别</a>
        <a href="/" class="btn">返回首页</a>
    '''
    return render_base_layout('OCR 智能导入', content, current_user)


@app.route('/achievement/ocr_import/confirm', methods=['POST'])
def ocr_import_confirm():
    """确认导入 OCR 识别结果"""
    current_user = get_current_user()
    if not current_user:
        flash('请先登录！', 'danger')
        return redirect(url_for('login'))

    try:
        # 获取表单数据
        ocr_text = request.form.get('raw_text', '').strip()
        type_name = request.form.get('type_name', '').strip()
        title = request.form.get('title', '').strip()
        ai_data_str = request.form.get('ai_data', '{}')

        if not ocr_text or not type_name or not title:
            flash('成果类型、名称和原始文本不能为空！', 'danger')
            return redirect('/achievement/ocr_import')

        # ========== 新增：查重校验 ==========
        duplicate_check = check_achievement_duplicate(title, type_name, current_user.id)
        if duplicate_check['exists']:
            flash(f'❌ 检测到重复成果：《{title}》已存在于数据库中（{duplicate_check["type_name"]}），无法重复导入！', 'danger')
            return redirect('/achievement/ocr_import')
        # ===================================

        # 解析 AI 数据
        try:
            ai_data = json.loads(ai_data_str) if ai_data_str else {}
        except:
            ai_data = {}

        # 构建成果信息
        achievement_info = {
            'type_name': type_name,
            'title': title,
            'raw_text': ocr_text,
            'ai_data': ai_data,
            'extra_fields': {}
        }

        # 从详细字段中提取信息
        for key, value in request.form.items():
            if key.startswith('detail_') and value.strip():
                field_name = key.replace('detail_', '')
                achievement_info['extra_fields'][field_name] = value.strip()

            # 同时从原始文本中提取未识别的字段（作为补充）
        achievement_info['extra_fields']['raw_text_backup'] = ocr_text

        # 创建成果记录
        success, msg, result_type_name, achievement_id = create_achievement_from_ocr(achievement_info, current_user)

        if success:
            # 映射成果类型到对应路由
            type_route_mapping = {
                '期刊论文': '/achievement/journal_paper',
                '会议论文': '/achievement/conference_paper',
                '教材': '/achievement/textbook',
                '专著': '/achievement/monograph',
                '发明专利': '/achievement/patent',
                '实用新型专利': '/achievement/patent',
                '软著': '/achievement/software_copyright',
                '教学成果获奖': '/achievement/teaching_achievement_award',
                '教学竞赛获奖': '/achievement/teaching_competition_award',
                '指导学生获奖': '/achievement/student_guidance_award',
                '教研教改和课程建设项目': '/achievement/teaching_project'
            }
            target_route = type_route_mapping.get(result_type_name, '/')

            flash(f'✅ {msg}', 'success')

            # 显示成功页面
            content = f'''
            <div class="alert alert-success">
                <h4>✅ 导入成功！</h4>
                <p>成果类型：<strong>{result_type_name}</strong></p>
                <p>成果名称：<strong>{title}</strong></p>
                <p>{msg}</p>
            </div>
            <div style="margin-top:20px;">
                <a href="{target_route}" class="btn">📋 查看成果列表</a>
                <a href="/achievement/ocr_import" class="btn" style="background:#27ae60; margin-left:10px;">📷 继续识别</a>
                <a href="/" class="btn" style="background:#3498db; margin-left:10px;">🏠 返回首页</a>
            </div>
            '''
            return render_base_layout('OCR 导入成功', content, current_user)
        else:
            flash(f'❌ {msg}', 'danger')
            return redirect('/achievement/ocr_import')

    except Exception as e:
        logger.error(f"确认导入 OCR 结果失败：{str(e)}")
        flash(f'导入失败：{str(e)}', 'danger')
        return redirect('/achievement/ocr_import')


@app.route('/achievement/check_duplicate', methods=['POST'])
def check_duplicate_api():
    """前端 AJAX 调用查重接口"""
    current_user = get_current_user()
    if not current_user:
        return json.dumps({'exists': False, 'error': '未登录'})

    try:
        data = request.get_json()
        title = data.get('title', '').strip()
        type_name = data.get('type_name', '').strip()

        if not title or not type_name:
            return json.dumps({'exists': False, 'error': '参数错误'})

        duplicate_check = check_achievement_duplicate(title, type_name, current_user.id)
        return json.dumps(duplicate_check)

    except Exception as e:
        logger.error(f"查重接口调用失败：{str(e)}")
        return json.dumps({'exists': False, 'error': str(e)})


@app.route('/achievement/voice_export', methods=['GET', 'POST'])
def voice_export():
    """语音导出成果（支持手动修改识别文字）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 检查百度API配置（语音识别必需）
    api_config = current_user.get_api_config()
    if not api_config.get('baidu', {}).get('api_key') or not api_config.get('baidu', {}).get('secret_key'):
        content = '''
        <div class="alert alert-danger">
            未配置百度API Key/Secret Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型API配置</a> 配置百度语音识别API。
        </div>
        <a href="/" class="btn">返回首页</a>
        '''
        return render_base_layout('语音导出成果', content, current_user)

    # 处理音频数据或手动修改后的文字提交
    if request.method == 'POST':
        try:
            # 情况1：上传录音文件（语音识别）
            if 'audio_blob' in request.files:
                audio_data = request.files['audio_blob'].read()
                # 音频转文字
                voice_text, err = audio_to_text(audio_data, current_user)
                if err:
                    return json.dumps({'status': 'error', 'msg': f'语音识别失败：{err}'})

                return json.dumps({
                    'status': 'success',
                    'voice_text': voice_text,
                    'export_url': '',
                    'msg': '语音识别完成，可手动修改后导出'
                })

            # 情况2：手动修改文字后提交导出
            elif 'voice_text' in request.form:
                voice_text = request.form.get('voice_text', '').strip()
                if not voice_text:
                    return json.dumps({'status': 'error', 'msg': '请输入导出指令'})

                # 解析语音指令
                cmd = parse_voice_command(voice_text)

                # 根据指令生成导出链接（核心修改：添加时间参数）
                export_url = ''
                export_msg = ''
                type_route_mapping = {
                    '期刊论文': '/achievement/journal_paper',
                    '会议论文': '/achievement/conference_paper',
                    '教材': '/achievement/textbook',
                    '专著': '/achievement/monograph',
                    '专利': '/achievement/patent',
                    '软著': '/achievement/software_copyright',
                    '教学成果获奖': '/achievement/teaching_achievement_award',
                    '教学竞赛获奖': '/achievement/teaching_competition_award',
                    '指导学生获奖': '/achievement/student_guidance_award'
                }

                if cmd['action'] == 'export' and cmd['type_name']:
                    base_url = type_route_mapping.get(cmd['type_name'], '')
                    if base_url:
                        # 拼接时间参数
                        export_url = f"{base_url}?action=export"
                        if cmd['start_date']:
                            export_url += f"&start_date={cmd['start_date']}"
                        if cmd['end_date']:
                            export_url += f"&end_date={cmd['end_date']}"
                        export_msg = f'已识别指令：导出{cmd["start_date"] if cmd["start_date"] else ""}{cmd["end_date"] if cmd["end_date"] else ""}的{cmd["type_name"]}'
                    else:
                        export_msg = f'暂不支持导出{cmd["type_name"]}类型成果'
                else:
                    export_msg = f'未识别有效导出指令，识别文本：{voice_text}'

                return json.dumps({
                    'status': 'success',
                    'voice_text': voice_text,
                    'export_url': export_url,
                    'msg': export_msg
                })

        except Exception as e:
            return json.dumps({'status': 'error', 'msg': f'处理失败：{str(e)}'})

    # 渲染语音导出页面（增加手动修改功能）
    # 原有HTML代码保持不变...
    form_html = '''
    <h2>语音导出成果</h2>
    <div class="alert alert-info">
        支持语音指令示例：<br>
        - 导出2024年的期刊论文<br>
        - 导出团队的教学竞赛获奖<br>
        - 导出我的所有教材
    </div>

    <!-- 录音区域 -->
    <div style="margin:20px 0; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h4>网页录音导出</h4>
        <button id="recordBtn" class="btn" style="background:#27ae60;">开始录音</button>
        <button id="stopBtn" class="btn" style="background:#e74c3c; display:none;">停止录音</button>
        <div id="recordStatus" style="margin-top:10px; color:#666;"></div>

        <!-- 识别结果展示 + 手动修改 -->
        <div id="resultArea" style="margin-top:20px; display:none;">
            <div class="alert alert-info">
                <h5>语音识别结果（可手动修改）：</h5>
                <textarea id="voiceTextInput" style="width:100%; height:100px; margin:10px 0; padding:10px;" placeholder="请输入导出指令..."></textarea>
                <button id="submitTextBtn" class="btn">确认导出</button>
            </div>
            <div id="exportArea"></div>
        </div>
    </div>

    <script>
        let recorder = null;
        let audioBlob = null;

        // 开始录音
        document.getElementById('recordBtn').addEventListener('click', async () => {
            try {
                // 获取麦克风权限
                const stream = await navigator.mediaDevices.getUserMedia({ audio: true });
                recorder = new MediaRecorder(stream);
                const chunks = [];

                // 收集录音数据
                recorder.ondataavailable = (e) => chunks.push(e.data);

                // 录音停止后处理
                recorder.onstop = async () => {
                    audioBlob = new Blob(chunks, { type: 'audio/webm' });
                    document.getElementById('recordStatus').textContent = '录音完成，正在识别...';

                    // 创建FormData并提交录音数据
                    const formData = new FormData();
                    formData.append('audio_blob', audioBlob, 'record.webm');

                    // 发送请求识别语音
                    const response = await fetch('/achievement/voice_export', {
                        method: 'POST',
                        body: formData
                    });

                    const result = await response.json();
                    document.getElementById('recordStatus').textContent = '';

                    // 展示结果（可修改）
                    document.getElementById('resultArea').style.display = 'block';
                    if (result.status === 'success') {
                        // 填充识别结果到文本框
                        document.getElementById('voiceTextInput').value = result.voice_text;
                        document.getElementById('exportArea').innerHTML = 
                            `<div class="alert alert-success">${result.msg}</div>`;
                    } else {
                        document.getElementById('exportArea').innerHTML = 
                            `<div class="alert alert-danger">${result.msg}</div>`;
                    }

                    // 停止所有音轨
                    stream.getTracks().forEach(track => track.stop());
                };

                // 开始录音
                recorder.start();
                document.getElementById('recordBtn').style.display = 'none';
                document.getElementById('stopBtn').style.display = 'inline-block';
                document.getElementById('recordStatus').textContent = '正在录音...（点击停止按钮结束）';

            } catch (err) {
                document.getElementById('recordStatus').textContent = `录音权限获取失败：${err.message}`;
            }
        });

        // 停止录音
        document.getElementById('stopBtn').addEventListener('click', () => {
            if (recorder && recorder.state === 'recording') {
                recorder.stop();
                document.getElementById('recordBtn').style.display = 'inline-block';
                document.getElementById('stopBtn').style.display = 'none';
            }
        });

        // 手动修改文字后提交导出
        document.getElementById('submitTextBtn').addEventListener('click', async () => {
            const voiceText = document.getElementById('voiceTextInput').value.trim();
            if (!voiceText) {
                alert('请输入导出指令');
                return;
            }

            document.getElementById('exportArea').innerHTML = '<div class="alert alert-info">正在解析导出指令...</div>';

            // 提交手动修改后的文字
            const formData = new FormData();
            formData.append('voice_text', voiceText);

            const response = await fetch('/achievement/voice_export', {
                method: 'POST',
                body: formData
            });

            const result = await response.json();
            if (result.status === 'success') {
                // 显示导出链接
                let exportHtml = '';
                if (result.export_url) {
                    exportHtml = `
                    <div class="alert alert-success">
                        <p>${result.msg}</p>
                        <a href="${result.export_url}" class="btn">点击导出成果</a>
                    </div>
                    `;
                } else {
                    exportHtml = `<div class="alert alert-warning">${result.msg}</div>`;
                }
                document.getElementById('exportArea').innerHTML = exportHtml;
            } else {
                document.getElementById('exportArea').innerHTML = 
                    `<div class="alert alert-danger">${result.msg}</div>`;
            }
        });
    </script>
    '''
    return render_base_layout('语音导出成果', form_html, current_user)

@app.route('/team/voice_export', methods=['GET', 'POST'])
def team_voice_export():
    """团队负责人专属：语音导出团队公开成果（支持指定老师+成果类型）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('仅团队负责人可使用此功能！', 'danger')
        return redirect(url_for('index'))

    api_config = current_user.get_api_config()
    if not api_config.get('baidu', {}).get('api_key') or not api_config.get('baidu', {}).get('secret_key'):
        content = '''
        <div class="alert alert-danger">
            未配置百度API Key/Secret Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型API配置</a> 配置百度语音识别API。
        </div>
        <a href="/team/list" class="btn">返回团队列表</a>
        '''
        return render_base_layout('团队语音导出成果', content, current_user)

    managed_teams = Team.query.filter_by(leader_id=current_user.id).all()
    if not managed_teams:
        content = '''
            <div class="alert alert-warning">
                您尚未管理任何团队！
            </div>
            <a href="/team/list" class="btn">创建团队</a>
            '''
        return render_base_layout('团队语音导出成果', content, current_user)

    # GET 请求：显示团队选择页面
    if request.method == 'GET':
        team_options = ''.join([
            f'<option value="{team.id}">{team.name}</option>'
            for team in managed_teams
        ])

        form_html = f'''
    <h2>团队语音导出成果</h2>
    <div class="alert alert-info">
        支持语音指令示例：<br>
        - 导出张三老师的 2024 年期刊论文<br>
        - 导出李四老师的专利<br>
        - 导出所有成员的教学竞赛获奖<br>
        - 导出王五老师的近三年教研教改和课程建设项目
    </div>

    <!-- 第一步：选择团队 -->
    <div style="margin:20px 0; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h4>步骤 1：选择要导出哪个团队的成果</h4>
        <div class="form-group">
            <label>选择团队</label>
            <select id="teamSelect" onchange="showVoiceInput()" style="width:100%; padding:10px; font-size:14px;">
                <option value="">请选择团队...</option>
                {team_options}
            </select>
        </div>
    </div>

    <!-- 第二步：语音输入（选择团队后显示） -->
    <div id="voiceInputArea" style="display:none; margin:20px 0; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h4>步骤 2：语音或文字输入导出指令</h4>
        <button id="recordBtn" class="btn" style="background:#27ae60;">开始录音</button>
        <button id="stopBtn" class="btn" style="background:#e74c3c; display:none;">停止录音</button>
        <div id="recordStatus" style="margin-top:10px; color:#666;"></div>

        <div id="resultArea" style="margin-top:20px; display:none;">
            <div class="alert alert-info">
                <h5>语音识别结果（可手动修改）：</h5>
                <textarea id="voiceTextInput" style="width:100%; height:100px; margin:10px 0; padding:10px;" placeholder="请输入导出指令..."></textarea>
                <button id="submitTextBtn" class="btn">确认导出</button>
            </div>
            <div id="exportArea"></div>
        </div>
    </div>

    <script>
        let recorder = null;
        let audioBlob = null;
        let selectedTeamId = null;
        let selectedTeamName = '';

        function showVoiceInput() {{
            const teamSelect = document.getElementById('teamSelect');
            selectedTeamId = teamSelect.value;
            selectedTeamName = teamSelect.options[teamSelect.selectedIndex].text;

            if (selectedTeamId) {{
                document.getElementById('voiceInputArea').style.display = 'block';
            }} else {{
                document.getElementById('voiceInputArea').style.display = 'none';
            }}
        }}

        document.getElementById('recordBtn').addEventListener('click', async () => {{
            try {{
                const stream = await navigator.mediaDevices.getUserMedia({{ audio: true }});
                recorder = new MediaRecorder(stream);
                const chunks = [];

                recorder.ondataavailable = (e) => chunks.push(e.data);

                recorder.onstop = async () => {{
                    audioBlob = new Blob(chunks, {{ type: 'audio/webm' }});
                    document.getElementById('recordStatus').textContent = '录音完成，正在识别...';

                    const formData = new FormData();
                    formData.append('audio_blob', audioBlob, 'record.webm');

                    const response = await fetch('/team/voice_export', {{
                        method: 'POST',
                        body: formData
                    }});

                    const result = await response.json();
                    document.getElementById('recordStatus').textContent = '';

                    document.getElementById('resultArea').style.display = 'block';
                    if (result.status === 'success') {{
                        document.getElementById('voiceTextInput').value = result.voice_text;
                        document.getElementById('exportArea').innerHTML =
                            `<div class="alert alert-success">${{result.msg}}</div>`;
                    }} else {{
                        document.getElementById('exportArea').innerHTML =
                            `<div class="alert alert-danger">${{result.msg}}</div>`;
                    }}

                    stream.getTracks().forEach(track => track.stop());
                }};

                recorder.start();
                document.getElementById('recordBtn').style.display = 'none';
                document.getElementById('stopBtn').style.display = 'inline-block';
                document.getElementById('recordStatus').textContent = '正在录音...（点击停止按钮结束）';

            }} catch (err) {{
                document.getElementById('recordStatus').textContent = `录音权限获取失败：${{err.message}}`;
            }}
        }});

        document.getElementById('stopBtn').addEventListener('click', () => {{
            if (recorder && recorder.state === 'recording') {{
                recorder.stop();
                document.getElementById('recordBtn').style.display = 'inline-block';
                document.getElementById('stopBtn').style.display = 'none';
            }}
        }});

        document.getElementById('submitTextBtn').addEventListener('click', async () => {{
            const voiceText = document.getElementById('voiceTextInput').value.trim();
            if (!voiceText) {{
                alert('请输入导出指令');
                return;
            }}

            document.getElementById('exportArea').innerHTML = '<div class="alert alert-info">正在解析导出指令...</div>';

            const formData = new FormData();
            formData.append('voice_text', voiceText);
            formData.append('team_id', selectedTeamId);

            const response = await fetch('/team/voice_export', {{
                method: 'POST',
                body: formData
            }});

            const result = await response.json();
            if (result.status === 'success') {{
                let exportHtml = '';
                if (result.export_url) {{
                    exportHtml = `
    <div class="alert alert-success">
    <p>${{result.msg}}</p>
    <a href="${{result.export_url}}" class="btn">点击导出成果</a>
    </div>`;
                }} else {{
                    exportHtml = `<div class="alert alert-warning">${{result.msg}}</div>`;
                }}
                document.getElementById('exportArea').innerHTML = exportHtml;
            }} else if (result.status === 'warning') {{
                document.getElementById('exportArea').innerHTML =
                    `<div class="alert alert-warning">${{result.msg}}</div>`;
            }} else {{
                document.getElementById('exportArea').innerHTML =
                    `<div class="alert alert-danger">${{result.msg}}</div>`;
            }}
        }});
    </script>
    '''
        return render_base_layout('团队语音导出成果', form_html, current_user)

    # POST 请求：处理语音识别和指令解析
    try:
        if 'audio_blob' in request.files:
            audio_data = request.files['audio_blob'].read()
            voice_text, err = audio_to_text(audio_data, current_user)
            if err:
                return json.dumps({'status': 'error', 'msg': f'语音识别失败：{err}'})

            return json.dumps({
                'status': 'success',
                'voice_text': voice_text,
                'export_url': '',
                'msg': '语音识别完成，可手动修改后导出'
            })

        elif 'voice_text' in request.form:
            voice_text = request.form.get('voice_text', '').strip()
            team_id = request.form.get('team_id', type=int)

            if not team_id:
                return json.dumps({'status': 'error', 'msg': '请先选择团队'})

            # 验证团队权限
            team = db.session.get(Team, team_id)
            if not team or team.leader_id != current_user.id:
                return json.dumps({'status': 'error', 'msg': '无权限导出该团队成果'})

            if not voice_text:
                return json.dumps({'status': 'error', 'msg': '请输入导出指令'})

            cmd = parse_voice_command(voice_text)

            export_url = ''
            export_msg = ''
            error_msg = ''

            target_teacher = None
            if cmd.get('teacher_name'):
                teacher_name = cmd['teacher_name']
                target_teacher = User.query.filter(
                    or_(
                        User.username.like(f'%{teacher_name}%'),
                        User.employee_id.like(f'%{teacher_name}%')
                    ),
                    User.id.in_([ut.user_id for ut in UserTeam.query.filter_by(team_id=team_id).all()])
                ).first()

                if not target_teacher:
                    export_msg = f'未找到团队成员：{teacher_name}'
                    return json.dumps({
                        'status': 'warning',
                        'voice_text': voice_text,
                        'export_url': '',
                        'msg': export_msg
                    })

            type_mapping = {
                '期刊论文': (JournalPaper, 'journal_paper'),
                '会议论文': (ConferencePaper, 'conference_paper'),
                '教材': (Textbook, 'textbook'),
                '专著': (Monograph, 'monograph'),
                '专利': (Patent, 'patent'),
                '软著': (SoftwareCopyright, 'software_copyright'),
                '教学成果获奖': (TeachingAchievementAward, 'teaching_achievement_award'),
                '教学竞赛获奖': (TeachingCompetitionAward, 'teaching_competition_award'),
                '指导学生获奖': (StudentGuidanceAward, 'student_guidance_award'),
                '教研教改和课程建设项目': (TeachingProject, 'teaching_project')
            }

            if cmd['action'] == 'export' and cmd['type_name']:
                if cmd['type_name'] not in type_mapping:
                    export_msg = f'暂不支持导出{cmd["type_name"]}类型成果'
                else:
                    model, type_key = type_mapping[cmd['type_name']]
                    export_url = f"/team/export_specified?team_id={team_id}"
                    export_url += f"&type={type_key}"

                    if target_teacher:
                        export_url += f"&teacher_id={target_teacher.id}"
                        teacher_info = f"{target_teacher.username}（{target_teacher.employee_id}）"
                    else:
                        teacher_info = "所有成员"

                    if cmd['start_date']:
                        export_url += f"&start_date={cmd['start_date']}"
                    if cmd['end_date']:
                        export_url += f"&end_date={cmd['end_date']}"

                    time_info = ""
                    if cmd['start_date'] and cmd['end_date']:
                        time_info = f"{cmd['start_date'][:4]}-{cmd['end_date'][:4]}年"
                    elif cmd['start_date']:
                        time_info = f"{cmd['start_date'][:4]}年"

                    export_msg = f'已识别指令：导出{time_info}{teacher_info}的{cmd["type_name"]}（仅公开给{team.name}的成果）'

            else:
                export_msg = f'未识别有效导出指令，示例：导出张三老师的 2024 年期刊论文'

            return json.dumps({
                'status': 'success',
                'voice_text': voice_text,
                'export_url': export_url,
                'msg': export_msg
            })

    except Exception as e:
        return json.dumps({'status': 'error', 'msg': f'处理失败：{str(e)}'})

    # 修复：修正JavaScript模板字符串语法，移除多余的$符号
    form_html = f'''
<h2>团队语音导出成果（{current_team.name}）</h2>
<div class="alert alert-info">
    支持语音指令示例：<br>
    - 导出张三老师的2024年期刊论文<br>
    - 导出李四老师的专利<br>
    - 导出所有成员的教学竞赛获奖<br>
    - 导出王五老师的近三年教研教改和课程建设项目
</div>

<div style="margin:20px 0; padding:20px; border:1px solid #eee; border-radius:8px;">
    <h4>网页录音导出</h4>
    <button id="recordBtn" class="btn" style="background:#27ae60;">开始录音</button>
    <button id="stopBtn" class="btn" style="background:#e74c3c; display:none;">停止录音</button>
    <div id="recordStatus" style="margin-top:10px; color:#666;"></div>

    <div id="resultArea" style="margin-top:20px; display:none;">
        <div class="alert alert-info">
            <h5>语音识别结果（可手动修改）：</h5>
            <textarea id="voiceTextInput" style="width:100%; height:100px; margin:10px 0; padding:10px;" placeholder="请输入导出指令..."></textarea>
            <button id="submitTextBtn" class="btn">确认导出</button>
        </div>
        <div id="exportArea"></div>
    </div>
</div>

<script>
    let recorder = null;
    let audioBlob = null;

    document.getElementById('recordBtn').addEventListener('click', async () => {{
        try {{
            const stream = await navigator.mediaDevices.getUserMedia({{ audio: true }});
            recorder = new MediaRecorder(stream);
            const chunks = [];

            recorder.ondataavailable = (e) => chunks.push(e.data);

            recorder.onstop = async () => {{
                audioBlob = new Blob(chunks, {{ type: 'audio/webm' }});
                document.getElementById('recordStatus').textContent = '录音完成，正在识别...';

                const formData = new FormData();
                formData.append('audio_blob', audioBlob, 'record.webm');

                const response = await fetch('/team/voice_export', {{
                    method: 'POST',
                    body: formData
                }});

                const result = await response.json();
                document.getElementById('recordStatus').textContent = '';

                document.getElementById('resultArea').style.display = 'block';
                if (result.status === 'success') {{
                    document.getElementById('voiceTextInput').value = result.voice_text;
                    document.getElementById('exportArea').innerHTML =
                        `<div class="alert alert-success">${{result.msg}}</div>`;
                }} else {{
                    document.getElementById('exportArea').innerHTML =
                        `<div class="alert alert-danger">${{result.msg}}</div>`;
                }}

                stream.getTracks().forEach(track => track.stop());
            }};

            recorder.start();
            document.getElementById('recordBtn').style.display = 'none';
            document.getElementById('stopBtn').style.display = 'inline-block';
            document.getElementById('recordStatus').textContent = '正在录音...（点击停止按钮结束）';

        }} catch (err) {{
            document.getElementById('recordStatus').textContent = `录音权限获取失败：${{err.message}}`;
        }}
    }});

    document.getElementById('stopBtn').addEventListener('click', () => {{
        if (recorder && recorder.state === 'recording') {{
            recorder.stop();
            document.getElementById('recordBtn').style.display = 'inline-block';
            document.getElementById('stopBtn').style.display = 'none';
        }}
    }});

    document.getElementById('submitTextBtn').addEventListener('click', async () => {{
        const voiceText = document.getElementById('voiceTextInput').value.trim();
        if (!voiceText) {{
            alert('请输入导出指令');
            return;
        }}

        document.getElementById('exportArea').innerHTML = '<div class="alert alert-info">正在解析导出指令...</div>';

        const formData = new FormData();
        formData.append('voice_text', voiceText);

        const response = await fetch('/team/voice_export', {{
            method: 'POST',
            body: formData
        }});

        const result = await response.json();
        if (result.status === 'success') {{
            let exportHtml = '';
            if (result.export_url) {{
                exportHtml = `
<div class="alert alert-success">
<p>${{result.msg}}</p>
<a href="${{result.export_url}}" class="btn">点击导出成果</a>
</div>`;
            }} else {{
                exportHtml = `<div class="alert alert-warning">${{result.msg}}</div>`;
            }}
            document.getElementById('exportArea').innerHTML = exportHtml;
        }} else if (result.status === 'warning') {{
            document.getElementById('exportArea').innerHTML =
                `<div class="alert alert-warning">${{result.msg}}</div>`;
        }} else {{
            document.getElementById('exportArea').innerHTML =
                `<div class="alert alert-danger">${{result.msg}}</div>`;
        }}
    }});
</script>
'''

    return render_base_layout(f'团队语音导出成果 - {current_team.name}', form_html, current_user)

@app.route('/team/export_specified')
def team_export_specified():
    """导出团队指定老师的指定类型公开成果"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('仅团队负责人可使用此功能！', 'danger')
        return redirect(url_for('index'))

    team_id = request.args.get('team_id', type=int)
    type_key = request.args.get('type')
    teacher_id = request.args.get('teacher_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not team_id or not type_key:
        flash('导出参数缺失！', 'danger')
        return redirect(url_for('team_voice_export'))

    team = db.session.get(Team, team_id)
    if not team or team.leader_id != current_user.id:
        flash('无权限导出该团队成果！', 'danger')
        return redirect(url_for('team_voice_export'))
    team_id_str = str(team_id)

    type_model_mapping = {
        'journal_paper': (JournalPaper, '期刊论文', [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'corresponding_authors', 'label': '通讯作者'},
            {'name': 'journal_name', 'label': '期刊名称'},
            {'name': 'inclusion_status', 'label': '收录情况'},
            {'name': 'year', 'label': '年'},
            {'name': 'volume', 'label': '卷'},
            {'name': 'issue', 'label': '期'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'publish_date', 'label': '发表日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'conference_paper': (ConferencePaper, '会议论文', [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'corresponding_authors', 'label': '通讯作者'},
            {'name': 'conference_name', 'label': '会议名称'},
            {'name': 'conference_start_date', 'label': '会议开始日期'},
            {'name': 'conference_end_date', 'label': '会议结束日期'},
            {'name': 'conference_place', 'label': '会议地点'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'textbook': (Textbook, '教材', [
            {'name': 'title', 'label': '教材名称'},
            {'name': 'textbook_series', 'label': '教材系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'cip_number', 'label': 'CIP 核字号'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'publication_month', 'label': '出版月份'},
            {'name': 'publish_date', 'label': '出版日期'},
            {'name': 'edition', 'label': '版次'},
            {'name': 'word_count', 'label': '字数'},
            {'name': 'price', 'label': '定价'},
            {'name': 'textbook_level_id', 'label': '教材级别'},
            {'name': 'textbook_type', 'label': '教材类型'},
            {'name': 'applicable_majors', 'label': '适用专业'},
            {'name': 'remarks', 'label': '备注'},
            {'name': 'textbook_attachment', 'label': '附件'}
        ]),
        'monograph': (Monograph, '专著', [
            {'name': 'title', 'label': '专著名称'},
            {'name': 'textbook_series', 'label': '专著系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'cip_number', 'label': 'CIP 核字号'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'publication_month', 'label': '出版月份'},
            {'name': 'publish_date', 'label': '出版日期'},
            {'name': 'edition', 'label': '版次'},
            {'name': 'word_count', 'label': '字数'},
            {'name': 'price', 'label': '定价'},
            {'name': 'monograph_type', 'label': '专著类型'},
            {'name': 'applicable_majors', 'label': '适用专业'},
            {'name': 'remarks', 'label': '备注'},
            {'name': 'monograph_attachment', 'label': '附件'}
        ]),
        'patent': (Patent, '专利', [
            {'name': 'title', 'label': '专利名称'},
            {'name': 'patent_type_id', 'label': '专利类型'},
            {'name': 'patentee', 'label': '专利权人'},
            {'name': 'address', 'label': '地址'},
            {'name': 'inventors', 'label': '发明人'},
            {'name': 'patent_status_id', 'label': '专利状态'},
            {'name': 'patent_number', 'label': '专利号'},
            {'name': 'grant_announcement_number', 'label': '授权公告号'},
            {'name': 'apply_date', 'label': '专利申请日'},
            {'name': 'grant_announcement_date', 'label': '授权公告日'},
            {'name': 'applicant_at_apply_date', 'label': '申请日时申请人'},
            {'name': 'inventor_at_apply_date', 'label': '申请日时发明人'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'software_copyright': (SoftwareCopyright, '软著', [
            {'name': 'title', 'label': '软件名称'},
            {'name': 'copyright_owner', 'label': '著作权人'},
            {'name': 'completion_date', 'label': '开发完成日期'},
            {'name': 'first_publication_date', 'label': '首次发表日期'},
            {'name': 'right_acquisition_method', 'label': '权利取得方式'},
            {'name': 'right_scope', 'label': '权利范围'},
            {'name': 'copyright_number', 'label': '登记号'},
            {'name': 'certificate_number', 'label': '证书号'},
            {'name': 'register_date', 'label': '登记日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'teaching_achievement_award': (TeachingAchievementAward, '教学成果获奖', [
            {'name': 'title', 'label': '成果名称'},
            {'name': 'achievement_type_id', 'label': '教学成果奖类型'},
            {'name': 'achievement_level_id', 'label': '成果等级'},
            {'name': 'main_contributors', 'label': '主要完成人'},
            {'name': 'completing_units', 'label': '成果完成单位'},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'award_rank_id', 'label': '获奖等级'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'awarding_unit', 'label': '颁奖单位'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'teaching_competition_award': (TeachingCompetitionAward, '教学竞赛获奖', [
            {'name': 'title', 'label': '竞赛名称'},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'competition_level_id', 'label': '竞赛等级'},
            {'name': 'award_rank_id', 'label': '获奖等级'},
            {'name': 'winners', 'label': '获奖人'},
            {'name': 'winner_unit', 'label': '获奖人所在单位'},
            {'name': 'competition_name', 'label': '竞赛主办方'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'student_guidance_award': (StudentGuidanceAward, '指导学生获奖', [
            {'name': 'title', 'label': '获奖名称'},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'competition_name', 'label': '竞赛名称'},
            {'name': 'competition_level_id', 'label': '竞赛等级'},
            {'name': 'award_rank_id', 'label': '获奖等级'},
            {'name': 'student_name', 'label': '获奖学生'},
            {'name': 'project_name', 'label': '获奖项目名称'},
            {'name': 'teacher_name', 'label': '指导教师'},
            {'name': 'student_unit', 'label': '获奖学生所在单位'},
            {'name': 'organizer', 'label': '竞赛主办方'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'teaching_project': (TeachingProject, '教研教改和课程建设项目', [
            {'name': 'title', 'label': '项目名称'},
            {'name': 'project_code', 'label': '项目编号'},
            {'name': 'project_type_id', 'label': '项目类型'},
            {'name': 'project_leader', 'label': '项目负责人'},
            {'name': 'project_members', 'label': '项目参与人'},
            {'name': 'approval_department', 'label': '批准部门'},
            {'name': 'approval_date', 'label': '立项时间'},
            {'name': 'project_level_id', 'label': '项目级别'},
            {'name': 'project_category_id', 'label': '项目类别'},
            {'name': 'funding', 'label': '经费'},
            {'name': 'start_date', 'label': '开始时间'},
            {'name': 'end_date', 'label': '结束时间'},
            {'name': 'project_status_id', 'label': '项目状态'},
            {'name': 'attachment', 'label': '附件'}
        ])
    }

    if type_key not in type_model_mapping:
        flash('不支持的成果类型！', 'danger')
        return redirect(url_for('team_voice_export'))

    model, type_name, fields_config = type_model_mapping[type_key]

    query = model.query.filter(
        func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id_str, ',')) > 0
    )

    if teacher_id:
        is_team_member = UserTeam.query.filter_by(team_id=team_id, user_id=teacher_id).first()
        if not is_team_member:
            flash('该老师不属于本团队！', 'danger')
            return redirect(url_for('team_voice_export'))
        query = query.filter(model.user_id == teacher_id)

    date_field_map = {
        JournalPaper: 'publish_date',
        ConferencePaper: 'conference_time',
        Textbook: 'publish_date',
        Monograph: 'publish_date',
        TeachingProject: 'start_date',
        Patent: 'apply_date',
        SoftwareCopyright: 'register_date',
        TeachingAchievementAward: 'award_date',
        TeachingCompetitionAward: 'award_date',
        StudentGuidanceAward: 'award_date'
    }
    date_field = date_field_map.get(model)
    if date_field:
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(getattr(model, date_field) >= start_date_obj)
            except:
                pass
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(getattr(model, date_field) <= end_date_obj)
            except:
                pass

    items = query.all()
    if not items:
        flash('暂无符合条件的公开成果！', 'warning')
        return redirect(url_for('team_voice_export'))

    wb = openpyxl.Workbook()
    ws = wb.active

    teacher_name = "所有成员"
    if teacher_id:
        teacher = db.session.get(User, teacher_id)
        teacher_name = teacher.username if teacher else "未知老师"
    ws.title = f"{team.name}-{teacher_name}-{type_name}"

    headers = [f['label'] for f in fields_config]
    ws.append(headers)

    for item in items:
        row = []
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            if value is None:
                value = ''
            elif isinstance(value, (date, datetime)):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field_name == 'attachment' and value:
                value = os.path.basename(value) if value else ''
            # 处理外键对象（如专利类型、项目类型等）
            elif hasattr(value, '__class__') and hasattr(value, 'type_name'):
                value = value.type_name if value else ''
            elif hasattr(value, '__class__') and hasattr(value, 'level_name'):
                value = value.level_name if value else ''
            elif hasattr(value, '__class__') and hasattr(value, 'status_name'):
                value = value.status_name if value else ''
            elif hasattr(value, '__class__') and hasattr(value, 'rank_name'):
                value = value.rank_name if value else ''
            elif hasattr(value, '__class__') and hasattr(value, 'category_name'):
                value = value.category_name if value else ''
            row.append(value)
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{team.name}_{teacher_name}_{type_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.errorhandler(413)
def request_entity_too_large(error):
    user = get_current_user()
    content = '''
    <div class="alert alert-danger">
        <h4>上传失败！</h4>
        <p>文件大小超过限制（当前限制：100MB），请压缩或拆分文件后重新上传。</p>
    </div>
    <a href="/achievement/ocr_import" class="btn">重新上传</a>
    '''
    return render_base_layout('上传失败', content, user), 413


# ---------------------- 使用示例：添加期刊论文并关联多个作者 ----------------------

@app.route('/achievement/journal_paper/add_author', methods=['POST'])
def add_journal_paper_with_authors():
    """添加期刊论文并关联多个系统用户作者"""
    current_user = get_current_user()

    # 获取表单数据
    title = request.form.get('title')
    authors_str = request.form.get('authors')  # 逗号分隔的作者名字符串
    corresponding_authors_str = request.form.get('corresponding_authors')
    journal_name = request.form.get('journal_name')
    year = request.form.get('year', type=int)

    # 创建论文记录
    paper = JournalPaper(
        user_id=current_user.id,
        title=title,
        authors=authors_str,
        corresponding_authors=corresponding_authors_str,
        journal_name=journal_name,
        year=year
    )
    db.session.add(paper)
    db.session.flush()  # 获取 paper.id

    # 解析作者列表并关联系统用户
    import re
    author_names = [name.strip() for name in re.split(r'[;,;,,]', authors_str)]
    corresponding_names = [name.strip() for name in
                           re.split(r'[;,;,,]', corresponding_authors_str or '')] if corresponding_authors_str else []

    db.session.commit()
    flash('期刊论文添加成功！', 'success')
    return redirect(url_for('journal_paper_list'))


# ---------------------- 查询示例：获取某用户参与的所有期刊论文 ----------------------

def get_user_journal_papers(user_id):
    """获取用户参与的所有期刊论文（包含作为作者和关联人）"""
    owned_papers = JournalPaper.query.filter_by(user_id=user_id).all()

    contributed_papers = JournalPaper.query.join(AchievementContributor).filter(
        AchievementContributor.user_id == user_id,
        AchievementContributor.achievement_type == 'journal_paper'
    ).all()

    all_papers = {p.id: p for p in owned_papers}
    for p in contributed_papers:
        all_papers[p.id] = p

    return list(all_papers.values())


# ---------------------- 查询示例：获取论文的详细信息（包含作者信息） ----------------------

def get_paper_detail_with_authors(paper_id):
    """获取期刊论文详情及作者信息"""
    paper = JournalPaper.query.get(paper_id)
    if not paper:
        return None

    # 从 authors 字段解析作者信息（修复：支持多种分隔符）
    import re
    authors_info = []
    if paper.authors:
        author_names = [name.strip() for name in re.split(r'[;,;,,]', paper.authors)]
        corresponding_names = [name.strip() for name in
                               re.split(r'[;,;,,]', paper.corresponding_authors or '')] if paper.corresponding_authors else []

        for idx, name in enumerate(author_names, start=1):
            authors_info.append({
                'name': name,
                'author_order': idx,
                'is_corresponding': name in corresponding_names
            })

    return {
        'paper': paper,
        'authors': authors_info
    }


@app.route('/achievement/journal_paper/submit', methods=['POST'])
def submit_journal_paper():
    """提交期刊论文（支持多作者关联）"""
    current_user = get_current_user()

    title = request.form.get('title')
    authors_str = request.form.get('authors')
    corresponding_authors_str = request.form.get('corresponding_authors')
    journal_name = request.form.get('journal_name')
    year = request.form.get('year', type=int)
    volume = request.form.get('volume')
    issue = request.form.get('issue')
    page_range = request.form.get('page_range')
    doi = request.form.get('doi')

    paper = JournalPaper(
        user_id=current_user.id,
        title=title,
        authors=authors_str,
        corresponding_authors=corresponding_authors_str,
        journal_name=journal_name,
        year=year,
        volume=volume,
        issue=issue,
        page_range=page_range,
        doi=doi
    )
    db.session.add(paper)
    db.session.flush()

    auto_link_contributors(paper, 'journal_paper', authors_str, current_user.id)

    db.session.commit()

    flash('期刊论文添加成功！', 'success')
    return redirect(url_for('journal_paper_list'))


@app.route('/my_achievements')
def my_achievements():
    """我的成果（包含作为作者参与的成果）"""
    current_user = get_current_user()

    # 查询期刊论文（仅查看拥有的）
    papers = JournalPaper.query.filter_by(user_id=current_user.id).all()

    # 标记是否拥有
    papers_data = []
    for paper in papers:
        papers_data.append({
            'paper': paper,
            'is_owner': paper.user_id == current_user.id,
        })

    content = f'''
    <h2>我的期刊论文</h2>
    <div class="alert alert-info">
        显示您拥有的论文（共{len(papers_data)}篇）
    </div>
    <table class="table">
        <thead>
            <tr>
                <th>序号</th>
                <th>论文名称</th>
                <th>期刊名称</th>
                <th>年份</th>
                <th>您的角色</th>
                <th>操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    for idx, item in enumerate(papers_data, start=1):
        paper = item['paper']
        role_label = '<span class="badge badge-success">拥有者</span>'

        action_btns = f'<a href="/achievement/journal_paper/edit?id={paper.id}" class="btn btn-sm">编辑</a>'

        content += f'''
            <tr>
                <td>{idx}</td>
                <td>{paper.title}</td>
                <td>{paper.journal_name}</td>
                <td>{paper.year or '-'}</td>
                <td>{role_label}</td>
                <td>{action_btns}</td>
            </tr>
        '''

    content += '''
        </tbody>
    </table>
    '''

    return render_base_layout('我的期刊论文', content, current_user)


@app.route('/admin/dict_manage/<dict_type>', methods=['GET', 'POST'])
def admin_dict_manage(dict_type):
    """管理员-字典表维护（增删改查）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'admin':
        flash('无管理员权限！', 'danger')
        return redirect(url_for('index'))

    # 字典表映射
    dict_mapping = {
        'achievement_type': (TeachingAchievementType, '教学成果奖类型', ['type_name']),
        'achievement_level': (AchievementLevel, '成果等级', ['level_name']),
        'award_rank': (AwardRank, '获奖等级', ['rank_name'])
    }

    if dict_type not in dict_mapping:
        flash('无效的字典类型！', 'danger')
        return redirect(url_for('index'))

    model, dict_name, name_fields = dict_mapping[dict_type]

    # 处理操作
    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'add':
                # 添加新记录
                name_value = request.form.get(name_fields[0])
                sort_order = request.form.get('sort_order', 0, type=int)

                if model.query.filter_by(**{name_fields[0]: name_value}).first():
                    flash(f'{name_fields[0]}已存在！', 'danger')
                else:
                    new_record = model(**{name_fields[0]: name_value, 'sort_order': sort_order})
                    db.session.add(new_record)
                    db.session.commit()
                    flash(f'{dict_name}添加成功！', 'success')

            elif action == 'edit':
                # 编辑记录
                record_id = request.form.get('id')
                record = db.session.get(model, record_id)
                if record:
                    for field in name_fields:
                        setattr(record, field, request.form.get(field))
                    record.sort_order = request.form.get('sort_order', 0, type=int)
                    db.session.commit()
                    flash(f'{dict_name}更新成功！', 'success')

            elif action == 'delete':
                # 删除记录
                record_id = request.form.get('id')
                record = db.session.get(model, record_id)
                if record:
                    db.session.delete(record)
                    db.session.commit()
                    flash(f'{dict_name}删除成功！', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 查询所有记录
    records = model.query.order_by(model.sort_order).all()

    # 渲染页面
    dict_html = f'''
    <h2>{dict_name}管理</h2>

    <!-- 添加记录表单 -->
    <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h3>添加{dict_name}</h3>
        <form method="POST">
            <input type="hidden" name="action" value="add">
            <div class="form-group">
                <label>{name_fields[0]} <span class="required">*</span></label>
                <input type="text" name="{name_fields[0]}" required>
            </div>
            <div class="form-group">
                <label>排序顺序</label>
                <input type="number" name="sort_order" value="0">
            </div>
            <button type="submit" class="btn">添加</button>
        </form>
    </div>

    <!-- 记录列表 -->
    <table style="width:100%; border-collapse:collapse;">
        <thead>
            <tr style="background:#f5f7fa;">
                <th style="padding:10px; border:1px solid #dee2e6;">{name_fields[0]}</th>
                <th style="padding:10px; border:1px solid #dee2e6;">排序</th>
                <th style="padding:10px; border:1px solid #dee2e6;">状态</th>
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    for record in records:
        status_text = '启用' if record.is_active else '禁用'
        status_style = 'color: #27ae60;' if record.is_active else 'color: #95a5a6;'

        dict_html += f'''
        <tr>
            <td style="padding:10px; border:1px solid #dee2e6;">{getattr(record, name_fields[0])}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{record.sort_order}</td>
            <td style="padding:10px; border:1px solid #dee2e6;"><span style="{status_style}">{status_text}</span></td>
            <td style="padding:10px; border:1px solid #dee2e6;">
                <button onclick="editRecord({record.id}, '{getattr(record, name_fields[0])}', {record.sort_order})" class="btn" style="padding:5px 10px; font-size:12px;">编辑</button>
                <form method="POST" style="display:inline;" onsubmit="return confirm('确定删除？')">
                    <input type="hidden" name="action" value="delete">
                    <input type="hidden" name="id" value="{record.id}">
                    <button type="submit" class="btn" style="padding:5px 10px; font-size:12px; background:#e74c3c;">删除</button>
                </form>
            </td>
        </tr>
        '''

    dict_html += '''
        </tbody>
    </table>

    <!-- 编辑弹窗 -->
    <div id="editModal" style="display:none; position:fixed; top:50%; left:50%; transform:translate(-50%, -50%); background:white; padding:30px; border-radius:8px; box-shadow:0 0 20px rgba(0,0,0,0.3); z-index:1000;">
        <h3>编辑''' + dict_name + '''</h3>
        <form method="POST">
            <input type="hidden" name="action" value="edit">
            <input type="hidden" name="id" id="edit_id">
            <div class="form-group">
                <label>''' + name_fields[0] + '''</label>
                <input type="text" name="''' + name_fields[0] + '''" id="edit_name" required>
            </div>
            <div class="form-group">
                <label>排序顺序</label>
                <input type="number" name="sort_order" id="edit_sort" value="0">
            </div>
            <button type="submit" class="btn">保存</button>
            <button type="button" onclick="closeEditModal()" class="btn" style="background:#95a5a6;">取消</button>
        </form>
    </div>

    <script>
        function editRecord(id, name, sort) {
            document.getElementById('edit_id').value = id;
            document.getElementById('edit_name').value = name;
            document.getElementById('edit_sort').value = sort;
            document.getElementById('editModal').style.display = 'block';
        }
        function closeEditModal() {
            document.getElementById('editModal').style.display = 'none';
        }
    </script>
    '''

    return render_base_layout(f'{dict_name}管理', dict_html, current_user)


def check_achievement_duplicate(title, type_name, user_id):
    """
    检查成果是否重复（根据成果名称和类型查询数据库）
    :param title: 成果名称
    :param type_name: 成果类型
    :param user_id: 用户 ID
    :return: {'exists': True/False, 'type_name': '实际类型名称', 'message': '提示信息'}
    """
    if not title:
        return {'exists': False, 'type_name': '', 'message': ''}

    # 成果类型与数据库表的映射关系
    type_model_mapping = {
        '期刊论文': JournalPaper,
        '会议论文': ConferencePaper,
        '教材': Textbook,
        '专著': Monograph,
        '发明专利': Patent,
        '实用新型专利': Patent,
        '外观设计专利': Patent,
        '软著': SoftwareCopyright,
        '教学成果获奖': TeachingAchievementAward,
        '教学竞赛获奖': TeachingCompetitionAward,
        '指导学生获奖': StudentGuidanceAward,
        '教研教改和课程建设项目': TeachingProject
    }

    # 如果类型不在映射中，返回不重复
    if type_name not in type_model_mapping:
        return {'exists': False, 'type_name': '', 'message': ''}

    model = type_model_mapping[type_name]

    try:
        existing = model.query.filter_by(title=title, user_id=user_id).first()

        if existing:
            # 找到重复成果，获取其所属用户信息
            owner = db.session.get(User, existing.user_id)
            owner_name = owner.username if owner else '未知用户'
            return {
                'exists': True,
                'type_name': type_name,
                'message': f'该成果已存在于数据库中（所有者：{owner_name}）'
            }
        else:
            return {'exists': False, 'type_name': '', 'message': ''}

    except Exception as e:
        logger.error(f"查重校验失败：{str(e)}")
        # 出错时默认不重复，允许导入
        return {'exists': False, 'type_name': '', 'message': ''}


@app.route('/achievement/batch_import', methods=['GET', 'POST'])
def batch_import():
    """批量导入成果页面"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    if request.method != 'POST':
        # 显示批量导入选择页面
        achievement_types = [
            ('journal_paper', '期刊论文'),
            ('conference_paper', '会议论文'),
            ('textbook', '教材'),
            ('monograph', '专著'),
            ('teaching_project', '教研教改和课程建设项目'),
            ('patent', '专利'),
            ('software_copyright', '软件著作'),
            ('teaching_achievement_award', '教学成果获奖'),
            ('teaching_competition_award', '教学竞赛获奖'),
            ('student_guidance_award', '指导学生获奖')
        ]

        content = f'''
        <h2>批量导入成果</h2>
        <div class="alert alert-info">
            <strong>使用说明：</strong><br>
            1. 选择要导入的成果类型<br>
            2. 下载对应的 Excel 模板<br>
            3. 按照模板格式填写数据<br>
            4. 上传填好的 Excel 文件<br>
            5. 系统会解析并显示数据，经您确认后导入数据库<br>
            6. 系统会自动检测重复成果，避免重复导入
        </div>

        <form method="GET" action="/achievement/batch_import/select_type" style="margin-top:20px;">
            <div class="form-group">
                <label>选择成果类型 <span style="color:red;">*</span></label>
                <select name="type" required style="width:300px;">
                    <option value="">请选择</option>
                    {''.join([f'<option value="{t[0]}">{t[1]}</option>' for t in achievement_types])}
                </select>
            </div>
            <button type="submit" class="btn">下一步：下载模板</button>
        </form>
        '''

        return render_base_layout('批量导入成果', content, current_user)

    # POST 请求处理（确认导入）
    return handle_batch_import_confirm(current_user)


@app.route('/achievement/batch_import/select_type')
def batch_import_select_type():
    """选择成果类型后跳转"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    achievement_type = request.args.get('type')
    if not achievement_type:
        flash('请选择成果类型！', 'danger')
        return redirect(url_for('batch_import'))

    # 重定向到带类型的批量导入页面
    return redirect(f'/achievement/batch_import/{achievement_type}')


@app.route('/achievement/batch_import/<achievement_type>', methods=['GET', 'POST'])
def batch_import_by_type(achievement_type):
    """按类型批量导入成果"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 验证成果类型
    type_mapping = {
        'journal_paper': ('JournalPaper', '期刊论文'),
        'conference_paper': ('ConferencePaper', '会议论文'),
        'textbook': ('Textbook', '教材'),
        'monograph': ('Monograph', '专著'),
        'teaching_project': ('TeachingProject', '教研教改和课程建设项目'),
        'patent': ('Patent', '专利'),
        'software_copyright': ('SoftwareCopyright', '软件著作'),
        'teaching_achievement_award': ('TeachingAchievementAward', '教学成果获奖'),
        'teaching_competition_award': ('TeachingCompetitionAward', '教学竞赛获奖'),
        'student_guidance_award': ('StudentGuidanceAward', '指导学生获奖')
    }

    if achievement_type not in type_mapping:
        flash('无效的成果类型！', 'danger')
        return redirect(url_for('batch_import'))

    model_name, type_name = type_mapping[achievement_type]

    # GET 请求：显示上传表单和模板下载
    if request.method != 'POST':
        content = f'''
        <h2>批量导入{type_name}</h2>
        <div class="alert alert-info">
            <strong>操作步骤：</strong><br>
            1. 点击下方按钮下载 Excel 模板<br>
            2. 按照模板格式填写数据<br>
            3. 上传填好的 Excel 文件<br>
            4. 系统解析后请仔细核对数据<br>
            5. 确认无误后点击导入数据库
        </div>

        <div style="margin:20px 0;">
            <a href="/achievement/batch_import/template/{achievement_type}" class="btn" style="background:#27ae60; margin-right:10px;">
                📥 下载{type_name}导入模板
            </a>
        </div>

        <form method="POST" enctype="multipart/form-data" style="margin-top:20px;">
            <div class="form-group">
                <label>上传 Excel 文件 <span style="color:red;">*</span></label>
                <input type="file" name="excel_file" accept=".xlsx,.xls" required>
                <p style="margin-top:5px; color:#666;">
                    支持格式：.xlsx 或 .xls 格式
                </p>
            </div>
            <button type="submit" class="btn">上传并解析</button>
            <a href="/achievement/batch_import" class="btn" style="background-color:#95a5a6; margin-left:10px;">返回上一步</a>
        </form>
        '''

        return render_base_layout(f'批量导入{type_name}', content, current_user)

    # POST 请求：处理 Excel 上传和解析
    return handle_excel_upload(achievement_type, model_name, type_name, current_user)


@app.route('/achievement/batch_import/template/<achievement_type>')
def download_batch_import_template(achievement_type):
    """下载批量导入 Excel 模板"""
    type_mapping = {
        'journal_paper': ('JournalPaper', '期刊论文'),
        'conference_paper': ('ConferencePaper', '会议论文'),
        'textbook': ('Textbook', '教材'),
        'monograph': ('Monograph', '专著'),
        'teaching_project': ('TeachingProject', '教研教改和课程建设项目'),
        'patent': ('Patent', '专利'),
        'software_copyright': ('SoftwareCopyright', '软件著作'),
        'teaching_achievement_award': ('TeachingAchievementAward', '教学成果获奖'),
        'teaching_competition_award': ('TeachingCompetitionAward', '教学竞赛获奖'),
        'student_guidance_award': ('StudentGuidanceAward', '指导学生获奖')
    }

    if achievement_type not in type_mapping:
        flash('无效的成果类型！', 'danger')
        return redirect(url_for('batch_import'))

    model_name, type_name = type_mapping[achievement_type]

    # 获取字段配置
    fields_config = get_batch_import_fields(achievement_type)

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = type_name

    # 写入表头
    headers = ['序号'] + [field['label'] for field in fields_config]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col)].width = 20

    # 写入示例数据（第二行）
    sample_data = get_sample_data(achievement_type)
    if sample_data:
        for row_idx, sample_row in enumerate(sample_data, 2):
            for col_idx, value in enumerate(sample_row, 2):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # 添加说明工作表
    ws_info = wb.create_sheet(title='填写说明')
    ws_info.cell(row=1, column=1, value='填写说明')
    ws_info.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    instructions = [
        '1. 请在下方工作表中填写数据',
        '2. 带*号的字段为必填项',
        '3. 日期格式：YYYY-MM-DD（如 2026-03-29）',
        '4. 多个作者/发明人等用逗号分隔',
        '5. 系统会自动检测重复成果',
        '6. 第一次导入建议先填少量数据测试',
        '',
        '各字段详细说明：'
    ]

    for idx, field in enumerate(fields_config, 9):
        required = '（必填）' if field.get('required', False) else ''
        instructions.append(f'{idx - 8}. {field["label"]}{required}')

    for row_idx, instruction in enumerate(instructions, 2):
        ws_info.cell(row=row_idx, column=1, value=instruction)

    ws_info.column_dimensions['A'].width = 50

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'{type_name}导入模板.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def get_batch_import_fields(achievement_type):
    """获取各成果类型的批量导入字段配置"""
    fields_map = {
        'journal_paper': [
            {'name': 'title', 'label': '论文名称', 'required': True},
            {'name': 'authors', 'label': '论文作者', 'required': True},
            {'name': 'corresponding_authors', 'label': '通讯作者'},
            {'name': 'journal_name', 'label': '期刊名称', 'required': True},
            {'name': 'inclusion_status', 'label': '收录情况'},
            {'name': 'year', 'label': '年'},
            {'name': 'volume', 'label': '卷'},
            {'name': 'issue', 'label': '期'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'publish_date', 'label': '发表日期'},
        ],
        'conference_paper': [
            {'name': 'title', 'label': '论文名称', 'required': True},
            {'name': 'authors', 'label': '论文作者', 'required': True},
            {'name': 'corresponding_authors', 'label': '通讯作者'},
            {'name': 'conference_name', 'label': '会议名称', 'required': True},
            {'name': 'conference_time', 'label': '会议时间'},
            {'name': 'conference_place', 'label': '会议地点'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
        ],
        'textbook': [
            {'name': 'title', 'label': '教材名称', 'required': True},
            {'name': 'textbook_series', 'label': '教材系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'cip_number', 'label': 'CIP 核字号'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'publication_month', 'label': '出版月份'},
            {'name': 'edition', 'label': '版次'},
            {'name': 'word_count', 'label': '字数'},
            {'name': 'price', 'label': '定价'},
            {'name': 'textbook_level', 'label': '教材级别'},
            {'name': 'textbook_type', 'label': '教材类型'},
            {'name': 'applicable_majors', 'label': '适用专业'},
            {'name': 'remarks', 'label': '备注'},
        ],
        'monograph': [
            {'name': 'title', 'label': '专著名称', 'required': True},
            {'name': 'textbook_series', 'label': '专著系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'cip_number', 'label': 'CIP 核字号'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'publication_month', 'label': '出版月份'},
            {'name': 'edition', 'label': '版次'},
            {'name': 'word_count', 'label': '字数'},
            {'name': 'price', 'label': '定价'},
            {'name': 'monograph_type', 'label': '专著类型'},
            {'name': 'applicable_majors', 'label': '适用专业'},
            {'name': 'remarks', 'label': '备注'},
        ],
        'teaching_project': [
            {'name': 'title', 'label': '项目名称', 'required': True},
            {'name': 'project_code', 'label': '项目编号'},
            {'name': 'project_leader', 'label': '项目负责人'},
            {'name': 'project_members', 'label': '项目参与人'},
            {'name': 'approval_department', 'label': '项目批准部门'},
            {'name': 'approval_date', 'label': '项目立项时间'},
            {'name': 'project_type', 'label': '项目类型'},
            {'name': 'project_level', 'label': '项目级别'},
            {'name': 'project_category', 'label': '项目类别'},
            {'name': 'project_status', 'label': '项目状态'},
            {'name': 'funding', 'label': '项目经费'},
            {'name': 'start_date', 'label': '项目开始时间'},
            {'name': 'end_date', 'label': '项目结束时间'},
        ],
        'patent': [
            {'name': 'title', 'label': '专利名称', 'required': True},
            {'name': 'patent_type', 'label': '专利类型'},
            {'name': 'patentee', 'label': '专利权人'},
            {'name': 'address', 'label': '地址'},
            {'name': 'inventors', 'label': '发明人'},
            {'name': 'patent_status', 'label': '专利状态'},
            {'name': 'patent_number', 'label': '专利号'},
            {'name': 'grant_announcement_number', 'label': '授权公告号'},
            {'name': 'apply_date', 'label': '专利申请日'},
            {'name': 'grant_announcement_date', 'label': '授权公告日'},
        ],
        'software_copyright': [
            {'name': 'title', 'label': '软件名称', 'required': True},
            {'name': 'copyright_owner', 'label': '著作权人'},
            {'name': 'completion_date', 'label': '开发完成日期'},
            {'name': 'first_publication_date', 'label': '首次发表日期'},
            {'name': 'right_acquisition_method', 'label': '权利取得方式'},
            {'name': 'right_scope', 'label': '权利范围'},
            {'name': 'copyright_number', 'label': '登记号'},
            {'name': 'certificate_number', 'label': '证书号'},
            {'name': 'register_date', 'label': '登记日期'},
        ],
        'teaching_achievement_award': [
            {'name': 'title', 'label': '成果名称', 'required': True},
            {'name': 'achievement_type', 'label': '教学成果奖类型'},
            {'name': 'achievement_level', 'label': '成果等级'},
            {'name': 'main_contributors', 'label': '主要完成人'},
            {'name': 'completing_units', 'label': '成果完成单位'},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'award_rank', 'label': '获奖等级'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'awarding_unit', 'label': '颁奖单位'},
            {'name': 'award_date', 'label': '获奖日期'},
        ],
        'teaching_competition_award': [
            {'name': 'title', 'label': '竞赛名称', 'required': True},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'competition_level', 'label': '竞赛等级'},
            {'name': 'award_rank', 'label': '获奖等级'},
            {'name': 'winners', 'label': '获奖人'},
            {'name': 'winner_unit', 'label': '获奖人所在单位'},
            {'name': 'competition_name', 'label': '竞赛主办方'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'certificate_number', 'label': '证书编号'},
        ],
        'student_guidance_award': [
            {'name': 'title', 'label': '获奖名称', 'required': True},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'competition_name', 'label': '竞赛名称'},
            {'name': 'competition_level', 'label': '竞赛等级'},
            {'name': 'award_rank', 'label': '获奖等级'},
            {'name': 'student_name', 'label': '获奖学生'},
            {'name': 'project_name', 'label': '获奖项目名称'},
            {'name': 'teacher_name', 'label': '指导教师'},
            {'name': 'student_unit', 'label': '获奖学生所在单位'},
            {'name': 'organizer', 'label': '竞赛主办方'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'award_date', 'label': '获奖日期'},
        ]
    }

    return fields_map.get(achievement_type, [])


def get_sample_data(achievement_type):
    """获取各成果类型的示例数据"""
    samples = {
        'journal_paper': [
            ['基于深度学习的图像识别研究', '张三，李四，王五', '张三', '计算机学报', 'SCI,EI', '2026', '45', '3',
             '100-110', '10.1234/journal.2026.001', '2026', '2026-03-01']
        ],
        'conference_paper': [
            ['人工智能在教育中的应用', '赵六，钱七', '赵六', '全国教育技术大会', '2026.03.15-2026.03-17', '北京',
             '50-60', '10.5678/conf.2026.002', '2026']
        ],
        'textbook': [
            ['Python 程序设计教程', '信息技术系列教材', '张三', '李四', '王五，赵六', '清华大学出版社',
             '978-7-302-123456', 'CIP2026001234', '2026', '3', '第 1 版', '300 千字', '59.00', '省级规划教材',
             '纸质教材', '计算机科学与技术', '适用于本科院校']
        ],
        'monograph': [
            ['深度学习理论与实践', '人工智能专著系列', '张三', '', '', '科学出版社', '978-7-03-123456', 'CIP2026005678',
             '2026', '1', '第 1 版', '450 千字', '89.00', '学术专著', '计算机科学与技术', '研究生及以上']
        ],
        'teaching_project': [
            ['新工科背景下的人工智能课程改革', 'JG2026001', '张三', '李四，王五', '教育部高教司', '2026-01-15',
             '教学改革', '省部级', '理工类', '在研', '100000.00', '2026-01-01', '2027-12-31']
        ],
        'patent': [
            ['一种基于机器学习的图像分类方法', '发明专利', 'XX 大学', '湖南省长沙市', '张三，李四', '已授权',
             'ZL202610000001.0', 'CN123456789A', '2026-01-10', '2026-03-20']
        ],
        'software_copyright': [
            ['智能教学管理系统 V1.0', '张三，李四', '2026-01-15', '2026-02-01', '原始取得', '全部权利', '2026SR001234',
             '软著登字第 1234567 号', '2026-03-01']
        ],
        'teaching_achievement_award': [
            ['地方高校一流本科专业建设创新与实践', '专业建设类', '省级', '张三，李四，王五', 'XX 大学教务处', '2026',
             '一等奖', '2026001', 'XX 省教育厅', '2026-03-15']
        ],
        'teaching_competition_award': [
            ['青年教师教学竞赛', '2026', '省部级', '一等奖', '张三', 'XX 大学教育学院', '全国高校教师教学竞赛',
             '2026-03-20', '2026002']
        ],
        'student_guidance_award': [
            ['大学生创新创业大赛', '2026', '互联网+', '国家级', '金奖', '李明', '智能医疗诊断系统', '张三',
             'XX 大学计算机学院', '教育部', '2026003', '2026-03-25']
        ]
    }

    return samples.get(achievement_type, None)


def handle_excel_upload(achievement_type, model_name, type_name, current_user):
    """处理 Excel 上传和解析"""
    if 'excel_file' not in request.files:
        flash('未选择文件！', 'danger')
        return redirect(request.referrer)

    file = request.files['excel_file']
    if file.filename == '':
        flash('未选择文件！', 'danger')
        return redirect(request.referrer)

    if not (file and allowed_file(file.filename)):
        flash('文件格式错误，请上传 Excel 文件！', 'danger')
        return redirect(request.referrer)

    try:
        # 读取 Excel 文件
        df = pd.read_excel(file)

        # 获取字段配置
        fields_config = get_batch_import_fields(achievement_type)
        field_labels = [field['label'] for field in fields_config]

        # 校验列名
        excel_columns = [col.strip() for col in df.columns.tolist()]
        if excel_columns[0] == '序号':
            excel_columns = excel_columns[1:]  # 跳过序号列

        # 检查必填列是否存在
        missing_cols = []
        for field in fields_config:
            if field.get('required', False) and field['label'] not in excel_columns:
                missing_cols.append(field['label'])

        if missing_cols:
            flash(f'Excel 文件缺少必填列：{", ".join(missing_cols)}', 'danger')
            return redirect(request.referrer)

        # 解析数据
        parsed_data = []
        for idx, row in df.iterrows():
            if idx == 0 or all(pd.isna(row[label]) for label in field_labels if label in excel_columns):
                continue  # 跳过表头行和空行

            item_data = {}
            for field in fields_config:
                label = field['label']
                if label in excel_columns:
                    value = row[label]
                    if pd.isna(value):
                        value = ''
                    else:
                        # 处理日期
                        if isinstance(value, datetime) or isinstance(value, date):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = str(value).strip()
                    item_data[field['name']] = value

            item_data['_row_num'] = idx + 1  # 记录行号用于提示
            parsed_data.append(item_data)

        if not parsed_data:
            flash('Excel 文件中没有有效数据！', 'danger')
            return redirect(request.referrer)

        # 检测重复成果
        duplicate_check_results = check_duplicate_achievements(achievement_type, parsed_data, current_user)

        # 生成确认页面
        return render_confirm_page(achievement_type, type_name, parsed_data, duplicate_check_results, fields_config,
                                   current_user)

    except Exception as e:
        flash(f'解析 Excel 文件失败：{str(e)}', 'danger')
        return redirect(request.referrer)


def check_duplicate_achievements(achievement_type, parsed_data, current_user):
    """检测重复成果"""
    model_map = {
        'journal_paper': JournalPaper,
        'conference_paper': ConferencePaper,
        'textbook': Textbook,
        'monograph': Monograph,
        'teaching_project': TeachingProject,
        'patent': Patent,
        'software_copyright': SoftwareCopyright,
        'teaching_achievement_award': TeachingAchievementAward,
        'teaching_competition_award': TeachingCompetitionAward,
        'student_guidance_award': StudentGuidanceAward
    }

    if achievement_type not in model_map:
        return {'duplicates': [], 'new_items': parsed_data}

    model = model_map[achievement_type]
    duplicates = []
    new_items = []

    for item_data in parsed_data:
        # 根据成果类型构建查询条件
        query_filters = {'user_id': current_user.id}

        # 核心字段查重（不同成果类型用不同字段）
        if achievement_type in ['journal_paper', 'conference_paper', 'textbook', 'monograph']:
            query_filters['title'] = item_data.get('title', '')
        elif achievement_type == 'teaching_project':
            query_filters['title'] = item_data.get('title', '')
        elif achievement_type == 'patent':
            query_filters['title'] = item_data.get('title', '')
        elif achievement_type == 'software_copyright':
            query_filters['title'] = item_data.get('title', '')
        elif achievement_type in ['teaching_achievement_award', 'teaching_competition_award', 'student_guidance_award']:
            query_filters['title'] = item_data.get('title', '')

        # 查询数据库
        existing = model.query.filter_by(**query_filters).first()

        if existing:
            item_data['_is_duplicate'] = True
            item_data['_duplicate_id'] = existing.id
            duplicates.append(item_data)
        else:
            item_data['_is_duplicate'] = False
            new_items.append(item_data)

    return {'duplicates': duplicates, 'new_items': new_items}


def render_confirm_page(achievement_type, type_name, parsed_data, duplicate_check_results, fields_config, current_user):
    """渲染确认页面"""
    duplicates = duplicate_check_results['duplicates']
    new_items = duplicate_check_results['new_items']

    # 生成表格 HTML
    table_html = '''
    <h2>数据解析结果 - 请确认</h2>
    <div class="alert alert-info">
        <strong>说明：</strong><br>
        1. 系统共解析出 <strong>{total}</strong> 条数据<br>
        2. 其中 <strong style="color:#27ae60;">{new_count}</strong> 条可导入（数据库中不存在）<br>
        3. 其中 <strong style="color:#e74c3c;">{duplicate_count}</strong> 条重复（数据库中已存在相同名称的成果）<br>
        4. 请勾选需要导入的数据（默认全选可导入的数据）<br>
        5. 重复的数据不能导入，仅供核对
    </div>
    '''.format(
        total=len(parsed_data),
        new_count=len(new_items),
        duplicate_count=len(duplicates)
    )

    if duplicates:
        table_html += '''
        <div class="alert alert-warning">
            <strong>以下成果重复，无法导入：</strong><br>
            <ul>
        '''
        for dup in duplicates:
            table_html += f'<li>第{dup["_row_num"]}行：{dup.get("title", "未知")}</li>'
        table_html += '''
            </ul>
        </div>
        '''

    table_html += '''
    <form method="POST" action="/achievement/batch_import/confirm">
        <input type="hidden" name="achievement_type" value="{achievement_type}">
        <input type="hidden" name="data_json" value="{data_json}">

        <table style="width:100%; border-collapse:collapse; margin:20px 0;">
            <thead>
                <tr style="background:#f5f7fa;">
                    <th style="padding:10px; border:1px solid #dee2e6;">选择</th>
                    <th style="padding:10px; border:1px solid #dee2e6;">序号</th>
        '''.format(
        achievement_type=achievement_type,
        data_json=json.dumps(parsed_data, ensure_ascii=False)
    )

    # 表头
    for field in fields_config[:8]:  # 只显示前 8 个字段避免太宽
        table_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{field["label"]}</th>'
    table_html += '<th style="padding:10px; border:1px solid #dee2e6;">状态</th></tr></thead><tbody>'

    # 表体
    for idx, item_data in enumerate(parsed_data, 1):
        is_duplicate = item_data.get('_is_duplicate', False)
        checkbox_disabled = 'disabled' if is_duplicate else ''
        checkbox_checked = '' if is_duplicate else 'checked'

        status_badge = '<span style="color:#27ae60;">✅ 可导入</span>' if not is_duplicate else '<span style="color:#e74c3c;">❌ 已存在</span>'

        table_html += f'''
        <tr>
            <td style="padding:10px; border:1px solid #dee2e6; text-align:center;">
                <input type="checkbox" name="selected_indices" value="{idx - 1}" {checkbox_checked} {checkbox_disabled}>
            </td>
            <td style="padding:10px; border:1px solid #dee2e6;">{item_data.get('_row_num', idx)}</td>
        '''

        # 显示前 8 个字段
        for field in fields_config[:8]:
            value = item_data.get(field['name'], '')
            if len(str(value)) > 30:
                value = str(value)[:28] + '...'
            table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{value}</td>'

        table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{status_badge}</td></tr>'

    table_html += '</tbody></table>'

    if new_items:
        table_html += '''
        <div style="margin:20px 0;">
            <button type="submit" class="btn" style="background:#27ae60;">✅ 确认导入选中的数据</button>
            <a href="/achievement/batch_import" class="btn" style="background-color:#95a5a6; margin-left:10px;">取消导入</a>
        </div>
        '''
    else:
        table_html += '''
        <div class="alert alert-danger">
            所有数据均已在数据库中存在，无需导入！
        </div>
        <a href="/achievement/batch_import" class="btn">返回</a>
        '''

    table_html += '</form>'

    return render_base_layout('批量导入确认', table_html, current_user)


@app.route('/achievement/batch_import/confirm', methods=['POST'])
def handle_batch_import_confirm(current_user):
    """处理批量导入确认"""
    achievement_type = request.form.get('achievement_type')
    data_json = request.form.get('data_json')
    selected_indices_str = request.form.getlist('selected_indices')

    if not achievement_type or not data_json:
        flash('参数错误！', 'danger')
        return redirect(url_for('batch_import'))

    try:
        parsed_data = json.loads(data_json)
    except:
        flash('数据解析失败！', 'danger')
        return redirect(url_for('batch_import'))

    # 获取选中的索引
    selected_indices = [int(idx) for idx in selected_indices_str]

    if not selected_indices:
        flash('未选择任何数据！', 'danger')
        return redirect(url_for('batch_import'))

    # 获取要导入的数据
    items_to_import = [parsed_data[i] for i in selected_indices if i < len(parsed_data)]

    # 执行导入
    success_count = 0
    error_count = 0
    errors = []

    for item_data in items_to_import:
        try:
            result = import_single_achievement(achievement_type, item_data, current_user)
            if result['success']:
                success_count += 1
            else:
                error_count += 1
                errors.append(f'第{item_data.get("_row_num", "?")}行：{result["error"]}')
        except Exception as e:
            error_count += 1
            errors.append(f'第{item_data.get("_row_num", "?")}行：{str(e)}')

    # 显示结果
    if success_count > 0:
        flash(f'成功导入 {success_count} 条数据！', 'success')

    if error_count > 0:
        flash(f'失败 {error_count} 条：{"；".join(errors[:3])}{"..." if error_count > 3 else ""}', 'danger')

    return redirect(url_for('batch_import'))


def import_single_achievement(achievement_type, item_data, current_user):
    """导入单条成果数据"""
    model_map = {
        'journal_paper': JournalPaper,
        'conference_paper': ConferencePaper,
        'textbook': Textbook,
        'monograph': Monograph,
        'teaching_project': TeachingProject,
        'patent': Patent,
        'software_copyright': SoftwareCopyright,
        'teaching_achievement_award': TeachingAchievementAward,
        'teaching_competition_award': TeachingCompetitionAward,
        'student_guidance_award': StudentGuidanceAward
    }

    if achievement_type not in model_map:
        return {'success': False, 'error': '无效的成果类型'}

    model = model_map[achievement_type]

    try:
        # 创建新对象
        item = model()
        item.user_id = current_user.id
        item.create_time = datetime.now()
        item.update_time = datetime.now()

        # 映射字段值
        field_mapping = get_field_mapping(achievement_type)
        for excel_field, db_field in field_mapping.items():
            if excel_field in item_data:
                value = item_data[excel_field]

                # 特殊处理
                if value == '' or value is None:
                    value = None

                # 日期字段处理
                if db_field in ['publish_date', 'conference_start_date', 'conference_end_date',
                                'approval_date', 'start_date', 'end_date', 'apply_date',
                                'grant_announcement_date', 'completion_date',
                                'first_publication_date', 'register_date', 'award_date']:
                    if value:
                        try:
                            value = datetime.strptime(str(value), '%Y-%m-%d').date()
                        except:
                            value = None
                    else:
                        value = None

                # 数字字段处理
                if db_field in ['publication_year', 'publication_month', 'publish_year',
                                'year', 'award_year']:
                    if value:
                        try:
                            value = int(value)
                        except:
                            value = None

                setattr(item, db_field, value)

        # 设置关联人员
        if hasattr(item, 'related_personnel_ids'):
            item.related_personnel_ids = str(current_user.id)

        db.session.add(item)
        db.session.commit()

        # 自动关联贡献者
        authors_field = None
        if achievement_type == 'journal_paper':
            authors_field = 'authors'
        elif achievement_type == 'conference_paper':
            authors_field = 'authors'
        elif achievement_type in ['textbook', 'monograph']:
            authors_field = 'chief_editor'

        if authors_field and authors_field in item_data:
            auto_link_contributors(item, achievement_type, item_data[authors_field], current_user.id)
            db.session.commit()

        return {'success': True}

    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}


def get_field_mapping(achievement_type):
    """获取 Excel 字段到数据库字段的映射"""
    mappings = {
        'journal_paper': {
            'title': 'title',
            'authors': 'authors',
            'corresponding_authors': 'corresponding_authors',
            'journal_name': 'journal_name',
            'inclusion_status': 'inclusion_status',
            'year': 'year',
            'volume': 'volume',
            'issue': 'issue',
            'page_range': 'page_range',
            'doi': 'doi',
            'publish_year': 'publish_year',
            'publish_date': 'publish_date',
        },
        'conference_paper': {
            'title': 'title',
            'authors': 'authors',
            'corresponding_authors': 'corresponding_authors',
            'conference_name': 'conference_name',
            'conference_time': 'conference_time',
            'conference_place': 'conference_place',
            'page_range': 'page_range',
            'doi': 'doi',
            'publish_year': 'publish_year',
        },
        'textbook': {
            'title': 'title',
            'textbook_series': 'textbook_series',
            'chief_editor': 'chief_editor',
            'associate_editors': 'associate_editors',
            'editorial_board': 'editorial_board',
            'publisher': 'publisher',
            'isbn': 'isbn',
            'cip_number': 'cip_number',
            'publication_year': 'publication_year',
            'publication_month': 'publication_month',
            'edition': 'edition',
            'word_count': 'word_count',
            'price': 'price',
            'textbook_level': 'textbook_level',
            'textbook_type': 'textbook_type',
            'applicable_majors': 'applicable_majors',
            'remarks': 'remarks',
        },
        'monograph': {
            'title': 'title',
            'textbook_series': 'textbook_series',
            'chief_editor': 'chief_editor',
            'associate_editors': 'associate_editors',
            'editorial_board': 'editorial_board',
            'publisher': 'publisher',
            'isbn': 'isbn',
            'cip_number': 'cip_number',
            'publication_year': 'publication_year',
            'publication_month': 'publication_month',
            'edition': 'edition',
            'word_count': 'word_count',
            'price': 'price',
            'monograph_type': 'monograph_type',
            'applicable_majors': 'applicable_majors',
            'remarks': 'remarks',
        },
        'teaching_project': {
            'title': 'title',
            'project_code': 'project_code',
            'project_leader': 'project_leader',
            'project_members': 'project_members',
            'approval_department': 'approval_department',
            'approval_date': 'approval_date',
            'project_type': 'project_type',
            'project_level': 'project_level',
            'project_category': 'project_category',
            'project_status': 'project_status',
            'funding': 'funding',
            'start_date': 'start_date',
            'end_date': 'end_date',
        },
        'patent': {
            'title': 'title',
            'patent_type': 'patent_type',
            'patentee': 'patentee',
            'address': 'address',
            'inventors': 'inventors',
            'patent_status': 'patent_status',
            'patent_number': 'patent_number',
            'grant_announcement_number': 'grant_announcement_number',
            'apply_date': 'apply_date',
            'grant_announcement_date': 'grant_announcement_date',
        },
        'software_copyright': {
            'title': 'title',
            'copyright_owner': 'copyright_owner',
            'completion_date': 'completion_date',
            'first_publication_date': 'first_publication_date',
            'right_acquisition_method': 'right_acquisition_method',
            'right_scope': 'right_scope',
            'copyright_number': 'copyright_number',
            'certificate_number': 'certificate_number',
            'register_date': 'register_date',
        },
        'teaching_achievement_award': {
            'title': 'title',
            'achievement_type': 'achievement_type',
            'achievement_level': 'achievement_level',
            'main_contributors': 'main_contributors',
            'completing_units': 'completing_units',
            'award_year': 'award_year',
            'award_rank': 'award_rank',
            'certificate_number': 'certificate_number',
            'awarding_unit': 'awarding_unit',
            'award_date': 'award_date',
        },
        'teaching_competition_award': {
            'title': 'title',
            'award_year': 'award_year',
            'competition_level': 'competition_level',
            'award_rank': 'award_rank',
            'winners': 'winners',
            'winner_unit': 'winner_unit',
            'competition_name': 'competition_name',
            'award_date': 'award_date',
            'certificate_number': 'certificate_number',
        },
        'student_guidance_award': {
            'title': 'title',
            'award_year': 'award_year',
            'competition_name': 'competition_name',
            'competition_level': 'competition_level',
            'award_rank': 'award_rank',
            'student_name': 'student_name',
            'project_name': 'project_name',
            'teacher_name': 'teacher_name',
            'student_unit': 'student_unit',
            'organizer': 'organizer',
            'certificate_number': 'certificate_number',
            'award_date': 'award_date',
        }
    }

    return mappings.get(achievement_type, {})


# ---------------------- 5. 初始化数据库（强制重建+防重复创建） ----------------------
def init_database():
    """初始化数据库（强制删除旧文件 + 创建新表 + 默认管理员）"""
    with app.app_context():
        if os.path.exists(DB_FILE):
            try:
                os.remove(DB_FILE)
            except Exception as e:
                raise Exception(f'无法删除旧数据库文件，请手动删除 {DB_FILE} 后重试')

        db.create_all()

        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                employee_id='000000',
                email='admin@example.com',
                role='admin'
            )
            admin.set_password('admin123')
            db.session.add(admin)

        inclusion_data = [
            {'type_name': 'SCI 期刊', 'type_code': 'SCI', 'description': '科学引文索引期刊', 'sort_order': 1},
            {'type_name': 'SSCI 期刊', 'type_code': 'SSCI', 'description': '社会科学引文索引期刊', 'sort_order': 2},
            {'type_name': 'EI 期刊', 'type_code': 'EI', 'description': '工程索引期刊', 'sort_order': 3},
            {'type_name': 'CSSCI 期刊', 'type_code': 'CSSCI', 'description': '中文社会科学引文索引期刊',
             'sort_order': 4},
            {'type_name': 'CSCD 核心库期刊', 'type_code': 'CSCD_CORE',
             'description': '中国科学引文数据库核心库期刊', 'sort_order': 5},
            {'type_name': 'CSCD 扩展库期刊', 'type_code': 'CSCD_EXT', 'description': '中国科学引文数据库扩展库期刊',
             'sort_order': 6},
            {'type_name': '北大中文核心期刊', 'type_code': 'PKU_CORE', 'description': '北京大学中文核心期刊',
             'sort_order': 7},
            {'type_name': '中国科技核心期刊', 'type_code': 'CSTPCD', 'description': '中国科技论文统计源期刊',
             'sort_order': 8},
            {'type_name': '普通期刊', 'type_code': 'GENERAL', 'description': '普通期刊', 'sort_order': 9},
            {'type_name': '其它', 'type_code': 'OTHER', 'description': '其他收录类型', 'sort_order': 10},
        ]

        for data in inclusion_data:
            existing = InclusionType.query.filter_by(type_code=data['type_code']).first()
            if not existing:
                inclusion_type = InclusionType(**data)
                db.session.add(inclusion_type)

        textbook_level_data = [
            {'level_name': '国家级规划教材', 'level_code': 'NATIONAL', 'description': '国家级规划教材',
             'sort_order': 1},
            {'level_name': '全国行业规划教材', 'level_code': 'INDUSTRY', 'description': '全国行业规划教材',
             'sort_order': 2},
            {'level_name': '协编教材', 'level_code': 'COEDIT', 'description': '协编教材', 'sort_order': 3},
            {'level_name': '自编教材', 'level_code': 'SELF', 'description': '自编教材', 'sort_order': 4},
            {'level_name': '其它', 'level_code': 'OTHER', 'description': '其它', 'sort_order': 5},
        ]

        for data in textbook_level_data:
            existing = TextbookLevel.query.filter_by(level_code=data['level_code']).first()
            if not existing:
                textbook_level = TextbookLevel(**data)
                db.session.add(textbook_level)

        db.session.commit()


# ---------------------- 6. 启动应用 ----------------------
if __name__ == '__main__':
    with app.app_context():
        db.create_all()

        # 创建默认管理员账户
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                employee_id='admin',
                email='admin@hntcm.edu.cn',
                role='admin'
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()

        # 核心修复：初始化收录类型字典数据
        if InclusionType.query.count() == 0:
            inclusion_data = [
                {'type_name': 'SCI 期刊', 'type_code': 'SCI', 'description': '科学引文索引期刊', 'sort_order': 1},
                {'type_name': 'SSCI 期刊', 'type_code': 'SSCI', 'description': '社会科学引文索引期刊', 'sort_order': 2},
                {'type_name': 'EI 期刊', 'type_code': 'EI', 'description': '工程索引期刊', 'sort_order': 3},
                {'type_name': 'CSSCI 期刊', 'type_code': 'CSSCI', 'description': '中文社会科学引文索引期刊',
                 'sort_order': 4},
                {'type_name': 'CSCD 核心库期刊', 'type_code': 'CSCD_CORE',
                 'description': '中国科学引文数据库核心库期刊', 'sort_order': 5},
                {'type_name': 'CSCD 扩展库期刊', 'type_code': 'CSCD_EXT', 'description': '中国科学引文数据库扩展库期刊',
                 'sort_order': 6},
                {'type_name': '北大中文核心期刊', 'type_code': 'PKU_CORE', 'description': '北京大学中文核心期刊',
                 'sort_order': 7},
                {'type_name': '中国科技核心期刊', 'type_code': 'CSTPCD', 'description': '中国科技论文统计源期刊',
                 'sort_order': 8},
                {'type_name': '普通期刊', 'type_code': 'GENERAL', 'description': '普通期刊', 'sort_order': 9},
                {'type_name': '其它', 'type_code': 'OTHER', 'description': '其他收录类型', 'sort_order': 10},
            ]

            for data in inclusion_data:
                existing = InclusionType.query.filter_by(type_code=data['type_code']).first()
                if not existing:
                    inclusion_type = InclusionType(**data)
                    db.session.add(inclusion_type)

            db.session.commit()
            print("[OK] 初始化收录类型数据完成")

            # 初始化教材级别字典数据
        if TextbookLevel.query.count() == 0:
            textbook_level_data = [
                {'level_name': '国家级规划教材', 'level_code': 'NATIONAL', 'description': '国家级规划教材',
                 'sort_order': 1},
                {'level_name': '全国行业规划教材', 'level_code': 'INDUSTRY', 'description': '全国行业规划教材',
                 'sort_order': 2},
                {'level_name': '协编教材', 'level_code': 'COEDIT', 'description': '协编教材', 'sort_order': 3},
                {'level_name': '自编教材', 'level_code': 'SELF', 'description': '自编教材', 'sort_order': 4},
                {'level_name': '其它', 'level_code': 'OTHER', 'description': '其它', 'sort_order': 5},
            ]

            for data in textbook_level_data:
                existing = TextbookLevel.query.filter_by(level_code=data['level_code']).first()
                if not existing:
                    textbook_level = TextbookLevel(**data)
                    db.session.add(textbook_level)

            db.session.commit()
            print("初始化教材级别数据完成")

        # 初始化教材级别字典数据
        if TextbookLevel.query.count() == 0:
            textbook_level_data = [
                {'level_name': '国家级规划教材', 'level_code': 'NATIONAL', 'description': '国家级规划教材',
                 'sort_order': 1},
                {'level_name': '全国行业规划教材', 'level_code': 'INDUSTRY', 'description': '全国行业规划教材',
                 'sort_order': 2},
                {'level_name': '协编教材', 'level_code': 'COEDIT', 'description': '协编教材', 'sort_order': 3},
                {'level_name': '自编教材', 'level_code': 'SELF', 'description': '自编教材', 'sort_order': 4},
                {'level_name': '其它', 'level_code': 'OTHER', 'description': '其它', 'sort_order': 5},
            ]

            for data in textbook_level_data:
                existing = TextbookLevel.query.filter_by(level_code=data['level_code']).first()
                if not existing:
                    textbook_level = TextbookLevel(**data)
                    db.session.add(textbook_level)

            db.session.commit()
            print("✅ 初始化教材级别数据完成")

        # 初始化教学成果奖类型
        if TeachingAchievementType.query.count() == 0:
            default_types = [
                TeachingAchievementType(type_name='湖南中医药大学教学成果奖', sort_order=1),
                TeachingAchievementType(type_name='湖南中医药大学研究生教学成果奖', sort_order=2),
                TeachingAchievementType(type_name='湖南省计算机学会高等教育教学成果奖', sort_order=3),
                TeachingAchievementType(type_name='其它', sort_order=4)
            ]
            db.session.add_all(default_types)
            db.session.commit()
            print("初始化教学成果奖类型数据")

        # 初始化成果等级
        if AchievementLevel.query.count() == 0:
            default_levels = [
                AchievementLevel(level_name='国家级', sort_order=1),
                AchievementLevel(level_name='省部级', sort_order=2),
                AchievementLevel(level_name='市厅级', sort_order=3),
                AchievementLevel(level_name='校级', sort_order=4),
                AchievementLevel(level_name='院级', sort_order=5),
                AchievementLevel(level_name='其它', sort_order=6)
            ]
            db.session.add_all(default_levels)
            db.session.commit()
            print("初始化成果等级数据")

        # 初始化获奖等级
        if AwardRank.query.count() == 0:
            default_ranks = [
                AwardRank(rank_name='特等奖', sort_order=1),
                AwardRank(rank_name='一等奖', sort_order=2),
                AwardRank(rank_name='二等奖', sort_order=3),
                AwardRank(rank_name='三等奖', sort_order=4),
                AwardRank(rank_name='优秀奖', sort_order=5),
                AwardRank(rank_name='其它', sort_order=6)
            ]
            db.session.add_all(default_ranks)
            db.session.commit()
            print("初始化获奖等级数据")

        print("数据库初始化完成")

    app.run(debug=True, host='0.0.0.0', port=5000)

